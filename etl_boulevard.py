#!/usr/bin/env python3
"""
ETL: Boulevard Project → Supabase
==================================
Extracts data from "BOULEVARD 5" (actuals) and "BOULEVARD PPTO" (budget)
sheets in the Reservas workbook and loads into Supabase production database.

Usage:
    python etl_boulevard.py path/to/Reservas.xlsx                  # Dry-run (default)
    python etl_boulevard.py path/to/Reservas.xlsx --execute         # Live insert
    python etl_boulevard.py path/to/Reservas.xlsx --execute --verbose

Environment:
    SUPABASE_URL  — Supabase project URL
    SUPABASE_KEY  — Supabase service_role key (NOT anon key)
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

PROJECT_NAME = "boulevard"
PROJECT_DISPLAY_NAME = "Boulevard 5"
SHEET_ACTUALS = "BOULEVARD 5"
SHEET_BUDGET = "BOULEVARD PPTO"

# Row boundaries (used only as sanity-check limits, NOT for column lookup)
B5_HEADER_ROW = 6
B5_DATA_START = 7
B5_DATA_END = 304
B5_DESIST_HEADER = 319
B5_DESIST_START = 320
B5_DESIST_END = 351
B5_PROB_DESIST_HEADER = 357
B5_PROB_DESIST_START = 358
B5_PROB_DESIST_END = 366

PPTO_HEADER_ROW = 1
PPTO_DATA_START = 2
PPTO_DATA_END = 299

BATCH_SIZE = 500

# ─────────────────────────────────────────────────────────────────────────────
# Status normalization maps
# ─────────────────────────────────────────────────────────────────────────────

STATUS_TO_UNIT_STATUS: dict[str, str] = {
    "1.disponible": "available",
    "2.reserva": "reserved",
    "2. reserva": "reserved",
    "2,reserva": "reserved",
    "4.plan de pagos": "sold",
    "desistimiento": "cancelled",
}

STATUS_TO_SALE_STATUS: dict[str, str] = {
    "2.reserva": "active",
    "2. reserva": "active",
    "2,reserva": "active",
    "4.plan de pagos": "active",
    "desistimiento": "cancelled",
}

VENDEDOR_CANONICAL: dict[str, str] = {
    "ronaldo": "Ronaldo Ogaldez",
    "ronaldo ogaldez": "Ronaldo Ogaldez",
    "anahi cisneros": "Anahí Cisneros",
    "anahí cisneros": "Anahí Cisneros",
    "**": None,
    "sin datos": None,
}

log = logging.getLogger("etl_boulevard")


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class ParsedUnit:
    """Raw unit data extracted from BOULEVARD 5."""

    apto: str
    tipo: str | None = None
    vendedor: str | None = None
    cliente: str | None = None
    fecha_reserva: date | None = None
    estatus_raw: str | None = None
    precio_venta: float | None = None
    enganche: float | None = None
    saldo_financiar: float | None = None
    cuotas_pactadas: int | None = None
    monto_reserva_pactado: float | None = None
    monto_cuota_pactada: float | None = None
    cuotas_pagadas: int | None = None
    caso_especial: str | None = None
    observaciones: str | None = None
    notas: str | None = None
    iva: float | None = None
    timbres: float | None = None
    actual_payments: list[tuple[date, float]] = field(default_factory=list)
    source_section: str = "main"  # main | desistimiento | prob_desistimiento


@dataclass
class ParsedExpected:
    """Raw expected payment from BOULEVARD PPTO."""

    apto: str
    due_date: date
    amount: float


@dataclass
class ValidationReport:
    """Accumulates warnings and errors during validation."""

    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0

    def print_summary(self) -> None:
        if self.errors:
            log.error("═══ VALIDATION ERRORS (%d) ═══", len(self.errors))
            for e in self.errors:
                log.error("  ✗ %s", e)
        if self.warnings:
            log.warning("═══ VALIDATION WARNINGS (%d) ═══", len(self.warnings))
            for w in self.warnings:
                log.warning("  ⚠ %s", w)
        if not self.errors and not self.warnings:
            log.info("═══ VALIDATION: ALL CLEAN ═══")


# ─────────────────────────────────────────────────────────────────────────────
# Header discovery — finds columns by name, not by index
# ─────────────────────────────────────────────────────────────────────────────


def _normalize_header(val: Any) -> str | None:
    """Normalize a header cell value to a comparable string key."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m")
    s = str(val).strip().lower()
    s = re.sub(r"\s+", " ", s)  # collapse whitespace/newlines
    return s


# Patterns to match header names (lowercase, whitespace-collapsed)
# Each tuple: (logical_name, list_of_patterns)
B5_HEADER_PATTERNS: list[tuple[str, list[str]]] = [
    ("apto", ["apto", "apto."]),
    ("tipo", ["tipo", "tipo.", "modelo"]),
    ("notas", ["notas"]),
    ("vendedor", ["vendedor"]),
    ("cliente", ["cliente"]),
    ("fecha_reserva", ["fecha reserva", "fecha"]),
    ("estatus", ["estatus", "tipo de plan"]),
    ("precio_venta", ["precio de venta", "p. venta", "p.venta"]),
    ("enganche", ["enganche"]),
    ("total_enganches", ["total enganches y reservas"]),
    ("saldo_financiar", ["saldo a financiar por el banco"]),
    ("cuotas_pactadas", ["cuotas pactadas"]),
    ("monto_reserva_pactado", ["monto de reserva pactado"]),
    ("monto_cuota_pactada", ["monto de cuota pactada"]),
    ("cuotas_pagadas", ["cuotas pagadas"]),
    ("caso_especial", ["caso especial / f&f", "caso especial"]),
    ("observaciones", ["observaciones"]),
    ("iva", ["iva"]),
    ("timbres", ["timbres"]),
]

PPTO_HEADER_PATTERNS: list[tuple[str, list[str]]] = [
    ("apto", ["apto", "apto."]),
    ("estatus", ["estatus"]),
]


def discover_headers(
    ws: Worksheet,
    header_row: int,
    col_start: int,
    col_end: int,
    patterns: list[tuple[str, list[str]]],
) -> tuple[dict[str, int], dict[str, int]]:
    """
    Scan a header row and return two dicts:
      - named_cols: {logical_name: col_index} for pattern-matched headers
      - month_cols: {YYYY-MM: col_index} for datetime (monthly) headers
    """
    named_cols: dict[str, int] = {}
    month_cols: dict[str, int] = {}

    for col in range(col_start, col_end + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw is None:
            continue

        # Check if it's a date → monthly column
        if isinstance(raw, datetime):
            key = raw.strftime("%Y-%m")
            month_cols[key] = col
            continue

        norm = _normalize_header(raw)
        if norm is None:
            continue

        # Match against known patterns.
        # Two-pass: exact matches first, then substring matches.
        # This prevents short patterns (e.g. "tipo") from greedily
        # claiming longer headers (e.g. "tipo de plan").
        matched = False
        for logical_name, pats in patterns:
            if logical_name in named_cols:
                continue
            if any(p == norm for p in pats):
                named_cols[logical_name] = col
                matched = True
                break
        if not matched:
            for logical_name, pats in patterns:
                if logical_name in named_cols:
                    continue
                if any(len(p) > 4 and p in norm for p in pats):
                    named_cols[logical_name] = col
                    break

    return named_cols, month_cols


# ─────────────────────────────────────────────────────────────────────────────
# Parsers
# ─────────────────────────────────────────────────────────────────────────────


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return round(f, 2)
    except (ValueError, TypeError):
        return None


def _safe_int(val: Any) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _safe_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _safe_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _month_end_to_date(year_month: str) -> date:
    """Convert 'YYYY-MM' to the last day of that month (used as due_date)."""
    y, m = int(year_month[:4]), int(year_month[5:])
    if m == 12:
        return date(y + 1, 1, 1).replace(day=1) - __import__("datetime").timedelta(days=1)
    return date(y, m + 1, 1) - __import__("datetime").timedelta(days=1)


def parse_b5_section(
    ws: Worksheet,
    header_row: int,
    data_start: int,
    data_end: int,
    col_start: int,
    col_end: int,
    section_name: str,
    month_cols_override: dict[str, int] | None = None,
) -> list[ParsedUnit]:
    """Parse a section of BOULEVARD 5 (main, desistimientos, or problematic).

    For sub-sections (desistimientos), monthly columns are inherited from the
    main section via month_cols_override since they share the same column
    positions but only the main header row (row 6) has datetime headers.
    """
    named, months = discover_headers(ws, header_row, col_start, col_end, B5_HEADER_PATTERNS)
    if month_cols_override is not None:
        months = month_cols_override

    log.info(
        "  [%s] Headers found: %s | Monthly columns: %d (%s → %s)",
        section_name,
        list(named.keys()),
        len(months),
        min(months.keys(), default="?"),
        max(months.keys(), default="?"),
    )

    missing = {"apto"} - named.keys()
    if missing:
        log.error("  CRITICAL: Missing required headers: %s", missing)
        return []

    units: list[ParsedUnit] = []

    for row in range(data_start, data_end + 1):
        apto_raw = ws.cell(row=row, column=named["apto"]).value
        if apto_raw is None:
            continue

        apto = str(int(apto_raw)) if isinstance(apto_raw, (int, float)) else str(apto_raw).strip()
        if not apto:
            continue

        # Extract named fields (gracefully skip missing columns)
        def _get(name: str) -> Any:
            col = named.get(name)
            return ws.cell(row=row, column=col).value if col else None

        # Extract monthly payments
        payments: list[tuple[date, float]] = []
        for ym, col in sorted(months.items()):
            val = _safe_float(ws.cell(row=row, column=col).value)
            if val is not None and val != 0:
                payments.append((_month_end_to_date(ym), val))

        unit = ParsedUnit(
            apto=apto,
            tipo=_safe_str(_get("tipo")),
            vendedor=_safe_str(_get("vendedor")),
            cliente=_safe_str(_get("cliente")),
            fecha_reserva=_safe_date(_get("fecha_reserva")),
            estatus_raw=_safe_str(_get("estatus")),
            precio_venta=_safe_float(_get("precio_venta")),
            enganche=_safe_float(_get("enganche")),
            saldo_financiar=_safe_float(_get("saldo_financiar")),
            cuotas_pactadas=_safe_int(_get("cuotas_pactadas")),
            monto_reserva_pactado=_safe_float(_get("monto_reserva_pactado")),
            monto_cuota_pactada=_safe_float(_get("monto_cuota_pactada")),
            cuotas_pagadas=_safe_int(_get("cuotas_pagadas")),
            caso_especial=_safe_str(_get("caso_especial")),
            observaciones=_safe_str(_get("observaciones")),
            notas=_safe_str(_get("notas")),
            iva=_safe_float(_get("iva")),
            timbres=_safe_float(_get("timbres")),
            actual_payments=payments,
            source_section=section_name,
        )
        units.append(unit)

    log.info("  [%s] Parsed %d units, %d with payments", section_name, len(units), sum(1 for u in units if u.actual_payments))
    return units


def parse_ppto(ws: Worksheet) -> list[ParsedExpected]:
    """Parse BOULEVARD PPTO for expected installment schedule."""
    named, months = discover_headers(ws, PPTO_HEADER_ROW, 1, 62, PPTO_HEADER_PATTERNS)

    log.info(
        "  [PPTO] Headers found: %s | Monthly columns: %d (%s → %s)",
        list(named.keys()),
        len(months),
        min(months.keys(), default="?"),
        max(months.keys(), default="?"),
    )

    if "apto" not in named:
        log.error("  CRITICAL: 'Apto' header not found in PPTO")
        return []

    apto_col = named["apto"]
    records: list[ParsedExpected] = []

    for row in range(PPTO_DATA_START, PPTO_DATA_END + 1):
        apto_raw = ws.cell(row=row, column=apto_col).value
        if apto_raw is None:
            continue

        apto = str(int(apto_raw)) if isinstance(apto_raw, (int, float)) else str(apto_raw).strip()

        for ym, col in sorted(months.items()):
            val = _safe_float(ws.cell(row=row, column=col).value)
            if val is not None and val != 0:
                records.append(ParsedExpected(apto=apto, due_date=_month_end_to_date(ym), amount=val))

    log.info("  [PPTO] Parsed %d expected payment records", len(records))
    return records


# ─────────────────────────────────────────────────────────────────────────────
# Normalization
# ─────────────────────────────────────────────────────────────────────────────


def normalize_unit_status(raw: str | None) -> str:
    if raw is None:
        return "available"
    key = raw.strip().lower()
    return STATUS_TO_UNIT_STATUS.get(key, "available")


def normalize_sale_status(raw: str | None) -> str | None:
    """Returns None if no sale should exist (e.g., Disponible)."""
    if raw is None:
        return None
    key = raw.strip().lower()
    return STATUS_TO_SALE_STATUS.get(key)


def normalize_vendedor(raw: str | None) -> str | None:
    if raw is None:
        return None
    key = raw.strip().lower()
    canonical = VENDEDOR_CANONICAL.get(key)
    if canonical is not None:
        return canonical
    if key in VENDEDOR_CANONICAL and VENDEDOR_CANONICAL[key] is None:
        return None
    return raw.strip()


# ─────────────────────────────────────────────────────────────────────────────
# Validation
# ─────────────────────────────────────────────────────────────────────────────


def validate(
    main_units: list[ParsedUnit],
    desist_units: list[ParsedUnit],
    prob_desist_units: list[ParsedUnit],
    expected: list[ParsedExpected],
) -> ValidationReport:
    report = ValidationReport()

    # 1. Check for duplicate aptos within main
    main_aptos = [u.apto for u in main_units]
    dupes = [a for a in set(main_aptos) if main_aptos.count(a) > 1]
    if dupes:
        report.errors.append(f"Duplicate Apto in main data: {dupes}")

    # 2. Desistimiento + main overlap = expected (re-sold units).
    #    Flag as INFO, not warning.
    desist_aptos = {u.apto for u in desist_units + prob_desist_units}
    active_main = {u.apto for u in main_units if normalize_unit_status(u.estatus_raw) != "available"}
    resold = desist_aptos & active_main
    if resold:
        log.info("  ℹ %d units have lifecycle history (desisted then re-sold): %s",
                 len(resold), sorted(resold)[:10])

    # 3. Desistimientos pointing to units NOT in main data
    desist_only = desist_aptos - set(main_aptos)
    if desist_only:
        report.warnings.append(
            f"Desistimiento aptos not in main unit list ({len(desist_only)}): "
            f"{sorted(desist_only)[:10]}... — will create unit from desist data"
        )

    # 4. Cross-check PPTO aptos vs B5 aptos
    ppto_aptos = {r.apto for r in expected}
    all_b5_aptos = set(main_aptos) | desist_aptos
    ppto_only = ppto_aptos - all_b5_aptos
    b5_only = all_b5_aptos - ppto_aptos
    if ppto_only:
        report.warnings.append(f"Aptos in PPTO but not in B5 ({len(ppto_only)}): {sorted(ppto_only)[:10]}...")
    if b5_only:
        report.warnings.append(f"Aptos in B5 but not in PPTO ({len(b5_only)}): {sorted(b5_only)[:10]}...")

    # 5. Sales without client
    sold_no_client = [u.apto for u in main_units if normalize_sale_status(u.estatus_raw) and not u.cliente]
    if sold_no_client:
        report.warnings.append(f"Sold/reserved units missing client ({len(sold_no_client)}): {sold_no_client[:10]}")

    # 6. Units with unknown status
    for u in main_units:
        if u.estatus_raw and u.estatus_raw.strip().lower() not in STATUS_TO_UNIT_STATUS:
            report.errors.append(f"Apto {u.apto}: Unknown status '{u.estatus_raw}'")

    return report


# ─────────────────────────────────────────────────────────────────────────────
# DB insertion via Supabase client
# ─────────────────────────────────────────────────────────────────────────────


def _batched(items: list, size: int):
    """Yield successive batches from a list."""
    for i in range(0, len(items), size):
        yield items[i : i + size]


class SupabaseLoader:
    """Handles all DB writes in dependency order with idempotent upserts.

    Lifecycle model:
        unit (1) ←── (N) sale (N) ──→ (1) client
                           │
                           └── (N) payment

    A unit can cycle through sold → cancelled → re-sold unlimited times.
    Each cycle is a separate `sale` record. Payments belong to the sale,
    not the unit — preserving full history per client ownership period.

    Constraints:
        - units: UNIQUE(project_id, unit_number)
        - sales: UNIQUE(unit_id) WHERE status = 'active' (partial)
        - clients: UNIQUE(full_name)
    """

    def __init__(self, url: str, key: str):
        from supabase import create_client

        self.client = create_client(url, key)
        self.project_id: str | None = None
        self.unit_ids: dict[str, str] = {}    # apto → uuid
        self.client_ids: dict[str, str] = {}  # client_name → uuid
        # sale_ids keyed by Python object id — each ParsedUnit maps to its
        # own sale, even if multiple ParsedUnits share the same apto.
        self._sale_ids: dict[int, str] = {}   # id(ParsedUnit) → sale uuid

    def get_sale_id(self, unit: ParsedUnit) -> str | None:
        return self._sale_ids.get(id(unit))

    # ── Phase 1: Project ──────────────────────────────────────────────────

    def upsert_project(self) -> str:
        log.info("  Upserting project '%s'...", PROJECT_NAME)
        result = (
            self.client.table("projects")
            .upsert(
                {"name": PROJECT_NAME, "display_name": PROJECT_DISPLAY_NAME},
                on_conflict="name",
            )
            .execute()
        )
        self.project_id = result.data[0]["id"]
        log.info("  Project ID: %s", self.project_id)
        return self.project_id

    # ── Phase 2: Units ────────────────────────────────────────────────────

    def upsert_units(self, main_units: list[ParsedUnit]) -> None:
        """Upsert units from MAIN section ONLY.

        The unit row represents the CURRENT state of the physical asset.
        Desistimientos don't change the unit — they only create cancelled
        sale records. The unit's price/status reflects whoever owns it now.
        """
        log.info("  Upserting %d units (main section only)...", len(main_units))

        rows = []
        for u in main_units:
            precio = u.precio_venta or 0
            tax = round((u.iva or 0) + (u.timbres or 0), 2)
            rows.append(
                {
                    "project_id": self.project_id,
                    "unit_number": u.apto,
                    "price_with_tax": precio,
                    "price_without_tax": round(precio - tax, 2),
                    "down_payment_amount": u.enganche or 0,
                    "status": normalize_unit_status(u.estatus_raw),
                }
            )

        for batch in _batched(rows, BATCH_SIZE):
            result = (
                self.client.table("units")
                .upsert(batch, on_conflict="project_id,unit_number")
                .execute()
            )
            for row in result.data:
                self.unit_ids[row["unit_number"]] = row["id"]

        log.info("  Upserted %d units, got %d IDs", len(rows), len(self.unit_ids))

    # ── Phase 3: Clients ──────────────────────────────────────────────────

    def upsert_clients(self, all_units: list[ParsedUnit]) -> None:
        """Upsert ALL unique clients from every section — including
        desistimiento clients who no longer own a unit but have history."""
        unique_clients: dict[str, str] = {}
        for u in all_units:
            if u.cliente and u.cliente not in unique_clients:
                unique_clients[u.cliente] = u.cliente

        log.info("  Upserting %d clients...", len(unique_clients))

        rows = [{"full_name": name} for name in unique_clients.values()]

        for batch in _batched(rows, BATCH_SIZE):
            result = (
                self.client.table("clients")
                .upsert(batch, on_conflict="full_name")
                .execute()
            )
            for row in result.data:
                self.client_ids[row["full_name"]] = row["id"]

        log.info("  Upserted %d clients, got %d IDs", len(rows), len(self.client_ids))

    # ── Phase 3.5: Sales reps ────────────────────────────────────────────

    def upsert_sales_reps(self, all_units: list[ParsedUnit]) -> None:
        """Upsert sales_reps so sales.sales_rep_id FK is satisfied."""
        name_map = {
            "walk_in": "Puerta Abierta",
            "unknown": "Unknown / Directo",
            "05": "Sales Rep 05",
            "06": "Sales Rep 06",
            "35": "Sales Rep 35",
            "GV1": "Sales Rep GV1",
        }
        unique_reps: dict[str, str] = {}
        for u in all_units:
            rid = normalize_vendedor(u.vendedor) or "unknown"
            if rid not in unique_reps:
                unique_reps[rid] = name_map.get(rid, rid)

        if not unique_reps:
            return

        rows = [{"id": rid, "name": name} for rid, name in unique_reps.items()]

        log.info("  Upserting %d sales reps...", len(rows))
        for batch in _batched(rows, BATCH_SIZE):
            self.client.table("sales_reps").upsert(batch, on_conflict="id").execute()
        log.info("  Upserted %d sales reps", len(rows))

    # ── Phase 4: Sales ────────────────────────────────────────────────────

    def _build_sale_row(self, u: ParsedUnit) -> dict | None:
        """Build a sale dict from a ParsedUnit. Returns None if missing deps."""
        unit_id = self.unit_ids.get(u.apto)
        if not unit_id:
            log.warning("  Skipping sale for Apto %s: unit_id not found", u.apto)
            return None

        client_id = self.client_ids.get(u.cliente) if u.cliente else None
        if not client_id:
            log.warning("  Apto %s: no client_id for '%s', skipping sale", u.apto, u.cliente)
            return None

        financed = u.saldo_financiar or (
            round((u.precio_venta or 0) - (u.enganche or 0), 2) if u.precio_venta else 0
        )

        sale_status = normalize_sale_status(u.estatus_raw) or "active"

        notes_parts = []
        if u.source_section != "main":
            notes_parts.append(f"Source: {u.source_section}")

        precio = u.precio_venta or 0
        tax = round((u.iva or 0) + (u.timbres or 0), 2)

        return {
            "project_id": self.project_id,
            "unit_id": unit_id,
            "client_id": client_id,
            "sales_rep_id": normalize_vendedor(u.vendedor) or "unknown",
            "sale_date": u.fecha_reserva.isoformat() if u.fecha_reserva else "2023-03-01",
            "price_with_tax": precio,
            "price_without_tax": round(precio - tax, 2),
            "down_payment_amount": u.enganche or 0,
            "financed_amount": financed,
            "status": sale_status,
            "referral_applies": False,
            "caso_especial": u.caso_especial is not None,
            "caso_especial_type": u.caso_especial,
            "observaciones": u.observaciones,
            "notas": "; ".join(filter(None, [u.notas, ("; ".join(notes_parts) if notes_parts else None)])) or None,
        }

    def insert_sales(self, all_units: list[ParsedUnit]) -> None:
        """Insert sales preserving complete unit lifecycle history.

        Model:
          - MAIN active/reserved → upsert on partial index (1 active per unit)
          - DESISTIMIENTO cancelled → idempotent insert (N cancelled per unit)

        Each ParsedUnit gets its own sale_id via self._sale_ids[id(unit)].
        This is critical because payments bind to the sale, not the unit.
        """
        sellable = [u for u in all_units if normalize_sale_status(u.estatus_raw) is not None]
        log.info("  Processing %d sales (active + cancelled history)...", len(sellable))

        # Build reverse lookup: unit_id → apto
        uid_to_apto = {uid: apto for apto, uid in self.unit_ids.items()}

        active_units: list[ParsedUnit] = []
        cancelled_units: list[ParsedUnit] = []

        for u in sellable:
            row = self._build_sale_row(u)
            if row is None:
                continue
            # Stash the built row on the unit for later use
            u._sale_row = row  # type: ignore[attr-defined]
            if row["status"] == "active":
                active_units.append(u)
            else:
                cancelled_units.append(u)

        # ── Active sales: upsert on partial unique index ──────────────────
        # UNIQUE(unit_id) WHERE status = 'active' → at most 1 active per unit

        if active_units:
            active_rows = [u._sale_row for u in active_units]  # type: ignore[attr-defined]
            for batch_units, batch_rows in zip(
                _batched(active_units, BATCH_SIZE),
                _batched(active_rows, BATCH_SIZE),
            ):
                result = (
                    self.client.table("sales")
                    .upsert(batch_rows, on_conflict="unit_id")
                    .execute()
                )
                # Map returned sale IDs back to ParsedUnits
                returned = {r["unit_id"]: r["id"] for r in result.data}
                for u in batch_units:
                    unit_id = self.unit_ids.get(u.apto)
                    if unit_id and unit_id in returned:
                        self._sale_ids[id(u)] = returned[unit_id]

        log.info("  Active sales upserted: %d", len(active_units))

        # ── Cancelled sales: idempotent insert per (unit + client) ────────
        # No unique constraint on cancelled sales, so we check before insert
        # to make re-runs safe. Key: (unit_id, client_id, status=cancelled)

        inserted_cancelled = 0
        for u in cancelled_units:
            row = u._sale_row  # type: ignore[attr-defined]

            existing = (
                self.client.table("sales")
                .select("id")
                .eq("unit_id", row["unit_id"])
                .eq("client_id", row["client_id"])
                .eq("status", "cancelled")
                .limit(1)
                .execute()
            )

            if existing.data:
                self._sale_ids[id(u)] = existing.data[0]["id"]
            else:
                result = self.client.table("sales").insert(row).execute()
                self._sale_ids[id(u)] = result.data[0]["id"]
                inserted_cancelled += 1

        log.info(
            "  Cancelled sales: %d inserted, %d already existed",
            inserted_cancelled,
            len(cancelled_units) - inserted_cancelled,
        )
        log.info(
            "  Total sales mapped: %d (active: %d, cancelled: %d)",
            len(self._sale_ids),
            len(active_units),
            len(cancelled_units),
        )

    # ── Phase 5: Actual Payments ──────────────────────────────────────────

    def insert_payments(self, all_units: list[ParsedUnit]) -> None:
        """Insert actual payments, each linked to its specific sale.

        Payments bind to the SALE, not the unit. This means:
          - Active sale payments → linked to the current owner's sale
          - Cancelled sale payments → linked to the old owner's sale
          - The unit itself doesn't "remember" cancelled payments
        """
        rows: list[dict] = []
        skipped_no_sale = 0

        for u in all_units:
            sale_id = self.get_sale_id(u)
            if not u.actual_payments:
                continue
            if not sale_id:
                skipped_no_sale += 1
                log.debug("  Apto %s (%s): %d payments orphaned — no sale_id",
                          u.apto, u.source_section, len(u.actual_payments))
                continue

            for idx, (pdate, amount) in enumerate(sorted(u.actual_payments)):
                ptype = "reservation" if idx == 0 else "down_payment"

                rows.append(
                    {
                        "sale_id": sale_id,
                        "payment_date": pdate.isoformat(),
                        "amount": amount,
                        "payment_type": ptype,
                        "notes": f"ETL import from {u.source_section}",
                    }
                )

        log.info("  Inserting %d actual payments (%d units skipped, no sale)...",
                 len(rows), skipped_no_sale)

        inserted = 0
        for batch in _batched(rows, BATCH_SIZE):
            result = self.client.table("payments").insert(batch).execute()
            inserted += len(result.data)

        log.info("  Inserted %d actual payments", inserted)

    # ── Phase 6: Expected Payments ────────────────────────────────────────

    def upsert_expected_payments(self, expected: list[ParsedExpected]) -> None:
        log.info("  Upserting %d expected payments...", len(expected))

        # Group by apto to assign installment numbers
        by_apto: dict[str, list[ParsedExpected]] = {}
        for r in expected:
            by_apto.setdefault(r.apto, []).append(r)

        rows = []
        for apto, records in by_apto.items():
            for idx, r in enumerate(sorted(records, key=lambda x: x.due_date), start=1):
                rows.append(
                    {
                        "project_id": self.project_id,
                        "unit_number": apto,
                        "due_date": r.due_date.isoformat(),
                        "amount": r.amount,
                        "installment_number": idx,
                        "schedule_type": "budget",
                    }
                )

        inserted = 0
        for batch in _batched(rows, BATCH_SIZE):
            result = (
                self.client.table("expected_payments")
                .upsert(
                    batch,
                    on_conflict="project_id,unit_number,due_date,schedule_type",
                )
                .execute()
            )
            inserted += len(result.data)

        log.info("  Upserted %d expected payments", inserted)

    # ── Phase 7: Verify ──────────────────────────────────────────────────

    def verify(self) -> None:
        log.info("═══ VERIFICATION ═══")

        for table in ["projects", "units", "clients", "sales", "payments", "expected_payments"]:
            if table == "projects":
                result = self.client.table(table).select("id", count="exact").eq("name", PROJECT_NAME).execute()
            elif table in ("units", "expected_payments"):
                result = self.client.table(table).select("id", count="exact").eq("project_id", self.project_id).execute()
            elif table == "sales":
                result = self.client.table(table).select("id", count="exact").eq("project_id", self.project_id).execute()
            elif table == "payments":
                # Count payments linked to our sales
                result = self.client.rpc(
                    "get_count",  # fallback: just count all
                    {},
                ).execute()
                # Simpler: just count via join
                log.info("  payments: (use DB query to verify)")
                continue
            else:
                result = self.client.table(table).select("id", count="exact").execute()

            count = result.count if hasattr(result, "count") and result.count is not None else len(result.data)
            log.info("  %s: %d rows", table, count)


# ─────────────────────────────────────────────────────────────────────────────
# Dry-run report
# ─────────────────────────────────────────────────────────────────────────────


def print_dry_run(
    main_units: list[ParsedUnit],
    desist_units: list[ParsedUnit],
    prob_desist: list[ParsedUnit],
    expected: list[ParsedExpected],
) -> None:
    all_units = main_units + desist_units + prob_desist

    # Units breakdown
    sold = [u for u in main_units if normalize_unit_status(u.estatus_raw) == "sold"]
    reserved = [u for u in main_units if normalize_unit_status(u.estatus_raw) == "reserved"]
    available = [u for u in main_units if normalize_unit_status(u.estatus_raw) == "available"]
    clients_set = {u.cliente for u in all_units if u.cliente}
    main_payments = sum(len(u.actual_payments) for u in main_units)
    desist_payments = sum(len(u.actual_payments) for u in desist_units + prob_desist)

    active_sales = len([u for u in main_units if normalize_sale_status(u.estatus_raw)])
    cancelled_sales = len(desist_units) + len(prob_desist)

    # Lifecycle stats
    desist_aptos = {u.apto for u in desist_units + prob_desist}
    active_aptos = {u.apto for u in main_units if normalize_unit_status(u.estatus_raw) != "available"}
    resold = desist_aptos & active_aptos

    log.info("╔════════════════════════════════════════════════════════╗")
    log.info("║            DRY-RUN SUMMARY — BOULEVARD               ║")
    log.info("╠════════════════════════════════════════════════════════╣")
    log.info("║  projects              →     1 row                    ║")
    log.info("║  units (from main)     →  %4d (sold:%d res:%d avail:%d)  ║",
             len(main_units), len(sold), len(reserved), len(available))
    log.info("║  clients               →  %4d unique names            ║", len(clients_set))
    log.info("║  sales (active)        →  %4d                         ║", active_sales)
    log.info("║  sales (cancelled)     →  %4d (desist:%d prob:%d)       ║",
             cancelled_sales, len(desist_units), len(prob_desist))
    log.info("║  payments (active)     →  %4d records                 ║", main_payments)
    log.info("║  payments (cancelled)  →  %4d records                 ║", desist_payments)
    log.info("║  expected_payments     →  %4d records                 ║", len(expected))
    log.info("╠════════════════════════════════════════════════════════╣")
    log.info("║  Units with lifecycle history (desisted + re-sold): %d  ║", len(resold))
    log.info("╚════════════════════════════════════════════════════════╝")

    # Sample: lifecycle unit
    if resold:
        sample_apto = sorted(resold)[0]
        main_u = next(u for u in main_units if u.apto == sample_apto)
        desist_u = next((u for u in desist_units + prob_desist if u.apto == sample_apto), None)
        log.info("\n── Sample lifecycle: Apto %s ──", sample_apto)
        if desist_u:
            log.info("  CANCELLED: client=%s, payments=%d", desist_u.cliente, len(desist_u.actual_payments))
        log.info("  ACTIVE:    client=%s, payments=%d", main_u.cliente, len(main_u.actual_payments))

    # Sample: expected payments
    if expected:
        sample_apto = expected[0].apto
        apto_expected = [r for r in expected if r.apto == sample_apto]
        log.info("\n── Sample: Expected payments ──")
        log.info("  Apto %s: %d installments (%s → %s)",
                 sample_apto, len(apto_expected),
                 apto_expected[0].due_date, apto_expected[-1].due_date)

    log.info("\n→ Run with --execute to insert into Supabase")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(description="ETL: Boulevard → Supabase")
    parser.add_argument("file", type=Path, help="Path to Reservas.xlsx")
    parser.add_argument("--execute", action="store_true", help="Actually insert into DB (default: dry-run)")
    parser.add_argument("--verbose", action="store_true", help="Debug logging")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    if not args.file.exists():
        log.error("File not found: %s", args.file)
        return 1

    # ── Parse ──────────────────────────────────────────────────────────────

    log.info("═══ PARSING %s ═══", args.file.name)
    wb = openpyxl.load_workbook(str(args.file), data_only=True)

    log.info("Parsing BOULEVARD 5...")
    ws_b5 = wb[SHEET_ACTUALS]

    # Parse main section first to capture monthly column positions
    _, main_month_cols = discover_headers(ws_b5, B5_HEADER_ROW, 2, 84, B5_HEADER_PATTERNS)

    main_units = parse_b5_section(ws_b5, B5_HEADER_ROW, B5_DATA_START, B5_DATA_END, 2, 84, "main")
    desist_units = parse_b5_section(ws_b5, B5_DESIST_HEADER, B5_DESIST_START, B5_DESIST_END, 2, 84, "desistimiento", main_month_cols)
    prob_desist = parse_b5_section(ws_b5, B5_PROB_DESIST_HEADER, B5_PROB_DESIST_START, B5_PROB_DESIST_END, 2, 84, "prob_desistimiento", main_month_cols)

    # Force desistimiento status for cancelled sales
    for u in desist_units + prob_desist:
        if u.estatus_raw and "disponible" in u.estatus_raw.lower():
            u.estatus_raw = "Desistimiento"
        elif u.estatus_raw != "Desistimiento":
            u.estatus_raw = "Desistimiento"

    log.info("Parsing BOULEVARD PPTO...")
    ws_ppto = wb[SHEET_BUDGET]
    expected = parse_ppto(ws_ppto)

    # ── Validate ──────────────────────────────────────────────────────────

    log.info("═══ VALIDATING ═══")
    report = validate(main_units, desist_units, prob_desist, expected)
    report.print_summary()

    if report.has_errors:
        log.error("Aborting due to validation errors.")
        return 1

    # ── Dry-run or Execute ────────────────────────────────────────────────

    if not args.execute:
        print_dry_run(main_units, desist_units, prob_desist, expected)
        return 0

    # Verify env
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        log.error("Set SUPABASE_URL and SUPABASE_KEY environment variables")
        return 1

    # ── Insert ────────────────────────────────────────────────────────────

    log.info("═══ INSERTING INTO SUPABASE ═══")
    all_units = main_units + desist_units + prob_desist

    loader = SupabaseLoader(url, key)
    loader.upsert_project()
    loader.upsert_units(main_units)          # Only main → unit is current state
    loader.upsert_clients(all_units)          # All clients including desisted
    loader.upsert_sales_reps(all_units)      # FK for sales.sales_rep_id
    loader.insert_sales(all_units)            # Active + cancelled history
    loader.insert_payments(all_units)         # Each payment → its own sale
    loader.upsert_expected_payments(expected)

    # ── Verify ────────────────────────────────────────────────────────────

    loader.verify()

    log.info("═══ DONE ═══")
    return 0


if __name__ == "__main__":
    sys.exit(main())

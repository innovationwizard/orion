#!/usr/bin/env python3
"""
Production Database Backfill — Reservation History from SSOT xlsx files.

Reads the 4 "Reporte de Ventas" xlsx workbooks and the 4 "Disponibilidad" xlsx
workbooks. Extracts client names, reservation dates, salespeople, and desistimiento
info. Generates a single idempotent SQL file that populates rv_clients, reservations,
reservation_clients, unit_status_log, and freeze_requests.

Safety:
  - All INSERTs into empty tables (pre-flight check in SQL)
  - Wrapped in a single transaction (BEGIN / COMMIT)
  - No DELETEs or UPDATEs
  - Historical created_at timestamps preserved
  - Review backfill_reservations.sql before executing

Usage:
  python3 -u scripts/backfill_reservations.py
  # Review scripts/backfill_reservations.sql
  # Then execute via Supabase Management API
"""

from __future__ import annotations

import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
SSOT = ROOT / "origin" / "SSOT" / "Reservas y Ventas"

# Reporte de Ventas files (primary source — transactional history)
REPORTE_FILES = {
    "blt": SSOT / "Bosque Las Tapias" / "2. Reporte  de Ventas Bosque Las Tapias septiembre 2025.xlsx",
    "ben": SSOT / "Benestare" / "2. Reporte  de Ventas BENESTARE septiembre 2025.xlsx",
    "b5":  SSOT / "Boulevard 5" / "2. Reporte  de Ventas BOULEVARD 5 septiembre 2025.xlsx",
    "ce":  SSOT / "Casa Elisa" / "2 .Reporte de Ventas CASA ELISA septiembre 2025.xlsx",
}

# Disponibilidad files (secondary source — current snapshot with client+date for BEN)
DISP_FILES = {
    "blt": SSOT / "Bosque Las Tapias" / "Disponibilidad y cotizador Bosque Las Tapias septiembre 2025.xlsx",
    "ben": SSOT / "Benestare" / "Precios y Disponibilidad BENESTARE TA septiembre 25.xlsx",
    "b5":  SSOT / "Boulevard 5" / "DisponibilIdad e Integración B5 septiembre 2025.xlsx",
    "ce":  SSOT / "Casa Elisa" / "CASA ELISA con trazabilidad precios septiembre 2025.xlsx",
}

PROJECT_IDS = {
    "blt": "019c7d10-8ee5-7999-9881-2cd5ad038aa9",
    "ce":  "019c7d10-8e7b-7db6-93be-6f42d0538233",
    "b5":  "019c7d10-8e01-720f-942f-cac0017d83a8",
    "ben": "019c7d10-8f5a-74c7-b3df-c2151ad8a376",
}

PROJECT_NAMES = {
    "blt": "Bosque Las Tapias",
    "ce":  "Casa Elisa",
    "b5":  "Boulevard 5",
    "ben": "Benestare",
}

# Tower configurations — which projects have multiple towers
SINGLE_TOWER_PROJECTS = {"b5", "ce"}
DEFAULT_TOWER = {
    "b5": "Principal",
    "ce": "Principal",
}

# Salesperson canonical names (from seed_prod.py)
SALESPERSON_CANONICAL: dict[str, str] = {
    "paula hernandez":  "Paula Hernandez",
    "paula hernadez":   "Paula Hernandez",
    "paula hernande":   "Paula Hernandez",
    "jose gutierrez":   "Jose Gutierrez",
    "antonio rada":     "Antonio Rada",
    "brenda bucaro":    "Brenda Bucaro",
    "alexander franco": "Alexander Franco",
    "erwin cardona":    "Erwin Cardona",
    "anahi cisneros":   "Anahi Cisneros",
    "diana alvarez":    "Diana Alvarez",
    "sofia paredes":    "Sofia Paredes",
    "laura molina":     "Laura Molina",
    "efren sanchez":    "Efren Sanchez",
    "eder veliz":       "Eder Veliz",
    "pedro pablo sarti":"Pedro Pablo Sarti",
    "ronaldo ogaldez":  "Ronaldo Ogaldez",
    "pablo marroquin":  "Pablo Marroquin",
    "rony ramirez":     "Rony Ramirez",
    "noemi menendez":   "Noemi Menendez",
    "mario rodriguez":  "Mario Rodriguez",
    "forma capital":    "Forma Capital",
    "junta directiva":  "Junta Directiva",
    "alejandra calderon": "Alejandra Calderon",
    "andrea gonzalez":  "Andrea Gonzalez",
    "abigail garcia":   "Abigail Garcia",
    "ivan castillo":    "Ivan Castillo",
    "karina fuentes":   "Karina Fuentes",
    "luis esteban":     "Luis Esteban",
    "otto h.":          "Otto Herrera",
    "otto herrera":     "Otto Herrera",
}

SALESPERSON_EXCLUDE = {"traslado", "jd"}

# Status mapping from Disponibilidad files
STATUS_MAP: dict[str, str] = {
    "disponible": "AVAILABLE",
    "reservado":  "RESERVED",
    "pcv":        "SOLD",
    "promesa":    "SOLD",
    "vendido":    "SOLD",
    "congelado":  "FROZEN",
    "congelado junta directiva": "FROZEN",
    "liberado":   "AVAILABLE",
}

# Fallback date for entries with no date at all
FALLBACK_DATE = "2025-09-01"


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class SaleRecord:
    """A single sale/reservation from a Ventas sheet."""
    project_key: str
    unit_number: str
    tower_name: str | None  # None if sheet lacks Torre column
    client_names: list[str]  # Split co-buyers
    salesperson: str | None
    fecha: date | None
    enganche: float | None
    promesa_firmada: bool
    month_label: str  # Sheet name, for debugging


@dataclass
class DesistimientoRecord:
    """A desistimiento/cancellation from the Desistimientos sheet."""
    project_key: str
    unit_number: str
    tower_name: str | None
    client_names: list[str]
    salesperson: str | None
    fecha_reserva: date | None
    fecha_desistimiento: date | None
    motivo: str | None
    month_label: str | None  # MES column


@dataclass
class DispRecord:
    """Current unit state from Disponibilidad file."""
    project_key: str
    tower_name: str
    unit_number: str
    status: str  # AVAILABLE, RESERVED, SOLD, FROZEN
    client_name: str | None
    salesperson: str | None
    fecha: date | None
    enganche: float | None


@dataclass
class ReservationOut:
    """Finalized reservation record for SQL generation."""
    project_key: str
    tower_name: str
    unit_number: str
    floor_number: int
    client_names: list[str]
    salesperson: str
    status: str  # CONFIRMED or DESISTED
    fecha: date | None
    enganche: float | None
    is_resale: bool
    desistimiento_reason: str | None
    desistimiento_date: date | None
    promesa_firmada: bool


@dataclass
class FreezeOut:
    """Finalized freeze request for SQL generation."""
    project_key: str
    tower_name: str
    unit_number: str
    salesperson: str
    reason: str
    vip_name: str | None
    fecha: date | None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_salesperson(raw: str | None) -> str | None:
    if raw is None:
        return None
    cleaned = str(raw).strip()
    if not cleaned:
        return None
    for sep in [" / ", "/", ","]:
        if sep in cleaned:
            cleaned = cleaned.split(sep)[0].strip()
            break
    if not cleaned:
        return None
    key = strip_accents(cleaned).lower()
    if key in SALESPERSON_EXCLUDE:
        return None
    return SALESPERSON_CANONICAL.get(key, cleaned)


def normalize_client_name(raw: str) -> str:
    """Clean a single client name."""
    name = raw.strip()
    # Remove leading/trailing punctuation
    name = name.strip(".,;:-")
    # Collapse whitespace
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def split_client_names(raw: str | None) -> list[str]:
    """Split co-buyers by '/' separator, normalize each."""
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    # Split by " / " or "/" — but be careful with initials like "J."
    if " / " in text:
        parts = text.split(" / ")
    elif "/" in text and not re.search(r'\b\w\./', text):
        parts = text.split("/")
    else:
        parts = [text]
    result = []
    for p in parts:
        name = normalize_client_name(p)
        if name and len(name) >= 3:
            result.append(name)
    return result


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def safe_date(val) -> date | None:
    """Extract a date from a cell value (datetime, date, or string)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try parsing common string formats
    text = str(val).strip()
    if not text:
        return None
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def sql_str(val: str | None) -> str:
    if val is None:
        return "NULL"
    escaped = val.replace("'", "''")
    return f"'{escaped}'"


def sql_num(val: float | int | None) -> str:
    if val is None:
        return "NULL"
    return str(val)


def sql_date(val: date | None) -> str:
    if val is None:
        return "NULL"
    return f"'{val.isoformat()}'"


def sql_bool(val: bool) -> str:
    return "true" if val else "false"


def normalize_unit_number(raw) -> str | None:
    """Normalize unit number to string for matching."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    # Handle locale IDs (L-1, L-2, etc.)
    if text.upper().startswith("L-"):
        return text.upper()
    # Integer unit numbers
    n = safe_int(text)
    if n is not None and n >= 100:
        return str(n)
    # Text-based like "L-1"
    if re.match(r'^[A-Z]-\d+$', text, re.IGNORECASE):
        return text.upper()
    return None


def normalize_tower(raw: str | None, project_key: str) -> str | None:
    """Normalize tower name for matching against seeded data."""
    if project_key in SINGLE_TOWER_PROJECTS:
        return DEFAULT_TOWER[project_key]
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    # "Torre C" → "Torre C", "C" → "Torre C", "TC" → "Torre C"
    text = text.strip()
    if re.match(r'^[A-E]$', text, re.IGNORECASE):
        return f"Torre {text.upper()}"
    if re.match(r'^T[A-E]$', text, re.IGNORECASE):
        return f"Torre {text[1].upper()}"
    if text.lower().startswith("torre"):
        letter = text.split()[-1].upper() if len(text.split()) > 1 else text[-1].upper()
        return f"Torre {letter}"
    return text


def floor_from_unit(unit_number: str) -> int:
    """Derive floor number from unit number."""
    if unit_number.upper().startswith("L-"):
        return 1
    n = safe_int(unit_number)
    if n is not None:
        return n // 100
    return 1


def normalize_status(raw: str | None) -> str:
    if raw is None:
        return "AVAILABLE"
    key = str(raw).strip().lower()
    try:
        float(key)
        return "SOLD"
    except ValueError:
        pass
    return STATUS_MAP.get(key, "AVAILABLE")


# ---------------------------------------------------------------------------
# Header detection for Ventas sheets
# ---------------------------------------------------------------------------

# Known header patterns — maps normalized header text → semantic field name
VENTAS_HEADER_MAP = {
    "fecha": "fecha",
    "cliente": "cliente",
    "no. apartamento": "unit",
    "no apartamento": "unit",
    "no. de apartamento": "unit",
    "torre": "torre",
    "asesor": "asesor",
    "enganche": "enganche",
    "promesa firmada": "promesa",
    "falta firma": "falta_firma",
    "medio": "medio",
    "precio de venta": "precio",
    "fecha de promesa": "fecha_promesa",
}


def detect_ventas_headers(ws, max_scan_rows: int = 5) -> tuple[int, dict[str, int]]:
    """Scan first rows to find header row and map semantic fields to column indices.

    Returns (header_row, {field_name: col_index}).
    """
    for row_num in range(1, max_scan_rows + 1):
        cols: dict[str, int] = {}
        for col in range(1, 30):
            val = cell(ws, row_num, col)
            if val is None:
                continue
            text = str(val).strip().lower()
            # Strip trailing spaces and periods
            text = text.rstrip(". ")
            for pattern, field_name in VENTAS_HEADER_MAP.items():
                if text == pattern or text.startswith(pattern):
                    if field_name not in cols:
                        cols[field_name] = col
                    break
        # A valid header row must have at least "cliente" and "unit" (or "asesor")
        if "cliente" in cols and ("unit" in cols or "asesor" in cols):
            return row_num, cols
    return 0, {}


# ---------------------------------------------------------------------------
# Desistimientos header detection
# ---------------------------------------------------------------------------

DESIST_HEADER_MAP = {
    "vendedor": "vendedor",
    "1 nombre del cliente": "cliente1",
    "2 nombre del cliente": "cliente2",
    "apartamento": "unit",
    "torre": "torre",
    "fecha de reserva": "fecha_reserva",
    "fecha deposito": "fecha_deposito",
    "precio de venta": "precio",
    "mes": "mes",
    "ano": "ano",
    "motivo": "motivo",
    "fecha de la carta": "fecha_carta",
    "fecha": "fecha",
    "observaciones": "observaciones",
}


def detect_desist_headers(ws, max_scan_rows: int = 5) -> tuple[int, dict[str, int]]:
    """Detect Desistimientos header row and column mapping."""
    for row_num in range(1, max_scan_rows + 1):
        cols: dict[str, int] = {}
        for col in range(1, 20):
            val = cell(ws, row_num, col)
            if val is None:
                continue
            text = strip_accents(str(val).strip().lower().rstrip(". "))
            for pattern, field_name in DESIST_HEADER_MAP.items():
                if text == pattern or (len(pattern) > 5 and text.startswith(pattern)):
                    if field_name not in cols:
                        cols[field_name] = col
                    break
        if "unit" in cols and ("vendedor" in cols or "cliente1" in cols):
            return row_num, cols
    return 0, {}


# ---------------------------------------------------------------------------
# Reporte de Ventas extraction — Ventas sheets
# ---------------------------------------------------------------------------

def extract_ventas_sheets(project_key: str, wb_path: Path) -> list[SaleRecord]:
    """Extract all sale records from monthly Ventas sheets."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    records: list[SaleRecord] = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.lower().startswith("ventas"):
            continue

        ws = wb[sheet_name]
        header_row, cols = detect_ventas_headers(ws)
        if header_row == 0:
            print(f"  WARN: No headers found in '{sheet_name}' — skipping")
            continue

        if "cliente" not in cols:
            print(f"  WARN: No 'Cliente' column in '{sheet_name}' — skipping")
            continue

        row_count = 0
        for row in range(header_row + 1, (ws.max_row or 0) + 1):
            # Get client — skip empty rows
            client_raw = cell(ws, row, cols["cliente"])
            if client_raw is None or str(client_raw).strip() == "":
                continue

            # Get unit number
            unit_raw = cell(ws, row, cols["unit"]) if "unit" in cols else None
            unit_num = normalize_unit_number(unit_raw)
            if unit_num is None:
                continue

            # Get tower
            torre_raw = cell(ws, row, cols["torre"]) if "torre" in cols else None
            torre = normalize_tower(str(torre_raw) if torre_raw else None, project_key)

            # Get salesperson
            sp_raw = cell(ws, row, cols["asesor"]) if "asesor" in cols else None
            sp = normalize_salesperson(str(sp_raw) if sp_raw else None)

            # Get date
            fecha_raw = cell(ws, row, cols["fecha"]) if "fecha" in cols else None
            fecha = safe_date(fecha_raw)

            # Get enganche
            eng_raw = cell(ws, row, cols["enganche"]) if "enganche" in cols else None
            enganche = safe_float(eng_raw)

            # Get promesa firmada
            promesa_raw = cell(ws, row, cols["promesa"]) if "promesa" in cols else None
            promesa = promesa_raw is not None and str(promesa_raw).strip().lower() not in ("", "0", "no", "false")

            # Split client names
            clients = split_client_names(str(client_raw))
            if not clients:
                continue

            records.append(SaleRecord(
                project_key=project_key,
                unit_number=unit_num,
                tower_name=torre,
                client_names=clients,
                salesperson=sp,
                fecha=fecha,
                enganche=enganche,
                promesa_firmada=promesa,
                month_label=sheet_name,
            ))
            row_count += 1

        if row_count > 0:
            print(f"    {sheet_name}: {row_count} records")

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Reporte de Ventas extraction — Desistimientos sheet
# ---------------------------------------------------------------------------

def extract_desistimientos(project_key: str, wb_path: Path) -> list[DesistimientoRecord]:
    """Extract desistimiento records from the Desistimientos sheet."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # Find the Desistimientos sheet
    desist_sheet = None
    for name in wb.sheetnames:
        if "desistimiento" in name.lower():
            desist_sheet = name
            break

    if desist_sheet is None:
        print(f"  WARN: No Desistimientos sheet found")
        wb.close()
        return []

    ws = wb[desist_sheet]
    header_row, cols = detect_desist_headers(ws)
    if header_row == 0:
        print(f"  WARN: Could not detect headers in '{desist_sheet}'")
        wb.close()
        return []

    records: list[DesistimientoRecord] = []
    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        # Get unit number
        unit_raw = cell(ws, row, cols["unit"]) if "unit" in cols else None
        unit_num = normalize_unit_number(unit_raw)
        if unit_num is None:
            continue

        # Get client names
        c1_raw = cell(ws, row, cols["cliente1"]) if "cliente1" in cols else None
        c2_raw = cell(ws, row, cols["cliente2"]) if "cliente2" in cols else None
        clients = []
        for c in [c1_raw, c2_raw]:
            if c is not None:
                name = normalize_client_name(str(c))
                if name and len(name) >= 3:
                    clients.append(name)
        if not clients:
            continue

        # Get tower
        torre_raw = cell(ws, row, cols["torre"]) if "torre" in cols else None
        torre = normalize_tower(str(torre_raw) if torre_raw else None, project_key)

        # Get salesperson
        sp_raw = cell(ws, row, cols["vendedor"]) if "vendedor" in cols else None
        sp = normalize_salesperson(str(sp_raw) if sp_raw else None)

        # Get dates
        fecha_res_raw = cell(ws, row, cols["fecha_reserva"]) if "fecha_reserva" in cols else None
        fecha_reserva = safe_date(fecha_res_raw)

        # Desistimiento date: try "fecha" column, then "fecha_carta", then "mes"
        fecha_desist = None
        for col_name in ["fecha", "fecha_carta"]:
            if col_name in cols:
                val = cell(ws, row, cols[col_name])
                fecha_desist = safe_date(val)
                if fecha_desist:
                    break

        # If still no date, try to parse month from "mes" + "ano"
        if fecha_desist is None and "mes" in cols:
            mes_val = cell(ws, row, cols["mes"])
            if isinstance(mes_val, (datetime, date)):
                fecha_desist = safe_date(mes_val)

        # Motivo
        motivo_raw = cell(ws, row, cols["motivo"]) if "motivo" in cols else None
        motivo = str(motivo_raw).strip() if motivo_raw else None
        if motivo and len(motivo) < 3:
            motivo = None

        # Month label from MES column for debugging
        mes_raw = cell(ws, row, cols["mes"]) if "mes" in cols else None
        month_label = str(mes_raw).strip() if mes_raw else None

        records.append(DesistimientoRecord(
            project_key=project_key,
            unit_number=unit_num,
            tower_name=torre,
            client_names=clients,
            salesperson=sp,
            fecha_reserva=fecha_reserva,
            fecha_desistimiento=fecha_desist,
            motivo=motivo,
            month_label=month_label,
        ))

    print(f"    Desistimientos: {len(records)} records")
    wb.close()
    return records


# ---------------------------------------------------------------------------
# Disponibilidad extraction — current unit state with client info
# ---------------------------------------------------------------------------

def extract_disp_blt(wb_path: Path) -> list[DispRecord]:
    """Extract BLT Disponibilidad data."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    records: list[DispRecord] = []

    for sheet_name, tower_name in [("Precios Torre C", "Torre C"), ("Precios Torre B", "Torre B")]:
        ws = wb[sheet_name]
        is_c = tower_name == "Torre C"

        for row in range(3, (ws.max_row or 0) + 1):
            raw_unit = cell(ws, row, 2)
            unit_num = safe_int(raw_unit)
            if unit_num is None or unit_num < 100:
                continue

            if is_c:
                status_raw = cell(ws, row, 27)  # AA
                client_raw = cell(ws, row, 28)  # AB
                sales_raw = cell(ws, row, 29)   # AC
                eng_raw = cell(ws, row, 20)     # T = Reserva amount
            else:
                status_raw = cell(ws, row, 24)  # X
                client_raw = cell(ws, row, 25)  # Y
                sales_raw = cell(ws, row, 26)   # Z
                eng_raw = None

            status = normalize_status(str(status_raw) if status_raw else None)
            client = str(client_raw).strip() if client_raw else None
            sp = normalize_salesperson(str(sales_raw) if sales_raw else None)
            enganche = safe_float(eng_raw)

            records.append(DispRecord(
                project_key="blt", tower_name=tower_name,
                unit_number=str(unit_num), status=status,
                client_name=client if client and len(client) >= 3 else None,
                salesperson=sp, fecha=None, enganche=enganche,
            ))

    wb.close()
    return records


def extract_disp_ben(wb_path: Path) -> list[DispRecord]:
    """Extract Benestare Disponibilidad data."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Precios"]
    records: list[DispRecord] = []

    for row in range(4, (ws.max_row or 0) + 1):
        raw_unit = cell(ws, row, 4)  # D
        unit_num = safe_int(raw_unit)
        if unit_num is None or unit_num < 100:
            continue

        tower_letter = str(cell(ws, row, 11) or "").strip()  # K
        if not tower_letter or len(tower_letter) > 2:
            continue
        tower_name = f"Torre {tower_letter}"

        status_raw = cell(ws, row, 41)  # AO
        client_raw = cell(ws, row, 42)  # AP
        fecha_raw = cell(ws, row, 43)   # AQ = Fecha de Reserva
        sales_raw = cell(ws, row, 44)   # AR
        eng_raw = cell(ws, row, 26)     # Z = Reserva

        status = normalize_status(str(status_raw) if status_raw else None)
        client = str(client_raw).strip() if client_raw else None
        sp = normalize_salesperson(str(sales_raw) if sales_raw else None)
        fecha = safe_date(fecha_raw)
        enganche = safe_float(eng_raw)

        records.append(DispRecord(
            project_key="ben", tower_name=tower_name,
            unit_number=str(unit_num), status=status,
            client_name=client if client and len(client) >= 3 else None,
            salesperson=sp, fecha=fecha, enganche=enganche,
        ))

    wb.close()
    return records


def extract_disp_b5(wb_path: Path) -> list[DispRecord]:
    """Extract Boulevard 5 Disponibilidad data."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Matriz Precios A"]
    records: list[DispRecord] = []

    for row in range(4, (ws.max_row or 0) + 1):
        raw_unit = cell(ws, row, 2)  # B
        unit_num = safe_int(raw_unit)
        if unit_num is None or unit_num < 100:
            continue

        status_raw = cell(ws, row, 58)   # BF
        client_raw = cell(ws, row, 59)   # BG
        sales_raw = cell(ws, row, 60)    # BH
        eng_raw = cell(ws, row, 44)      # AR = Reserva

        status = normalize_status(str(status_raw) if status_raw else None)
        client = str(client_raw).strip() if client_raw else None
        sp = normalize_salesperson(str(sales_raw) if sales_raw else None)
        enganche = safe_float(eng_raw)

        records.append(DispRecord(
            project_key="b5", tower_name="Principal",
            unit_number=str(unit_num), status=status,
            client_name=client if client and len(client) >= 3 else None,
            salesperson=sp, fecha=None, enganche=enganche,
        ))

    wb.close()
    return records


def extract_disp_ce(wb_path: Path) -> list[DispRecord]:
    """Extract Casa Elisa Disponibilidad data."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Disponibilidad"]
    records: list[DispRecord] = []

    for row in range(5, (ws.max_row or 0) + 1):
        raw_unit = cell(ws, row, 2)  # B
        unit_num_str = str(raw_unit).strip() if raw_unit else ""
        unit_num_int = safe_int(raw_unit)

        is_locale = unit_num_str.startswith("L-")
        if unit_num_int is None and not is_locale:
            continue
        if unit_num_int is not None and unit_num_int < 100:
            continue

        unit_number = unit_num_str if is_locale else str(unit_num_int)

        status_raw = cell(ws, row, 49)  # AW
        client_raw = cell(ws, row, 48)  # AV
        sales_raw = cell(ws, row, 51)   # AY

        status = normalize_status(str(status_raw) if status_raw else None)
        client = str(client_raw).strip() if client_raw else None
        sp = normalize_salesperson(str(sales_raw) if sales_raw else None)

        records.append(DispRecord(
            project_key="ce", tower_name="Principal",
            unit_number=unit_number, status=status,
            client_name=client if client and len(client) >= 3 else None,
            salesperson=sp, fecha=None, enganche=None,
        ))

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Merge logic: combine Ventas + Desistimientos + Disponibilidad
# ---------------------------------------------------------------------------

def build_unit_key(project_key: str, tower_name: str, unit_number: str) -> str:
    return f"{project_key}:{tower_name}:{unit_number}"


def merge_data(
    all_ventas: dict[str, list[SaleRecord]],
    all_desist: dict[str, list[DesistimientoRecord]],
    all_disp: dict[str, list[DispRecord]],
) -> tuple[list[ReservationOut], list[FreezeOut]]:
    """Merge all data sources into finalized reservation and freeze records."""

    reservations: list[ReservationOut] = []
    freezes: list[FreezeOut] = []

    # Build indexed lookup from Disponibilidad
    disp_by_key: dict[str, DispRecord] = {}
    for project_key, disp_list in all_disp.items():
        for d in disp_list:
            key = build_unit_key(d.project_key, d.tower_name, d.unit_number)
            disp_by_key[key] = d

    # Build indexed lookup from Ventas (most recent record per unit wins)
    ventas_by_key: dict[str, SaleRecord] = {}
    for project_key, ventas_list in all_ventas.items():
        for v in ventas_list:
            # For multi-tower projects without torre, try to resolve from Disponibilidad
            tower = v.tower_name
            if tower is None and project_key not in SINGLE_TOWER_PROJECTS:
                # Try all towers for this project
                candidates = []
                for disp_key, disp_rec in disp_by_key.items():
                    if disp_rec.project_key == project_key and disp_rec.unit_number == v.unit_number:
                        candidates.append(disp_rec.tower_name)
                if len(candidates) == 1:
                    tower = candidates[0]
                else:
                    # Ambiguous or not found — skip
                    continue
            if tower is None:
                continue

            key = build_unit_key(project_key, tower, v.unit_number)
            existing = ventas_by_key.get(key)
            # Keep the most recent record (by fecha, or last-seen if no fecha)
            if existing is None:
                ventas_by_key[key] = SaleRecord(
                    project_key=v.project_key, unit_number=v.unit_number,
                    tower_name=tower, client_names=v.client_names,
                    salesperson=v.salesperson, fecha=v.fecha,
                    enganche=v.enganche, promesa_firmada=v.promesa_firmada,
                    month_label=v.month_label,
                )
            else:
                # Replace if newer date
                if v.fecha and (existing.fecha is None or v.fecha > existing.fecha):
                    ventas_by_key[key] = SaleRecord(
                        project_key=v.project_key, unit_number=v.unit_number,
                        tower_name=tower, client_names=v.client_names,
                        salesperson=v.salesperson, fecha=v.fecha,
                        enganche=v.enganche, promesa_firmada=v.promesa_firmada,
                        month_label=v.month_label,
                    )

    # Build desistimiento lookup
    desist_by_key: dict[str, DesistimientoRecord] = {}
    for project_key, desist_list in all_desist.items():
        for d in desist_list:
            tower = d.tower_name
            if tower is None and project_key not in SINGLE_TOWER_PROJECTS:
                candidates = []
                for disp_key, disp_rec in disp_by_key.items():
                    if disp_rec.project_key == project_key and disp_rec.unit_number == d.unit_number:
                        candidates.append(disp_rec.tower_name)
                if len(candidates) == 1:
                    tower = candidates[0]
                else:
                    continue
            if tower is None:
                continue

            key = build_unit_key(project_key, tower, d.unit_number)
            desist_by_key[key] = DesistimientoRecord(
                project_key=d.project_key, unit_number=d.unit_number,
                tower_name=tower, client_names=d.client_names,
                salesperson=d.salesperson, fecha_reserva=d.fecha_reserva,
                fecha_desistimiento=d.fecha_desistimiento, motivo=d.motivo,
                month_label=d.month_label,
            )

    # Now process every non-AVAILABLE unit from Disponibilidad
    processed_keys: set[str] = set()
    stats = defaultdict(int)

    for key, disp in sorted(disp_by_key.items()):
        if disp.status == "AVAILABLE":
            continue

        project_key = disp.project_key
        tower = disp.tower_name
        unit_num = disp.unit_number
        floor_num = floor_from_unit(unit_num)

        # Handle FROZEN units → freeze_requests
        if disp.status == "FROZEN":
            sp = disp.salesperson
            ventas_rec = ventas_by_key.get(key)
            if ventas_rec and ventas_rec.salesperson:
                sp = ventas_rec.salesperson
            if sp:
                freezes.append(FreezeOut(
                    project_key=project_key, tower_name=tower,
                    unit_number=unit_num, salesperson=sp,
                    reason="Congelado en inventario SSOT",
                    vip_name=disp.client_name,
                    fecha=ventas_rec.fecha if ventas_rec else None,
                ))
                stats["freeze"] += 1
            else:
                stats["freeze_no_sp"] += 1
            continue

        # RESERVED or SOLD → create reservation
        # Find best data source
        ventas_rec = ventas_by_key.get(key)
        desist_rec = desist_by_key.get(key)

        # Determine client names, salesperson, date
        client_names = None
        sp = None
        fecha = None
        enganche = None
        promesa = False

        # Priority: Ventas > Disponibilidad
        if ventas_rec:
            client_names = ventas_rec.client_names
            sp = ventas_rec.salesperson
            fecha = ventas_rec.fecha
            enganche = ventas_rec.enganche
            promesa = ventas_rec.promesa_firmada

        # Fill gaps from Disponibilidad
        if not client_names and disp.client_name:
            client_names = split_client_names(disp.client_name)
        if not sp and disp.salesperson:
            sp = disp.salesperson
        if not fecha and disp.fecha:
            fecha = disp.fecha
        if not enganche and disp.enganche:
            enganche = disp.enganche

        # Must have both client and salesperson to create a reservation
        if not client_names or not sp:
            stats["skip_no_data"] += 1
            continue

        # Check if this unit was desisted and re-reserved
        if desist_rec:
            # First: create the DESISTED reservation
            desist_sp = desist_rec.salesperson or sp
            desist_clients = desist_rec.client_names if desist_rec.client_names else client_names
            desist_fecha = desist_rec.fecha_reserva or desist_rec.fecha_desistimiento
            desist_date = desist_rec.fecha_desistimiento
            desist_motivo = desist_rec.motivo or "Desistimiento registrado en SSOT"

            # Ensure desistimiento_date exists (required by CHECK constraint)
            if desist_date is None:
                desist_date = desist_fecha  # Use reservation date as fallback

            if desist_date is not None:
                reservations.append(ReservationOut(
                    project_key=project_key, tower_name=tower,
                    unit_number=unit_num, floor_number=floor_num,
                    client_names=desist_clients, salesperson=desist_sp,
                    status="DESISTED", fecha=desist_fecha,
                    enganche=None, is_resale=False,
                    desistimiento_reason=desist_motivo,
                    desistimiento_date=desist_date,
                    promesa_firmada=False,
                ))
                stats["desisted"] += 1

                # Then: create the current CONFIRMED reservation (is_resale=true)
                reservations.append(ReservationOut(
                    project_key=project_key, tower_name=tower,
                    unit_number=unit_num, floor_number=floor_num,
                    client_names=client_names, salesperson=sp,
                    status="CONFIRMED", fecha=fecha,
                    enganche=enganche, is_resale=True,
                    desistimiento_reason=None, desistimiento_date=None,
                    promesa_firmada=promesa,
                ))
                stats["confirmed_resale"] += 1
            else:
                # No desistimiento date available — just create CONFIRMED
                reservations.append(ReservationOut(
                    project_key=project_key, tower_name=tower,
                    unit_number=unit_num, floor_number=floor_num,
                    client_names=client_names, salesperson=sp,
                    status="CONFIRMED", fecha=fecha,
                    enganche=enganche, is_resale=False,
                    desistimiento_reason=None, desistimiento_date=None,
                    promesa_firmada=promesa,
                ))
                stats["confirmed"] += 1
        else:
            # No desistimiento — simple CONFIRMED reservation
            reservations.append(ReservationOut(
                project_key=project_key, tower_name=tower,
                unit_number=unit_num, floor_number=floor_num,
                client_names=client_names, salesperson=sp,
                status="CONFIRMED", fecha=fecha,
                enganche=enganche, is_resale=False,
                desistimiento_reason=None, desistimiento_date=None,
                promesa_firmada=promesa,
            ))
            stats["confirmed"] += 1

        processed_keys.add(key)

    # Also create DESISTED reservations for units in Desistimientos
    # that are currently AVAILABLE (fully desisted, not re-reserved)
    for key, desist_rec in sorted(desist_by_key.items()):
        if key in processed_keys:
            continue
        disp = disp_by_key.get(key)
        if disp and disp.status != "AVAILABLE":
            continue  # Already handled above

        sp = desist_rec.salesperson
        if not sp:
            stats["desist_no_sp"] += 1
            continue
        if not desist_rec.client_names:
            stats["desist_no_client"] += 1
            continue

        desist_date = desist_rec.fecha_desistimiento
        desist_fecha = desist_rec.fecha_reserva or desist_date
        desist_motivo = desist_rec.motivo or "Desistimiento registrado en SSOT"

        if desist_date is None:
            desist_date = desist_fecha

        if desist_date is None:
            stats["desist_no_date"] += 1
            continue

        tower = desist_rec.tower_name
        if tower is None:
            continue
        floor_num = floor_from_unit(desist_rec.unit_number)

        reservations.append(ReservationOut(
            project_key=desist_rec.project_key, tower_name=tower,
            unit_number=desist_rec.unit_number, floor_number=floor_num,
            client_names=desist_rec.client_names, salesperson=sp,
            status="DESISTED", fecha=desist_fecha,
            enganche=None, is_resale=False,
            desistimiento_reason=desist_motivo,
            desistimiento_date=desist_date,
            promesa_firmada=False,
        ))
        stats["desisted_available"] += 1

    print(f"\n  Merge stats:")
    for k, v in sorted(stats.items()):
        print(f"    {k}: {v}")

    return reservations, freezes


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------

def generate_sql(
    reservations: list[ReservationOut],
    freezes: list[FreezeOut],
) -> str:
    lines: list[str] = []
    lines.append("-- ==========================================================================")
    lines.append("-- Production Backfill — Reservation History")
    lines.append("-- Generated by scripts/backfill_reservations.py")
    lines.append("-- ==========================================================================")
    lines.append("-- Safety: single transaction, INSERT only into empty tables")
    lines.append("-- Pre-flight: aborts if any target table already has data")
    lines.append("-- Review this file carefully before executing.")
    lines.append("-- ==========================================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # Pre-flight checks
    lines.append("-- Pre-flight: abort if any transactional table has data")
    lines.append("DO $$ BEGIN")
    for tbl in ["rv_clients", "reservations", "reservation_clients", "unit_status_log", "freeze_requests"]:
        lines.append(f"  IF (SELECT count(*) FROM {tbl}) > 0 THEN")
        lines.append(f"    RAISE EXCEPTION '{tbl} is not empty — aborting backfill';")
        lines.append(f"  END IF;")
    lines.append("END $$;")
    lines.append("")

    # Collect all unique salesperson names referenced by reservations/freezes
    all_sp_names: set[str] = set()
    for r in reservations:
        all_sp_names.add(r.salesperson)
    for f in freezes:
        all_sp_names.add(f.salesperson)
    sorted_sp = sorted(all_sp_names)

    # Step 0: Insert any salespeople not already in the DB
    lines.append(f"-- Step 0: Ensure all referenced salespeople exist ({len(sorted_sp)} names)")
    for sp in sorted_sp:
        display = sp.split()[0] if " " in sp else sp
        lines.append(
            f"INSERT INTO salespeople (full_name, display_name) "
            f"VALUES ({sql_str(sp)}, {sql_str(display)}) "
            f"ON CONFLICT (full_name) DO NOTHING;"
        )
    lines.append("")

    # Collect all unique client names across all reservations
    all_client_names: set[str] = set()
    for r in reservations:
        for name in r.client_names:
            all_client_names.add(name)

    # Step 1: Insert clients
    sorted_clients = sorted(all_client_names)
    lines.append(f"-- Step 1: Clients ({len(sorted_clients)} rows)")
    for name in sorted_clients:
        lines.append(
            f"INSERT INTO rv_clients (full_name) VALUES ({sql_str(name)});"
        )
    lines.append("")

    # Step 2: Insert reservations
    # Sort: DESISTED first (so the partial unique index on active reservations
    # doesn't block the CONFIRMED insert for the same unit)
    sorted_reservations = sorted(reservations, key=lambda r: (
        r.project_key, r.tower_name, r.unit_number,
        0 if r.status == "DESISTED" else 1,
    ))

    lines.append(f"-- Step 2: Reservations ({len(sorted_reservations)} rows)")
    for i, r in enumerate(sorted_reservations):
        pid = PROJECT_IDS[r.project_key]
        fecha_str = sql_date(r.fecha) if r.fecha else f"'{FALLBACK_DATE}'"

        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(r.tower_name)} "
            f"AND ru.unit_number = {sql_str(r.unit_number)})"
        )
        sp_subq = f"(SELECT id FROM salespeople WHERE full_name = {sql_str(r.salesperson)} LIMIT 1)"

        cols = [
            "unit_id", "salesperson_id", "status",
            "deposit_amount", "deposit_date",
            "is_resale", "created_at",
        ]
        vals = [
            unit_subq, sp_subq, f"'{r.status}'::rv_reservation_status",
            sql_num(r.enganche), fecha_str,
            sql_bool(r.is_resale), fecha_str,
        ]

        if r.status == "DESISTED":
            cols.extend(["desistimiento_reason", "desistimiento_date"])
            vals.extend([sql_str(r.desistimiento_reason), sql_date(r.desistimiento_date)])

        if r.status == "CONFIRMED":
            cols.extend(["reviewed_at"])
            vals.extend([fecha_str])

        lines.append(
            f"INSERT INTO reservations ({', '.join(cols)}) VALUES ({', '.join(vals)});"
        )
        # Add comment for readability every 50 rows
        if (i + 1) % 50 == 0:
            lines.append(f"-- ... {i + 1} / {len(sorted_reservations)}")
    lines.append("")

    # Step 3: Insert reservation_clients junction records
    lines.append(f"-- Step 3: Reservation ↔ Client links")
    link_count = 0
    for r in sorted_reservations:
        pid = PROJECT_IDS[r.project_key]
        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(r.tower_name)} "
            f"AND ru.unit_number = {sql_str(r.unit_number)})"
        )

        # For finding the right reservation: match on unit_id + status + is_resale
        if r.status == "DESISTED":
            res_subq = (
                f"(SELECT id FROM reservations "
                f"WHERE unit_id = {unit_subq} "
                f"AND status = 'DESISTED' "
                f"ORDER BY created_at ASC LIMIT 1)"
            )
        elif r.is_resale:
            res_subq = (
                f"(SELECT id FROM reservations "
                f"WHERE unit_id = {unit_subq} "
                f"AND status = 'CONFIRMED' "
                f"AND is_resale = true LIMIT 1)"
            )
        else:
            res_subq = (
                f"(SELECT id FROM reservations "
                f"WHERE unit_id = {unit_subq} "
                f"AND status = 'CONFIRMED' "
                f"AND is_resale = false LIMIT 1)"
            )

        for j, name in enumerate(r.client_names):
            client_subq = f"(SELECT id FROM rv_clients WHERE full_name = {sql_str(name)} LIMIT 1)"
            is_primary = j == 0

            lines.append(
                f"INSERT INTO reservation_clients (reservation_id, client_id, is_primary) "
                f"VALUES ({res_subq}, {client_subq}, {sql_bool(is_primary)});"
            )
            link_count += 1
    lines.append(f"-- Total links: {link_count}")
    lines.append("")

    # Step 4: Freeze requests
    lines.append(f"-- Step 4: Freeze requests ({len(freezes)} rows)")
    for fr in freezes:
        pid = PROJECT_IDS[fr.project_key]
        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(fr.tower_name)} "
            f"AND ru.unit_number = {sql_str(fr.unit_number)})"
        )
        sp_subq = f"(SELECT id FROM salespeople WHERE full_name = {sql_str(fr.salesperson)} LIMIT 1)"
        fecha_str = sql_date(fr.fecha) if fr.fecha else f"'{FALLBACK_DATE}'"

        lines.append(
            f"INSERT INTO freeze_requests (unit_id, salesperson_id, reason, vip_name, status, created_at) "
            f"VALUES ({unit_subq}, {sp_subq}, {sql_str(fr.reason)}, {sql_str(fr.vip_name)}, "
            f"'ACTIVE'::rv_freeze_request_status, {fecha_str});"
        )
    lines.append("")

    # Step 5: Unit status log — one entry per non-AVAILABLE unit
    # Gather all unique units from reservations + freezes
    unit_entries: list[tuple[str, str, str, str, str | None]] = []  # (project, tower, unit, status, fecha)
    seen_log_keys: set[str] = set()

    for r in sorted_reservations:
        if r.status != "CONFIRMED":
            continue  # Only log the active/current status
        key = build_unit_key(r.project_key, r.tower_name, r.unit_number)
        if key in seen_log_keys:
            continue
        seen_log_keys.add(key)
        fecha_str = r.fecha.isoformat() if r.fecha else FALLBACK_DATE
        unit_entries.append((r.project_key, r.tower_name, r.unit_number, "RESERVED", fecha_str))

    for fr in freezes:
        key = build_unit_key(fr.project_key, fr.tower_name, fr.unit_number)
        if key in seen_log_keys:
            continue
        seen_log_keys.add(key)
        fecha_str = fr.fecha.isoformat() if fr.fecha else FALLBACK_DATE
        unit_entries.append((fr.project_key, fr.tower_name, fr.unit_number, "FROZEN", fecha_str))

    lines.append(f"-- Step 5: Unit status log ({len(unit_entries)} rows)")
    for project_key, tower, unit_num, new_status, fecha_str in unit_entries:
        pid = PROJECT_IDS[project_key]
        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(tower)} "
            f"AND ru.unit_number = {sql_str(unit_num)})"
        )
        lines.append(
            f"INSERT INTO unit_status_log (unit_id, old_status, new_status, changed_by, reason, created_at) "
            f"VALUES ({unit_subq}, NULL, '{new_status}'::rv_unit_status, "
            f"'system:backfill', 'Backfill desde SSOT', '{fecha_str}');"
        )
    lines.append("")

    lines.append("COMMIT;")
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=== Backfill Reservation History ===\n")

    # Verify all files exist
    print("Checking Reporte de Ventas files...")
    for key, path in REPORTE_FILES.items():
        if not path.exists():
            print(f"  ERROR: Not found: {path}", file=sys.stderr)
            sys.exit(1)
        print(f"  Found: {path.name}")

    print("\nChecking Disponibilidad files...")
    for key, path in DISP_FILES.items():
        if not path.exists():
            print(f"  ERROR: Not found: {path}", file=sys.stderr)
            sys.exit(1)
        print(f"  Found: {path.name}")

    # Phase 1: Extract Ventas from Reporte de Ventas
    print("\n--- Phase 1: Extract sales records from Reporte de Ventas ---")
    all_ventas: dict[str, list[SaleRecord]] = {}
    all_desist: dict[str, list[DesistimientoRecord]] = {}

    for key in ["blt", "ben", "b5", "ce"]:
        print(f"\n  {PROJECT_NAMES[key]}:")
        all_ventas[key] = extract_ventas_sheets(key, REPORTE_FILES[key])
        all_desist[key] = extract_desistimientos(key, REPORTE_FILES[key])
        print(f"  Total ventas: {len(all_ventas[key])}, desistimientos: {len(all_desist[key])}")

    total_ventas = sum(len(v) for v in all_ventas.values())
    total_desist = sum(len(d) for d in all_desist.values())
    print(f"\n  Grand total: {total_ventas} ventas, {total_desist} desistimientos")

    # Phase 2: Extract Disponibilidad (current state)
    print("\n--- Phase 2: Extract current state from Disponibilidad ---")
    all_disp: dict[str, list[DispRecord]] = {}

    extractors = {
        "blt": extract_disp_blt,
        "ben": extract_disp_ben,
        "b5":  extract_disp_b5,
        "ce":  extract_disp_ce,
    }
    for key in ["blt", "ben", "b5", "ce"]:
        all_disp[key] = extractors[key](DISP_FILES[key])
        non_avail = sum(1 for d in all_disp[key] if d.status != "AVAILABLE")
        print(f"  {PROJECT_NAMES[key]}: {len(all_disp[key])} units, {non_avail} non-AVAILABLE")

    # Phase 3: Merge
    print("\n--- Phase 3: Merge data sources ---")
    reservations, freezes = merge_data(all_ventas, all_desist, all_disp)

    print(f"\n  Final counts:")
    print(f"    Reservations: {len(reservations)}")
    confirmed = sum(1 for r in reservations if r.status == "CONFIRMED")
    desisted = sum(1 for r in reservations if r.status == "DESISTED")
    print(f"      CONFIRMED: {confirmed}")
    print(f"      DESISTED:  {desisted}")
    print(f"    Freeze requests: {len(freezes)}")

    # Unique clients
    all_clients: set[str] = set()
    for r in reservations:
        for name in r.client_names:
            all_clients.add(name)
    print(f"    Unique clients: {len(all_clients)}")

    # Per-project breakdown
    for key in ["blt", "ben", "b5", "ce"]:
        count = sum(1 for r in reservations if r.project_key == key)
        frozen = sum(1 for f in freezes if f.project_key == key)
        print(f"    {PROJECT_NAMES[key]}: {count} reservations, {frozen} freezes")

    # Phase 4: Generate SQL
    print("\n--- Phase 4: Generate SQL ---")
    sql = generate_sql(reservations, freezes)

    out_path = ROOT / "scripts" / "backfill_reservations.sql"
    out_path.write_text(sql, encoding="utf-8")
    print(f"  SQL written to: {out_path}")
    print(f"  Lines: {len(sql.splitlines())}")
    print(f"  Size: {len(sql):,} bytes")
    print(f"\n  REVIEW the SQL file before executing!")


if __name__ == "__main__":
    main()

"""Extract the compliance officer's expediente database into JSON snapshots.

Reads CUMPLIMIENTO/Base de datos Apartamentos - *.xlsx (one sheet per project so
far: Boulevard 5). Emits:
  src/lib/cumplimiento/expedientes-snapshot.json  — aggregates (no PII)
  src/lib/cumplimiento/expedientes-rows.json      — per-apartment rows (client
    names, NO DPI numbers) — server-side import ONLY.

Headers are located BY CONTENT (row containing "No. APARTAMENTO"), never by
position. Buyer blocks (up to 4) are located from the "N NOMBRE DEL CLIENTE"
headers. DPI vencimiento dates before 2010 are flagged FECHA_ABSURDA (birth
dates typed into the wrong column — a real data-quality finding, not dropped).

Usage: python3 scripts/extract-cumplimiento-expedientes.py
"""
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SRC = "CUMPLIMIENTO/Base de datos Apartamentos - Boulevard 5.xlsx"
TODAY = date(2026, 8, 7)
OUT_AGG = "src/lib/cumplimiento/expedientes-snapshot.json"
OUT_ROWS = "src/lib/cumplimiento/expedientes-rows.json"

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb.worksheets[0]
rows = list(ws.iter_rows(values_only=True))

# Locate header row by content
header_i = next(i for i, r in enumerate(rows) if any(isinstance(v, str) and "No. APARTAMENTO" in v for v in r))
header = [str(v).strip() if v is not None else "" for v in rows[header_i]]

def col(name: str, after: int = 0) -> int | None:
    for j in range(after, len(header)):
        if header[j] == name:
            return j
    return None

APTO = col("No. APARTAMENTO")
MODELO = col("MODELO / TIPO")
NIVEL = col("NIVEL")
VENDEDOR = col("VENDEDOR")
PROMESA = col("FECHA FIRMA PROMESA")
BANCO_PRE = col("BANCO PRECALIFICACION")
FHA = col("FHA")
CONTADO = col("CONTADO")
FI_COLS = [c for c in (col("RELACIÓN DE DEPENDENCIA"), col("NEGOCIO PROPIO"), col("SERVICIOS PROFESIONALES"), col("OTROS")) if c is not None]
FI_LABELS = ["Relación de dependencia", "Negocio propio", "Servicios profesionales", "Otros"]
OBS_COLS = [j for j, h in enumerate(header) if h.startswith("OBSERVACIONES")]

# Buyer blocks: name col + the DPI/VENCIMIENTO/RTU cols that follow it
buyer_blocks = []
for j, h in enumerate(header):
    if re.match(r"^\d NOMBRE DEL CLIENTE$", h):
        buyer_blocks.append({"nombre": j, "venc": col("VENCIMIENTO", j), "rtu": next((k for k in range(j, min(j + 5, len(header))) if header[k].startswith("RTU")), None)})

def dpi_status(v) -> str:
    if isinstance(v, datetime):
        d = v.date()
        if d.year < 2010:
            return "FECHA_ABSURDA"
        return "VENCIDO" if d < TODAY else "VIGENTE"
    return "SIN_FECHA"

expedientes = []
agg = {"VIGENTE": 0, "VENCIDO": 0, "FECHA_ABSURDA": 0, "SIN_FECHA": 0}
con_promesa = con_fuente = con_banco = con_fha = con_contado = con_obs = total_compradores = rtu_poblado = 0

for r in rows[header_i + 1:]:
    if APTO is None or APTO >= len(r) or r[APTO] is None:
        continue
    clientes = []
    for b in buyer_blocks:
        nombre = r[b["nombre"]] if b["nombre"] < len(r) else None
        if nombre is None or str(nombre).strip() == "":
            continue
        status = dpi_status(r[b["venc"]] if b["venc"] is not None and b["venc"] < len(r) else None)
        rtu_v = r[b["rtu"]] if b["rtu"] is not None and b["rtu"] < len(r) else None
        rtu = rtu_v.date().isoformat() if isinstance(rtu_v, datetime) else None
        clientes.append({"nombre": str(nombre).strip(), "dpiStatus": status, "rtu": rtu})
        agg[status] += 1
        total_compradores += 1
        if rtu:
            rtu_poblado += 1
    promesa_v = r[PROMESA] if PROMESA is not None and PROMESA < len(r) else None
    promesa = promesa_v.date().isoformat() if isinstance(promesa_v, datetime) else None
    fuentes = [FI_LABELS[k] for k, c in enumerate(FI_COLS) if c < len(r) and r[c] is not None]
    banco = str(r[BANCO_PRE]).strip() if BANCO_PRE is not None and BANCO_PRE < len(r) and r[BANCO_PRE] is not None else None
    fha = FHA is not None and FHA < len(r) and r[FHA] is not None
    contado = CONTADO is not None and CONTADO < len(r) and r[CONTADO] is not None
    obs = [str(r[c]).strip() for c in OBS_COLS if c < len(r) and r[c] is not None and str(r[c]).strip()]
    if promesa:
        con_promesa += 1
    if fuentes:
        con_fuente += 1
    if banco:
        con_banco += 1
    if fha:
        con_fha += 1
    if contado:
        con_contado += 1
    if obs:
        con_obs += 1
    expedientes.append({
        "proyecto": "Boulevard 5",
        "apto": str(r[APTO]).strip(),
        "modelo": str(r[MODELO]).strip() if MODELO is not None and r[MODELO] is not None else None,
        "nivel": str(r[NIVEL]).strip() if NIVEL is not None and r[NIVEL] is not None else None,
        "vendedor": str(r[VENDEDOR]).strip() if VENDEDOR is not None and r[VENDEDOR] is not None else None,
        "clientes": clientes,
        "promesaFecha": promesa,
        "fuenteIngresos": fuentes,
        "bancoPrecalificacion": banco,
        "fha": fha,
        "contado": contado,
        "observaciones": obs,
    })

agg_out = {
    "_source": f"Extracted from {SRC} (base del oficial de cumplimiento). Regenerate with scripts/extract-cumplimiento-expedientes.py. Cutoff date for DPI status: {TODAY}.",
    "proyectos": ["Boulevard 5"],
    "totalExpedientes": len(expedientes),
    "totalCompradores": total_compradores,
    "dpi": agg,
    "rtuPoblado": rtu_poblado,
    "conPromesa": con_promesa,
    "conFuenteIngresos": con_fuente,
    "conBancoPrecalificacion": con_banco,
    "fha": con_fha,
    "contado": con_contado,
    "conObservaciones": con_obs,
    "notas": [
        "Solo Boulevard 5 — la data de los demás proyectos está completa en xlsx, pendiente de descarga (decisión de Jorge 2026-08-07: marcar completo con esta nota).",
        f"FECHA_ABSURDA = vencimiento de DPI anterior a 2010 ({agg['FECHA_ABSURDA']} compradores) — casi con certeza fechas de nacimiento tecleadas en la columna equivocada. El archivo fuente no puede responder confiablemente cuántos DPI están vencidos.",
        "Las observaciones del oficial se muestran por expediente pero NO son los 'casos específicos' (m3) — definición pendiente del equipo de cumplimiento.",
        "Números de DPI NO se extraen (dato sensible innecesario para monitoreo de status).",
    ],
}
Path("src/lib/cumplimiento").mkdir(parents=True, exist_ok=True)
json.dump(agg_out, open(OUT_AGG, "w"), ensure_ascii=False, indent=2)
json.dump(
    {"_source": agg_out["_source"] + " CONTAINS PII (client names) — server-side import only.", "expedientes": expedientes},
    open(OUT_ROWS, "w"), ensure_ascii=False, indent=2,
)
print(f"wrote {OUT_AGG} + {OUT_ROWS}: {len(expedientes)} expedientes, {total_compradores} compradores")
print("dpi:", agg, "| promesa:", con_promesa, "| fuente:", con_fuente, "| banco:", con_banco, "| obs:", con_obs)

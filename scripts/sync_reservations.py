#!/usr/bin/env python3
"""
Reservation Sync — Compare Reporte Excel files against Production DB.

Reads the 4 "Reporte" xlsx workbooks from Reservas/ at project root.
Extracts all Ventas tabs, queries the production DB for current state,
compares by (project, tower, unit_number), and outputs:
  1. SQL INSERT file for genuinely new reservations
  2. Markdown proposal for existing records with differences

Usage:
  python3 -u scripts/sync_reservations.py
"""

from __future__ import annotations

import base64
import json
import re
import subprocess
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
import requests

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
RESERVAS = ROOT / "Reservas"

REPORTE_FILES = {
    "blt": RESERVAS / "Bosque Las Tapias" / "ReporteBOSQUELASTAPIAS.xlsx",
    "ben": RESERVAS / "Benestare" / "ReporteBENESTARE.xlsx",
    "b5":  RESERVAS / "Boulevard 5" / "ReporteBOULEVARD5.xlsx",
    "ce":  RESERVAS / "Casa Elisa" / "ReporteCASAELISA.xlsx",
}

PROJECT_IDS = {
    "blt": "019c7d10-8ee5-7999-9881-2cd5ad038aa9",
    "ce":  "019c7d10-8e7b-7db6-93be-6f42d0538233",
    "b5":  "019c7d10-8e01-720f-942f-cac0017d83a8",
    "ben": "019c7d10-8f5a-74c7-b3df-c2151ad8a376",
}

PROJECT_NAMES = {
    "blt": "Bosque Las Tapias",
    "ce":  "Casa Elisa",
    "b5":  "Boulevard 5",
    "ben": "Benestare",
}

# Reverse lookup: project_name → project_key
PROJECT_KEY_BY_NAME = {v: k for k, v in PROJECT_NAMES.items()}

SUPABASE_REF = "nqaexbpteletuwdbpixq"

SINGLE_TOWER_PROJECTS = {"b5", "ce"}
DEFAULT_TOWER = {
    "b5": "Principal",
    "ce": "Principal",
}

# Salesperson canonical names
SALESPERSON_CANONICAL: dict[str, str] = {
    "paula hernandez":  "Paula Hernandez",
    "paula hernadez":   "Paula Hernandez",
    "paula hernande":   "Paula Hernandez",
    "jose gutierrez":   "Jose Gutierrez",
    "antonio rada":     "Antonio Rada",
    "brenda bucaro":    "Brenda Bucaro",
    "alexander franco": "Alexander Franco",
    "erwin cardona":    "Erwin Cardona",
    "anahi cisneros":   "Anahi Cisneros",
    "diana alvarez":    "Diana Alvarez",
    "sofia paredes":    "Sofia Paredes",
    "laura molina":     "Laura Molina",
    "efren sanchez":    "Efren Sanchez",
    "eder veliz":       "Eder Veliz",
    "pedro pablo sarti":"Pedro Pablo Sarti",
    "ronaldo ogaldez":  "Ronaldo Ogaldez",
    "pablo marroquin":  "Pablo Marroquin",
    "rony ramirez":     "Rony Ramirez",
    "noemi menendez":   "Noemi Menendez",
    "mario rodriguez":  "Mario Rodriguez",
    "forma capital":    "Forma Capital",
    "junta directiva":  "Junta Directiva",
    "alejandra calderon": "Alejandra Calderon",
    "andrea gonzalez":  "Andrea Gonzalez",
    "abigail garcia":   "Abigail Garcia",
    "ivan castillo":    "Ivan Castillo",
    "karina fuentes":   "Karina Fuentes",
    "luis esteban":     "Luis Esteban",
    "otto h.":          "Otto Herrera",
    "otto herrera":     "Otto Herrera",
}

SALESPERSON_EXCLUDE = {"traslado", "jd"}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class SaleRecord:
    """A single sale/reservation from a Ventas sheet."""
    project_key: str
    unit_number: str
    tower_name: str | None
    client_names: list[str]
    salesperson: str | None
    fecha: date | None
    enganche: float | None
    promesa_firmada: bool
    medio: str | None
    month_label: str


@dataclass
class DbRecord:
    """A reservation from the production database."""
    reservation_id: str
    status: str
    deposit_amount: float | None
    deposit_date: date | None
    lead_source: str | None
    is_resale: bool
    unit_number: str
    tower_name: str
    project_name: str
    project_key: str
    salesperson_name: str
    client_names: str | None  # "/" separated


@dataclass
class DiffField:
    """A single field difference between Excel and DB."""
    field_name: str
    excel_value: str
    db_value: str


@dataclass
class DiffRecord:
    """A matched record with field-level differences."""
    project_name: str
    tower_name: str
    unit_number: str
    db_reservation_id: str
    diffs: list[DiffField]


@dataclass
class EnrichmentRecord:
    """A DB record that can be enriched with Excel data."""
    project_name: str
    tower_name: str
    unit_number: str
    db_reservation_id: str
    field_name: str
    proposed_value: str


# ---------------------------------------------------------------------------
# Helpers (from backfill_reservations.py)
# ---------------------------------------------------------------------------

def strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_salesperson(raw: str | None) -> str | None:
    if raw is None:
        return None
    cleaned = str(raw).strip()
    if not cleaned:
        return None
    for sep in [" / ", "/", ","]:
        if sep in cleaned:
            cleaned = cleaned.split(sep)[0].strip()
            break
    if not cleaned:
        return None
    key = strip_accents(cleaned).lower()
    if key in SALESPERSON_EXCLUDE:
        return None
    return SALESPERSON_CANONICAL.get(key, cleaned)


def normalize_client_name(raw: str) -> str:
    name = raw.strip()
    name = name.strip(".,;:-")
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def split_client_names(raw: str | None) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    if " / " in text:
        parts = text.split(" / ")
    elif " - " in text and len(text.split(" - ")) == 2 and all(len(p.strip()) >= 5 for p in text.split(" - ")):
        parts = text.split(" - ")
    elif "/" in text and not re.search(r'\b\w\./', text):
        parts = text.split("/")
    else:
        parts = [text]
    result = []
    for p in parts:
        name = normalize_client_name(p)
        if name and len(name) >= 3:
            result.append(name)
    return result


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def safe_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = str(val).strip()
    if not text:
        return None
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def normalize_unit_number(raw) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if text.upper().startswith("L-"):
        return text.upper()
    n = safe_int(text)
    if n is not None and 100 <= n <= 99999:
        return str(n)
    if re.match(r'^[A-Z]-\d+$', text, re.IGNORECASE):
        return text.upper()
    return None


def normalize_tower(raw: str | None, project_key: str) -> str | None:
    if project_key in SINGLE_TOWER_PROJECTS:
        return DEFAULT_TOWER[project_key]
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if re.match(r'^[A-E]$', text, re.IGNORECASE):
        return f"Torre {text.upper()}"
    if re.match(r'^T[A-E]$', text, re.IGNORECASE):
        return f"Torre {text[1].upper()}"
    if text.lower().startswith("torre"):
        letter = text.split()[-1].upper() if len(text.split()) > 1 else text[-1].upper()
        return f"Torre {letter}"
    return text


def floor_from_unit(unit_number: str) -> int:
    if unit_number.upper().startswith("L-"):
        return 1
    n = safe_int(unit_number)
    if n is not None:
        return n // 100
    return 1


def sql_str(val: str | None) -> str:
    if val is None:
        return "NULL"
    escaped = val.replace("'", "''")
    return f"'{escaped}'"


def sql_num(val: float | int | None) -> str:
    if val is None:
        return "NULL"
    return str(val)


def sql_date(val: date | None) -> str:
    if val is None:
        return "NULL"
    return f"'{val.isoformat()}'"


def sql_bool(val: bool) -> str:
    return "true" if val else "false"


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

VENTAS_HEADER_MAP = {
    "fecha": "fecha",
    "cliente": "cliente",
    "no. apartamento": "unit",
    "no apartamento": "unit",
    "no. de apartamento": "unit",
    "torre": "torre",
    "asesor": "asesor",
    "enganche": "enganche",
    "promesa firmada": "promesa",
    "falta firma": "falta_firma",
    "medio": "medio",
    "precio de venta": "precio",
    "fecha de promesa": "fecha_promesa",
}


def detect_ventas_headers(ws, max_scan_rows: int = 5) -> tuple[int, dict[str, int]]:
    for row_num in range(1, max_scan_rows + 1):
        cols: dict[str, int] = {}
        for col in range(1, 30):
            val = cell(ws, row_num, col)
            if val is None:
                continue
            text = str(val).strip().lower()
            text = text.rstrip(". ")
            for pattern, field_name in VENTAS_HEADER_MAP.items():
                if text == pattern or text.startswith(pattern):
                    if field_name not in cols:
                        cols[field_name] = col
                    break
        if "cliente" in cols and ("unit" in cols or "asesor" in cols):
            return row_num, cols
    return 0, {}


# ---------------------------------------------------------------------------
# Phase 1: Extract from Excel
# ---------------------------------------------------------------------------

def extract_ventas_sheets(project_key: str, wb_path: Path) -> list[SaleRecord]:
    """Extract all sale records from monthly Ventas sheets."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    records: list[SaleRecord] = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.lower().startswith("ventas"):
            continue

        ws = wb[sheet_name]
        header_row, cols = detect_ventas_headers(ws)
        if header_row == 0:
            print(f"  WARN: No headers found in '{sheet_name}' — skipping")
            continue

        if "cliente" not in cols:
            print(f"  WARN: No 'Cliente' column in '{sheet_name}' — skipping")
            continue

        row_count = 0
        for row in range(header_row + 1, (ws.max_row or 0) + 1):
            client_raw = cell(ws, row, cols["cliente"])
            if client_raw is None or str(client_raw).strip() == "":
                continue

            unit_raw = cell(ws, row, cols["unit"]) if "unit" in cols else None
            unit_num = normalize_unit_number(unit_raw)
            if unit_num is None:
                continue

            torre_raw = cell(ws, row, cols["torre"]) if "torre" in cols else None
            torre = normalize_tower(str(torre_raw) if torre_raw else None, project_key)

            sp_raw = cell(ws, row, cols["asesor"]) if "asesor" in cols else None
            sp = normalize_salesperson(str(sp_raw) if sp_raw else None)

            fecha_raw = cell(ws, row, cols["fecha"]) if "fecha" in cols else None
            fecha = safe_date(fecha_raw)

            eng_raw = cell(ws, row, cols["enganche"]) if "enganche" in cols else None
            enganche = safe_float(eng_raw)

            promesa_raw = cell(ws, row, cols["promesa"]) if "promesa" in cols else None
            promesa = promesa_raw is not None and str(promesa_raw).strip().lower() not in ("", "0", "no", "false")

            medio_raw = cell(ws, row, cols["medio"]) if "medio" in cols else None
            medio = str(medio_raw).strip() if medio_raw and str(medio_raw).strip() else None

            clients = split_client_names(str(client_raw))
            if not clients:
                continue

            records.append(SaleRecord(
                project_key=project_key,
                unit_number=unit_num,
                tower_name=torre,
                client_names=clients,
                salesperson=sp,
                fecha=fecha,
                enganche=enganche,
                promesa_firmada=promesa,
                medio=medio,
                month_label=sheet_name,
            ))
            row_count += 1

        if row_count > 0:
            print(f"    {sheet_name}: {row_count} records")

    wb.close()
    return records


def build_unit_key(project_key: str, tower_name: str, unit_number: str) -> str:
    return f"{project_key}:{tower_name}:{unit_number}"


def deduplicate_ventas(all_ventas: dict[str, list[SaleRecord]]) -> dict[str, SaleRecord]:
    """Deduplicate: keep the most recent record per (project, tower, unit)."""
    by_key: dict[str, SaleRecord] = {}
    skipped_no_tower = 0

    for project_key, records in all_ventas.items():
        for v in records:
            tower = v.tower_name
            if tower is None:
                skipped_no_tower += 1
                continue

            key = build_unit_key(project_key, tower, v.unit_number)
            existing = by_key.get(key)
            if existing is None:
                by_key[key] = v
            elif v.fecha and (existing.fecha is None or v.fecha > existing.fecha):
                by_key[key] = v

    if skipped_no_tower > 0:
        print(f"  WARN: Skipped {skipped_no_tower} rows with no tower (multi-tower projects)")

    return by_key


# ---------------------------------------------------------------------------
# Phase 2: Query Production DB
# ---------------------------------------------------------------------------

def get_supabase_token() -> str:
    """Retrieve Supabase access token from macOS keychain or env var."""
    import os
    env_token = os.environ.get("SUPABASE_ACCESS_TOKEN")
    if env_token:
        return env_token

    try:
        result = subprocess.run(
            ["security", "find-generic-password", "-s", "Supabase CLI", "-w"],
            capture_output=True, text=True, timeout=5,
        )
        if result.returncode == 0:
            raw = result.stdout.strip()
            if raw.startswith("go-keyring-base64:"):
                raw = raw[len("go-keyring-base64:"):]
            return base64.b64decode(raw).decode("utf-8")
    except Exception as e:
        print(f"  WARN: Could not read keychain: {e}")

    print("ERROR: No Supabase access token found. Set SUPABASE_ACCESS_TOKEN env var.", file=sys.stderr)
    sys.exit(1)


def query_db(sql: str, token: str) -> list[dict]:
    """Execute a read-only SQL query via Supabase Management API."""
    resp = requests.post(
        f"https://api.supabase.com/v1/projects/{SUPABASE_REF}/database/query",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json={"query": sql},
        timeout=30,
    )
    if resp.status_code != 201:
        print(f"ERROR: DB query failed ({resp.status_code}): {resp.text}", file=sys.stderr)
        sys.exit(1)
    return resp.json()


def fetch_db_reservations(token: str) -> list[DbRecord]:
    """Fetch all reservations from production DB with full details."""
    sql = """
    SELECT
      r.id as reservation_id,
      r.status,
      r.deposit_amount,
      r.deposit_date,
      r.lead_source,
      r.is_resale,
      r.created_at,
      ru.unit_number,
      t.name as tower_name,
      p.name as project_name,
      p.id as project_id,
      sp.full_name as salesperson_name,
      (SELECT string_agg(rc2.full_name, ' / ' ORDER BY rc.is_primary DESC, rc2.full_name)
       FROM reservation_clients rc
       JOIN rv_clients rc2 ON rc.client_id = rc2.id
       WHERE rc.reservation_id = r.id) as client_names
    FROM reservations r
    JOIN rv_units ru ON r.unit_id = ru.id
    JOIN floors f ON ru.floor_id = f.id
    JOIN towers t ON f.tower_id = t.id
    JOIN projects p ON t.project_id = p.id
    JOIN salespeople sp ON r.salesperson_id = sp.id
    ORDER BY p.name, t.name, ru.unit_number;
    """
    rows = query_db(sql, token)
    records = []
    for row in rows:
        project_name = row["project_name"]
        project_key = PROJECT_KEY_BY_NAME.get(project_name)
        if project_key is None:
            continue

        deposit_date = None
        if row.get("deposit_date"):
            deposit_date = safe_date(row["deposit_date"])

        deposit_amount = None
        if row.get("deposit_amount") is not None:
            deposit_amount = safe_float(row["deposit_amount"])

        records.append(DbRecord(
            reservation_id=row["reservation_id"],
            status=row["status"],
            deposit_amount=deposit_amount,
            deposit_date=deposit_date,
            lead_source=row.get("lead_source"),
            is_resale=row.get("is_resale", False),
            unit_number=row["unit_number"],
            tower_name=row["tower_name"],
            project_name=project_name,
            project_key=project_key,
            salesperson_name=row["salesperson_name"],
            client_names=row.get("client_names"),
        ))
    return records


def build_db_map(db_records: list[DbRecord]) -> dict[str, list[DbRecord]]:
    """Build lookup: unit_key → list of DbRecords (may have desisted + confirmed)."""
    by_key: dict[str, list[DbRecord]] = defaultdict(list)
    for rec in db_records:
        key = build_unit_key(rec.project_key, rec.tower_name, rec.unit_number)
        by_key[key].append(rec)
    return dict(by_key)


# ---------------------------------------------------------------------------
# Phase 3: Compare
# ---------------------------------------------------------------------------

def fuzzy_eq(a: str | None, b: str | None) -> bool:
    """Case-insensitive, accent-insensitive string comparison."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return strip_accents(a).lower().strip() == strip_accents(b).lower().strip()


def fuzzy_client_eq(excel_clients: list[str], db_client_str: str | None) -> bool:
    """Compare client name lists (fuzzy)."""
    if db_client_str is None:
        return len(excel_clients) == 0
    db_clients = [c.strip() for c in db_client_str.split("/") if c.strip()]
    if len(excel_clients) != len(db_clients):
        return False
    excel_sorted = sorted(strip_accents(c).lower() for c in excel_clients)
    db_sorted = sorted(strip_accents(c).lower() for c in db_clients)
    return excel_sorted == db_sorted


def compare_records(
    excel_map: dict[str, SaleRecord],
    db_map: dict[str, list[DbRecord]],
) -> tuple[list[SaleRecord], list[DiffRecord], list[EnrichmentRecord], list[DbRecord]]:
    """Compare Excel records against DB records.

    Returns (new_records, diff_records, enrichments, db_only_records).
    """
    new_records: list[SaleRecord] = []
    diff_records: list[DiffRecord] = []
    enrichments: list[EnrichmentRecord] = []
    matched_keys: set[str] = set()

    for key, excel_rec in sorted(excel_map.items()):
        db_recs = db_map.get(key, [])

        # Find the active (CONFIRMED/PENDING_REVIEW) DB record
        active_db = None
        for db_rec in db_recs:
            if db_rec.status in ("CONFIRMED", "PENDING_REVIEW"):
                active_db = db_rec
                break

        if active_db is None:
            # Case A: NEW — no active reservation in DB
            new_records.append(excel_rec)
            continue

        matched_keys.add(key)

        # Case B: MATCH — compare fields
        diffs: list[DiffField] = []

        # Client names
        if not fuzzy_client_eq(excel_rec.client_names, active_db.client_names):
            diffs.append(DiffField(
                "Cliente",
                " / ".join(excel_rec.client_names),
                active_db.client_names or "(empty)",
            ))

        # Salesperson
        if excel_rec.salesperson and not fuzzy_eq(excel_rec.salesperson, active_db.salesperson_name):
            diffs.append(DiffField(
                "Asesor",
                excel_rec.salesperson or "(empty)",
                active_db.salesperson_name,
            ))

        # Date
        if excel_rec.fecha and active_db.deposit_date and excel_rec.fecha != active_db.deposit_date:
            diffs.append(DiffField(
                "Fecha",
                str(excel_rec.fecha),
                str(active_db.deposit_date),
            ))

        # Enganche (tolerance: Q1.00)
        if excel_rec.enganche is not None and active_db.deposit_amount is not None:
            if abs(excel_rec.enganche - active_db.deposit_amount) > 1.0:
                diffs.append(DiffField(
                    "Enganche",
                    f"Q{excel_rec.enganche:,.2f}",
                    f"Q{active_db.deposit_amount:,.2f}",
                ))

        # Lead source — only flag if BOTH have values and they differ
        if (excel_rec.medio and active_db.lead_source
                and not fuzzy_eq(excel_rec.medio, active_db.lead_source)):
            diffs.append(DiffField(
                "Medio/Lead Source",
                excel_rec.medio,
                active_db.lead_source,
            ))

        if diffs:
            diff_records.append(DiffRecord(
                project_name=PROJECT_NAMES[excel_rec.project_key],
                tower_name=excel_rec.tower_name or "",
                unit_number=excel_rec.unit_number,
                db_reservation_id=active_db.reservation_id,
                diffs=diffs,
            ))

        # Case C: ENRICHMENT — DB has NULL lead_source, Excel has value
        if excel_rec.medio and not active_db.lead_source:
            enrichments.append(EnrichmentRecord(
                project_name=PROJECT_NAMES[excel_rec.project_key],
                tower_name=excel_rec.tower_name or "",
                unit_number=excel_rec.unit_number,
                db_reservation_id=active_db.reservation_id,
                field_name="lead_source",
                proposed_value=excel_rec.medio,
            ))

    # Case D: DB-ONLY — active in DB but not in Excel
    db_only: list[DbRecord] = []
    for key, db_recs in sorted(db_map.items()):
        if key in matched_keys or key in excel_map:
            continue
        for db_rec in db_recs:
            if db_rec.status in ("CONFIRMED", "PENDING_REVIEW"):
                db_only.append(db_rec)

    return new_records, diff_records, enrichments, db_only


# ---------------------------------------------------------------------------
# Phase 4: Generate outputs
# ---------------------------------------------------------------------------

FALLBACK_DATE = "2025-09-01"


def generate_insert_sql(new_records: list[SaleRecord]) -> str:
    """Generate SQL INSERT statements for new reservations."""
    lines: list[str] = []
    lines.append("-- ==========================================================================")
    lines.append("-- Excel Sync — New Reservations")
    lines.append("-- Generated by scripts/sync_reservations.py")
    lines.append(f"-- Date: {date.today().isoformat()}")
    lines.append(f"-- New records: {len(new_records)}")
    lines.append("-- ==========================================================================")
    lines.append("")

    if not new_records:
        lines.append("-- No new records to insert.")
        return "\n".join(lines)

    lines.append("BEGIN;")
    lines.append("")

    # Collect unique client names
    all_client_names: set[str] = set()
    for r in new_records:
        for name in r.client_names:
            all_client_names.add(name)

    # Collect unique salesperson names
    all_sp_names: set[str] = set()
    for r in new_records:
        if r.salesperson:
            all_sp_names.add(r.salesperson)

    # Step 0: Ensure salespeople exist
    if all_sp_names:
        lines.append(f"-- Step 0: Ensure salespeople exist ({len(all_sp_names)})")
        for sp in sorted(all_sp_names):
            display = sp.split()[0] if " " in sp else sp
            lines.append(
                f"INSERT INTO salespeople (full_name, display_name) "
                f"VALUES ({sql_str(sp)}, {sql_str(display)}) "
                f"ON CONFLICT (full_name) DO NOTHING;"
            )
        lines.append("")

    # Step 1: Insert clients (skip if name already exists — no unique constraint on full_name)
    lines.append(f"-- Step 1: Clients ({len(all_client_names)} names, skip existing)")
    for name in sorted(all_client_names):
        lines.append(
            f"INSERT INTO rv_clients (full_name) "
            f"SELECT {sql_str(name)} "
            f"WHERE NOT EXISTS (SELECT 1 FROM rv_clients WHERE full_name = {sql_str(name)});"
        )
    lines.append("")

    # Step 2: Insert reservations
    lines.append(f"-- Step 2: Reservations ({len(new_records)} rows)")
    for r in new_records:
        pid = PROJECT_IDS[r.project_key]
        fecha_str = sql_date(r.fecha) if r.fecha else f"'{FALLBACK_DATE}'"

        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(r.tower_name)} "
            f"AND ru.unit_number = {sql_str(r.unit_number)})"
        )
        sp_subq = f"(SELECT id FROM salespeople WHERE full_name = {sql_str(r.salesperson)} LIMIT 1)"

        cols = [
            "unit_id", "salesperson_id", "status",
            "deposit_amount", "deposit_date",
            "lead_source", "is_resale",
            "created_at", "reviewed_at",
        ]
        vals = [
            unit_subq, sp_subq, "'CONFIRMED'::rv_reservation_status",
            sql_num(r.enganche), fecha_str,
            sql_str(r.medio), "false",
            fecha_str, fecha_str,
        ]

        lines.append(
            f"INSERT INTO reservations ({', '.join(cols)}) VALUES ({', '.join(vals)});"
        )
    lines.append("")

    # Step 3: Reservation ↔ Client links
    lines.append(f"-- Step 3: Reservation ↔ Client links")
    for r in new_records:
        pid = PROJECT_IDS[r.project_key]
        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(r.tower_name)} "
            f"AND ru.unit_number = {sql_str(r.unit_number)})"
        )
        res_subq = (
            f"(SELECT id FROM reservations "
            f"WHERE unit_id = {unit_subq} "
            f"AND status = 'CONFIRMED' "
            f"AND is_resale = false "
            f"ORDER BY created_at DESC LIMIT 1)"
        )

        for j, name in enumerate(r.client_names):
            client_subq = f"(SELECT id FROM rv_clients WHERE full_name = {sql_str(name)} LIMIT 1)"
            is_primary = j == 0
            lines.append(
                f"INSERT INTO reservation_clients (reservation_id, client_id, is_primary) "
                f"VALUES ({res_subq}, {client_subq}, {sql_bool(is_primary)});"
            )
    lines.append("")

    # Step 4: Unit status log
    lines.append(f"-- Step 4: Unit status log")
    for r in new_records:
        pid = PROJECT_IDS[r.project_key]
        fecha_str = r.fecha.isoformat() if r.fecha else FALLBACK_DATE
        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(r.tower_name)} "
            f"AND ru.unit_number = {sql_str(r.unit_number)})"
        )
        lines.append(
            f"INSERT INTO unit_status_log (unit_id, old_status, new_status, changed_by, reason, created_at) "
            f"VALUES ({unit_subq}, NULL, 'RESERVED'::rv_unit_status, "
            f"'system:excel-sync-{date.today().isoformat()}', 'Sync desde Reporte Excel', '{fecha_str}');"
        )
    lines.append("")

    # Step 5: Update rv_units status to RESERVED
    lines.append(f"-- Step 5: Update unit status to RESERVED")
    for r in new_records:
        pid = PROJECT_IDS[r.project_key]
        unit_subq = (
            f"(SELECT ru.id FROM rv_units ru "
            f"JOIN floors f ON f.id = ru.floor_id "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(r.tower_name)} "
            f"AND ru.unit_number = {sql_str(r.unit_number)})"
        )
        lines.append(
            f"UPDATE rv_units SET status = 'RESERVED'::rv_unit_status, "
            f"status_changed_at = now() "
            f"WHERE id = {unit_subq} AND status = 'AVAILABLE';"
        )
    lines.append("")

    lines.append("COMMIT;")
    lines.append("")
    return "\n".join(lines)


def generate_diff_markdown(
    new_records: list[SaleRecord],
    diff_records: list[DiffRecord],
    enrichments: list[EnrichmentRecord],
    db_only: list[DbRecord],
    excel_total: int,
    db_total: int,
    matched_no_diff: int,
) -> str:
    """Generate Markdown proposal document."""
    lines: list[str] = []

    lines.append("# Sync Diff Proposal — Reservas vs Production DB")
    lines.append(f"\n**Generated:** {date.today().isoformat()}")
    lines.append(f"**Script:** `scripts/sync_reservations.py`\n")

    # Summary
    lines.append("## Summary\n")
    lines.append(f"| Metric | Count |")
    lines.append(f"|---|---|")
    lines.append(f"| Excel records (deduplicated) | {excel_total} |")
    lines.append(f"| DB active reservations (CONFIRMED/PENDING_REVIEW) | {db_total} |")
    lines.append(f"| Matched (no differences) | {matched_no_diff} |")
    lines.append(f"| Matched (with differences) | {len(diff_records)} |")
    lines.append(f"| **New records** (auto-inserted) | **{len(new_records)}** |")
    lines.append(f"| Lead source enrichments | {len(enrichments)} |")
    lines.append(f"| DB-only (not in Excel) | {len(db_only)} |")
    lines.append("")

    # New records
    if new_records:
        lines.append("## New Records (Auto-Inserted)\n")
        lines.append("| # | Project | Tower | Unit | Client | Asesor | Fecha | Enganche | Medio |")
        lines.append("|---|---------|-------|------|--------|--------|-------|----------|-------|")
        for i, r in enumerate(new_records, 1):
            lines.append(
                f"| {i} | {PROJECT_NAMES[r.project_key]} | {r.tower_name or '-'} "
                f"| {r.unit_number} | {' / '.join(r.client_names)} "
                f"| {r.salesperson or '-'} | {r.fecha or '-'} "
                f"| {f'Q{r.enganche:,.2f}' if r.enganche else '-'} "
                f"| {r.medio or '-'} |"
            )
        lines.append("")

    # Differences
    if diff_records:
        lines.append("## Differences Found (Proposal for Review)\n")
        lines.append("These records exist in both Excel and DB but have field-level differences.\n")
        for i, dr in enumerate(diff_records, 1):
            lines.append(f"### {i}. {dr.project_name} — {dr.tower_name} — Unit {dr.unit_number}")
            lines.append(f"DB Reservation ID: `{dr.db_reservation_id}`\n")
            lines.append("| Field | Excel | DB |")
            lines.append("|-------|-------|----|")
            for d in dr.diffs:
                lines.append(f"| {d.field_name} | {d.excel_value} | {d.db_value} |")
            lines.append("")
    else:
        lines.append("## Differences Found\n")
        lines.append("None — all matched records are identical.\n")

    # Enrichments
    if enrichments:
        lines.append("## Lead Source Enrichments\n")
        lines.append("DB records with NULL `lead_source` that can be populated from Excel `Medio`.\n")
        lines.append("```sql")
        lines.append("-- Proposed UPDATE statements for lead_source enrichment")
        for e in enrichments:
            lines.append(
                f"UPDATE reservations SET lead_source = {sql_str(e.proposed_value)} "
                f"WHERE id = '{e.db_reservation_id}'; "
                f"-- {e.project_name} {e.tower_name} Unit {e.unit_number}"
            )
        lines.append("```\n")

    # DB-only
    if db_only:
        lines.append("## DB-Only Records (Not in Latest Excel)\n")
        lines.append("These active reservations exist in DB but have no matching row in the latest Ventas tabs.")
        lines.append("This is expected for app-created reservations or data from months not covered.\n")
        lines.append("| Project | Tower | Unit | Client | Asesor | Date | Status |")
        lines.append("|---------|-------|------|--------|--------|------|--------|")
        for rec in db_only:
            lines.append(
                f"| {rec.project_name} | {rec.tower_name} "
                f"| {rec.unit_number} | {rec.client_names or '-'} "
                f"| {rec.salesperson_name} | {rec.deposit_date or '-'} "
                f"| {rec.status} |"
            )
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=== Reservation Sync: Excel → Production DB ===\n")

    # Verify files exist
    print("Checking Reporte files...")
    for key, path in REPORTE_FILES.items():
        if not path.exists():
            print(f"  ERROR: Not found: {path}", file=sys.stderr)
            sys.exit(1)
        print(f"  Found: {path.name}")

    # Phase 1: Extract from Excel
    print("\n--- Phase 1: Extract sales records from Reporte files ---")
    all_ventas: dict[str, list[SaleRecord]] = {}
    for key in ["blt", "ben", "b5", "ce"]:
        print(f"\n  {PROJECT_NAMES[key]}:")
        all_ventas[key] = extract_ventas_sheets(key, REPORTE_FILES[key])
        print(f"  Total: {len(all_ventas[key])} records")

    total_raw = sum(len(v) for v in all_ventas.values())
    print(f"\n  Grand total (raw): {total_raw}")

    # Deduplicate
    print("\n  Deduplicating (most recent per unit)...")
    excel_map = deduplicate_ventas(all_ventas)
    print(f"  Deduplicated: {len(excel_map)} unique unit-reservations")

    # Phase 2: Query production DB
    print("\n--- Phase 2: Query production DB ---")
    token = get_supabase_token()
    print("  Token acquired.")

    db_records = fetch_db_reservations(token)
    print(f"  Fetched {len(db_records)} total DB reservations")

    db_map = build_db_map(db_records)
    active_count = sum(
        1 for recs in db_map.values()
        for r in recs if r.status in ("CONFIRMED", "PENDING_REVIEW")
    )
    print(f"  Active (CONFIRMED/PENDING_REVIEW): {active_count}")

    # Phase 3: Compare
    print("\n--- Phase 3: Compare ---")
    new_records, diff_records, enrichments, db_only = compare_records(excel_map, db_map)

    matched_no_diff = len(excel_map) - len(new_records) - len(diff_records)
    print(f"  NEW (not in DB):        {len(new_records)}")
    print(f"  MATCH (no diff):        {matched_no_diff}")
    print(f"  MATCH (with diff):      {len(diff_records)}")
    print(f"  ENRICHMENTS:            {len(enrichments)}")
    print(f"  DB-ONLY (not in Excel): {len(db_only)}")

    # Phase 4: Generate outputs
    print("\n--- Phase 4: Generate outputs ---")

    # SQL for new records
    sql = generate_insert_sql(new_records)
    sql_path = ROOT / "scripts" / "sync_new_reservations.sql"
    sql_path.write_text(sql, encoding="utf-8")
    print(f"  SQL: {sql_path} ({len(sql.splitlines())} lines)")

    # Markdown diff proposal
    md = generate_diff_markdown(
        new_records, diff_records, enrichments, db_only,
        excel_total=len(excel_map),
        db_total=active_count,
        matched_no_diff=matched_no_diff,
    )
    md_path = ROOT / "docs" / f"sync-diff-proposal-{date.today().isoformat()}.md"
    md_path.write_text(md, encoding="utf-8")
    print(f"  Proposal: {md_path}")

    print("\n=== Done ===")
    if new_records:
        print(f"  Review {sql_path.name} then execute via Management API.")
    print(f"  Review {md_path.name} for differences and enrichments.")


if __name__ == "__main__":
    main()

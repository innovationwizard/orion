#!/usr/bin/env python3
"""
Production Database Seed — ETL from SSOT xlsx files.

Reads the 4 SSOT xlsx workbooks and generates a single idempotent SQL file
that populates towers, floors, rv_units, and salespeople.

Safety:
  - All INSERTs use ON CONFLICT DO NOTHING
  - Wrapped in a single transaction (BEGIN / COMMIT)
  - No DELETEs or destructive UPDATEs
  - Only UPDATEs are on projects table (name/slug fix)
  - Review seed_prod.sql before executing

Usage:
  python3 scripts/seed_prod.py
  # Review scripts/seed_prod.sql
  # Then execute via Supabase Management API
"""

from __future__ import annotations

import os
import sys
import unicodedata
from pathlib import Path
from typing import NamedTuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
SSOT = ROOT / "origin" / "SSOT" / "Reservas y Ventas"

FILES = {
    "blt": SSOT / "Bosque Las Tapias" / "Disponibilidad y cotizador Bosque Las Tapias septiembre 2025.xlsx",
    "ben": SSOT / "Benestare" / "Precios y Disponibilidad BENESTARE TA septiembre 25.xlsx",
    "b5":  SSOT / "Boulevard 5" / "DisponibilIdad e Integración B5 septiembre 2025.xlsx",
    "ce":  SSOT / "Casa Elisa" / "CASA ELISA con trazabilidad precios septiembre 2025.xlsx",
}

PROJECT_IDS = {
    "blt": "019c7d10-8ee5-7999-9881-2cd5ad038aa9",
    "ce":  "019c7d10-8e7b-7db6-93be-6f42d0538233",
    "b5":  "019c7d10-8e01-720f-942f-cac0017d83a8",
    "ben": "019c7d10-8f5a-74c7-b3df-c2151ad8a376",
}

PROJECT_NAMES = {
    "blt": "Bosque Las Tapias",
    "ce":  "Casa Elisa",
    "b5":  "Boulevard 5",
    "ben": "Benestare",
}

PROJECT_SLUGS = {
    "blt": "bosque-las-tapias",
    "ce":  "casa-elisa",
    "b5":  "boulevard-5",
    "ben": "benestare",
}

STATUS_MAP: dict[str, str] = {
    "disponible": "AVAILABLE",
    "reservado":  "RESERVED",
    "pcv":        "SOLD",
    "promesa":    "SOLD",
    "congelado":  "FROZEN",
    "congelado junta directiva": "FROZEN",
    "liberado":   "AVAILABLE",
}

# Canonical salesperson names — maps (accent-stripped lowercase) to canonical form
SALESPERSON_CANONICAL: dict[str, str] = {
    "paula hernandez":  "Paula Hernandez",
    "paula hernadez":   "Paula Hernandez",
    "paula hernande":   "Paula Hernandez",
    "jose gutierrez":   "Jose Gutierrez",
    "antonio rada":     "Antonio Rada",
    "brenda bucaro":    "Brenda Bucaro",
    "alexander franco": "Alexander Franco",
    "erwin cardona":    "Erwin Cardona",
    "anahi cisneros":   "Anahi Cisneros",
    "diana alvarez":    "Diana Alvarez",
    "sofia paredes":    "Sofia Paredes",
    "laura molina":     "Laura Molina",
    "efren sanchez":    "Efren Sanchez",
    "eder veliz":       "Eder Veliz",
    "pedro pablo sarti":"Pedro Pablo Sarti",
    "ronaldo ogaldez":  "Ronaldo Ogaldez",
    "pablo marroquin":  "Pablo Marroquin",
    "rony ramirez":     "Rony Ramirez",
    "noemi menendez":   "Noemi Menendez",
    "mario rodriguez":  "Mario Rodriguez",
    "forma capital":    "Forma Capital",
    "junta directiva":  "Junta Directiva",
    "alejandra calderon": "Alejandra Calderon",
    "andrea gonzalez":  "Andrea Gonzalez",
    "abigail garcia":   "Abigail Garcia",
    "ivan castillo":    "Ivan Castillo",
    "karina fuentes":   "Karina Fuentes",
    "luis esteban":     "Luis Esteban",
    "otto h.":          "Otto Herrera",
    "otto herrera":     "Otto Herrera",
}

# Names that are not real salespeople — filter these out
SALESPERSON_EXCLUDE = {
    "traslado",
    "jd",
}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

class Tower(NamedTuple):
    project_key: str
    name: str
    is_default: bool


class Floor(NamedTuple):
    project_key: str
    tower_name: str
    number: int


class Unit(NamedTuple):
    project_key: str
    tower_name: str
    floor_number: int
    unit_number: str
    unit_code: str | None
    unit_type: str
    bedrooms: int
    is_local: bool
    area_interior: float | None
    area_balcony: float | None
    area_terrace: float | None
    area_garden: float | None
    area_total: float | None
    parking_car: int
    parking_tandem: int
    parking_moto: int
    parking_type: str | None
    parking_number: str | None
    parking_level: str | None
    bodega_number: str | None
    bodega_area: float | None
    price_list: float | None
    status: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sql_str(val: str | None) -> str:
    """Escape a string for SQL. Returns NULL for None."""
    if val is None:
        return "NULL"
    escaped = val.replace("'", "''")
    return f"'{escaped}'"


def sql_num(val: float | int | None) -> str:
    """Format a number for SQL. Returns NULL for None."""
    if val is None:
        return "NULL"
    return str(val)


def sql_bool(val: bool) -> str:
    return "true" if val else "false"


def normalize_status(raw: str | None) -> str:
    """Map xlsx status string to rv_unit_status enum value."""
    if raw is None:
        return "AVAILABLE"
    key = str(raw).strip().lower()
    # Handle numeric values (data errors in B5) — treat as SOLD
    try:
        float(key)
        return "SOLD"
    except ValueError:
        pass
    return STATUS_MAP.get(key, "AVAILABLE")


def strip_accents(s: str) -> str:
    """Remove diacritical marks (accents) from a string."""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_salesperson(raw: str | None) -> str | None:
    """Normalize salesperson name to canonical form."""
    if raw is None:
        return None
    cleaned = str(raw).strip()
    if not cleaned:
        return None
    # Strip suffixes: "Name / Traslado", "Name/Traslado", "Name, note"
    for sep in [" / ", "/", ","]:
        if sep in cleaned:
            cleaned = cleaned.split(sep)[0].strip()
            break
    if not cleaned:
        return None
    # Strip accents and lowercase for lookup
    key = strip_accents(cleaned).lower()
    # Exclude garbage entries
    if key in SALESPERSON_EXCLUDE:
        return None
    return SALESPERSON_CANONICAL.get(key, cleaned)


def safe_int(val) -> int | None:
    """Safely convert to int."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    """Safely convert to float."""
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def cell(ws: Worksheet, row: int, col: int):
    """Get cell value at (row, col) — 1-indexed."""
    return ws.cell(row=row, column=col).value


# ---------------------------------------------------------------------------
# Per-project extraction functions
# ---------------------------------------------------------------------------

def extract_blt(wb_path: Path) -> tuple[list[Tower], list[Floor], list[Unit], set[str]]:
    """Extract Bosque Las Tapias data from xlsx."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    towers: list[Tower] = []
    floors: list[Floor] = []
    units: list[Unit] = []
    salespeople: set[str] = set()

    for sheet_name, tower_name in [("Precios Torre C", "Torre C"), ("Precios Torre B", "Torre B")]:
        ws = wb[sheet_name]
        towers.append(Tower("blt", tower_name, False))

        is_torre_c = tower_name == "Torre C"

        seen_floors: set[int] = set()
        for row in range(3, ws.max_row + 1):
            # Unit number column is always B (col 2)
            raw_unit = cell(ws, row, 2)
            unit_num = safe_int(raw_unit)
            if unit_num is None or unit_num < 100:
                continue

            floor_num = unit_num // 100
            if floor_num not in seen_floors:
                seen_floors.add(floor_num)
                floors.append(Floor("blt", tower_name, floor_num))

            if is_torre_c:
                # Torre C layout
                unit_type = str(cell(ws, row, 3) or "").strip()  # C = Tipo
                bedrooms_raw = safe_int(cell(ws, row, 4))        # D = Habitaciones
                area_int = safe_float(cell(ws, row, 6))          # F = interior
                area_bal = safe_float(cell(ws, row, 7))          # G = balcony
                area_ter = safe_float(cell(ws, row, 8))          # H = terrace
                area_tot = safe_float(cell(ws, row, 9))          # I = total
                park_car = safe_int(cell(ws, row, 11)) or 0      # K = parqueo carro
                park_tan = safe_int(cell(ws, row, 12)) or 0      # L = parqueo tandem
                park_mot = safe_int(cell(ws, row, 13)) or 0      # M = carro+pq moto
                price = safe_float(cell(ws, row, 18))            # R = Aproximacion FHA
                status_raw = cell(ws, row, 27)                   # AA = Estatus
                sales_raw = cell(ws, row, 29)                    # AC = Asesor
                bedrooms = bedrooms_raw if bedrooms_raw is not None else (2 if unit_type.startswith("A") else 3)
            else:
                # Torre B layout — shifted columns, no Habitaciones column
                unit_type = str(cell(ws, row, 3) or "").strip()  # C = Tipo
                area_int = safe_float(cell(ws, row, 5))          # E = interior
                area_bal = None                                   # No balcony in Torre B
                area_ter = None                                   # No terrace in Torre B
                area_tot = safe_float(cell(ws, row, 8))          # H = total
                park_car = safe_int(cell(ws, row, 10)) or 0      # J = parqueo carro
                park_tan = safe_int(cell(ws, row, 11)) or 0      # K = parqueo tandem
                park_mot = safe_int(cell(ws, row, 12)) or 0      # L = carro+pq moto
                price = safe_float(cell(ws, row, 16))            # P = Aproximacion FHA
                status_raw = cell(ws, row, 24)                   # X = Estatus
                sales_raw = cell(ws, row, 26)                    # Z = Asesor
                # Derive bedrooms from type: A=2, B=3, C=3
                bedrooms = 2 if unit_type.startswith("A") else 3

            if not unit_type:
                continue

            status = normalize_status(str(status_raw) if status_raw else None)
            sp = normalize_salesperson(str(sales_raw) if sales_raw else None)
            if sp:
                salespeople.add(sp)

            units.append(Unit(
                project_key="blt",
                tower_name=tower_name,
                floor_number=floor_num,
                unit_number=str(unit_num),
                unit_code=None,
                unit_type=unit_type,
                bedrooms=bedrooms,
                is_local=False,
                area_interior=area_int,
                area_balcony=area_bal,
                area_terrace=area_ter,
                area_garden=None,
                area_total=area_tot,
                parking_car=park_car,
                parking_tandem=park_tan,
                parking_moto=park_mot,
                parking_type=None,
                parking_number=None,
                parking_level=None,
                bodega_number=None,
                bodega_area=None,
                price_list=price,
                status=status,
            ))

    wb.close()
    return towers, floors, units, salespeople


def extract_benestare(wb_path: Path) -> tuple[list[Tower], list[Floor], list[Unit], set[str]]:
    """Extract Benestare data from xlsx."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Precios"]

    towers: list[Tower] = []
    floors: list[Floor] = []
    units: list[Unit] = []
    salespeople: set[str] = set()

    seen_towers: set[str] = set()
    seen_floors: set[tuple[str, int]] = set()

    for row in range(4, ws.max_row + 1):
        # D = unit number, K = tower letter
        raw_unit = cell(ws, row, 4)   # D = NUMERO
        unit_num = safe_int(raw_unit)
        if unit_num is None or unit_num < 100:
            continue

        tower_letter = str(cell(ws, row, 11) or "").strip()  # K = Torre
        if not tower_letter or len(tower_letter) > 2:
            continue

        tower_name = f"Torre {tower_letter}"
        floor_num = safe_int(cell(ws, row, 3))  # C = Nivel
        if floor_num is None:
            floor_num = unit_num // 100

        if tower_name not in seen_towers:
            seen_towers.add(tower_name)
            towers.append(Tower("ben", tower_name, False))

        floor_key = (tower_name, floor_num)
        if floor_key not in seen_floors:
            seen_floors.add(floor_key)
            floors.append(Floor("ben", tower_name, floor_num))

        unit_type = str(cell(ws, row, 5) or "").strip()  # E = TIPO
        unit_code = str(cell(ws, row, 6) or "").strip() or None  # F = ID (e.g. C601)
        area_total = safe_float(cell(ws, row, 7))  # G = MTS2
        parking_number_raw = cell(ws, row, 8)  # H = PARQUEO (spot number)
        parking_level_raw = str(cell(ws, row, 9) or "").strip() or None  # I = NIVEL (N1, N2, etc.)
        bedrooms = safe_int(cell(ws, row, 10)) or 0  # J = HABITACIONES
        price = safe_float(cell(ws, row, 23))  # W = Precio de Venta
        status_raw = cell(ws, row, 41)  # AO = Estatus
        sales_raw = cell(ws, row, 44)   # AR = Asesor

        if not unit_type:
            continue

        status = normalize_status(str(status_raw) if status_raw else None)
        sp = normalize_salesperson(str(sales_raw) if sales_raw else None)
        if sp:
            salespeople.add(sp)

        parking_num_str = str(int(parking_number_raw)) if safe_int(parking_number_raw) is not None else None

        units.append(Unit(
            project_key="ben",
            tower_name=tower_name,
            floor_number=floor_num,
            unit_number=str(unit_num),
            unit_code=unit_code,
            unit_type=unit_type,
            bedrooms=bedrooms,
            is_local=False,
            area_interior=None,
            area_balcony=None,
            area_terrace=None,
            area_garden=None,
            area_total=area_total,
            parking_car=1,  # Every Benestare unit gets 1 parking
            parking_tandem=0,
            parking_moto=0,
            parking_type=None,
            parking_number=parking_num_str,
            parking_level=parking_level_raw,
            bodega_number=None,
            bodega_area=None,
            price_list=price,
            status=status,
        ))

    wb.close()
    return towers, floors, units, salespeople


def extract_b5(wb_path: Path) -> tuple[list[Tower], list[Floor], list[Unit], set[str]]:
    """Extract Boulevard 5 data from xlsx."""
    # B5 file has invalid filter XML — use read_only mode
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Matriz Precios A"]

    towers = [Tower("b5", "Principal", True)]
    floors: list[Floor] = []
    units: list[Unit] = []
    salespeople: set[str] = set()

    seen_floors: set[int] = set()

    for row in range(4, ws.max_row + 1):
        raw_unit = cell(ws, row, 2)  # B = Numero
        unit_num = safe_int(raw_unit)
        if unit_num is None or unit_num < 100:
            continue

        floor_num = unit_num // 100
        if floor_num not in seen_floors:
            seen_floors.add(floor_num)
            floors.append(Floor("b5", "Principal", floor_num))

        unit_type = str(cell(ws, row, 3) or "").strip()  # C = Tipo
        area_int = safe_float(cell(ws, row, 4))           # D = M2 (interior)
        bedrooms = safe_int(cell(ws, row, 5)) or 0        # E = Habitaciones
        bodega_count = safe_int(cell(ws, row, 6))         # F = No. Bodega 5mts
        area_bal = safe_float(cell(ws, row, 7))           # G = Terraza-Balcon
        park_car = safe_int(cell(ws, row, 8)) or 0        # H = Parqueos
        park_tan = safe_int(cell(ws, row, 9)) or 0        # I = Parqueo Tandem

        # Parking info from column O (new location) — format: S5_229
        parking_raw = str(cell(ws, row, 15) or "").strip()  # O = Sotano y parqueo NUEVA UBICACION
        parking_level = None
        parking_number = None
        if parking_raw and "_" in parking_raw:
            parts = parking_raw.split("_", 1)
            parking_level = parts[0].strip()
            parking_number = parts[1].strip()
        elif parking_raw:
            parking_level = parking_raw

        # Bodega from column K (new location)
        bodega_new = cell(ws, row, 11)  # K = Bodega NUEVA UBICACION
        bodega_number = str(int(bodega_new)) if safe_int(bodega_new) is not None else None
        bodega_area = 5.0 if (bodega_count and bodega_count > 0) else None

        price = safe_float(cell(ws, row, 41))     # AO = Precio Final REDONDEADO
        status_raw = cell(ws, row, 58)             # BF = Estatus
        sales_raw = cell(ws, row, 60)              # BH = Asesor

        if not unit_type:
            continue

        # Compute area_total
        area_tot = None
        if area_int is not None:
            area_tot = area_int + (area_bal or 0)

        status = normalize_status(str(status_raw) if status_raw else None)
        sp = normalize_salesperson(str(sales_raw) if sales_raw else None)
        if sp:
            salespeople.add(sp)

        units.append(Unit(
            project_key="b5",
            tower_name="Principal",
            floor_number=floor_num,
            unit_number=str(unit_num),
            unit_code=None,
            unit_type=unit_type,
            bedrooms=bedrooms,
            is_local=False,
            area_interior=area_int,
            area_balcony=area_bal,
            area_terrace=None,
            area_garden=None,
            area_total=area_tot,
            parking_car=park_car,
            parking_tandem=park_tan,
            parking_moto=0,
            parking_type=None,
            parking_number=parking_number,
            parking_level=parking_level,
            bodega_number=bodega_number,
            bodega_area=bodega_area,
            price_list=price,
            status=status,
        ))

    wb.close()
    return towers, floors, units, salespeople


def extract_ce(wb_path: Path) -> tuple[list[Tower], list[Floor], list[Unit], set[str]]:
    """Extract Casa Elisa data from xlsx."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Disponibilidad"]

    towers = [Tower("ce", "Principal", True)]
    floors: list[Floor] = []
    units: list[Unit] = []
    salespeople: set[str] = set()

    seen_floors: set[int] = set()

    for row in range(5, ws.max_row + 1):
        raw_unit = cell(ws, row, 2)  # B = Numero
        unit_num_int = safe_int(raw_unit)
        unit_num_str = str(raw_unit).strip() if raw_unit else ""

        # Accept integer unit numbers (101, 202, ...) OR locale text IDs (L-1, L-2, L-3)
        is_locale_id = unit_num_str.startswith("L-")
        if unit_num_int is None and not is_locale_id:
            continue
        if unit_num_int is not None and unit_num_int < 100:
            continue

        if is_locale_id:
            floor_num = 1  # Locales are on floor 1
            unit_number_text = unit_num_str
        else:
            floor_num = unit_num_int // 100
            unit_number_text = str(unit_num_int)

        if floor_num not in seen_floors:
            seen_floors.add(floor_num)
            floors.append(Floor("ce", "Principal", floor_num))

        unit_type = str(cell(ws, row, 3) or "").strip()   # C = Tipo
        area_int = safe_float(cell(ws, row, 5))            # E = M2
        bedrooms = safe_int(cell(ws, row, 6)) or 0         # F = Habitaciones
        bodega_raw = cell(ws, row, 7)                      # G = No. Bodega
        area_bal = safe_float(cell(ws, row, 8))            # H = Balcon m2
        area_ter = safe_float(cell(ws, row, 9))            # I = Terraza
        park_car = safe_int(cell(ws, row, 10)) or 0        # J = Parqueos
        parking_level_raw = str(cell(ws, row, 11) or "").strip() or None  # K = Sotano
        parking_number_raw = cell(ws, row, 13)             # M = No. Parqueo Actual (Nuevo)
        parking_type_raw = str(cell(ws, row, 14) or "").strip()  # N = Tipo de Parqueo
        price = safe_float(cell(ws, row, 39))              # AM = Precio FINAL
        status_raw = cell(ws, row, 49)                     # AW = Estatus
        sales_raw = cell(ws, row, 51)                      # AY = Vendedor

        if not unit_type:
            continue

        is_local = unit_type.lower() == "local"
        if is_local:
            bedrooms = 0

        # Compute area_total
        area_tot = None
        if area_int is not None:
            area_tot = area_int + (area_bal or 0) + (area_ter or 0)

        bodega_number = str(int(bodega_raw)) if safe_int(bodega_raw) is not None else None
        parking_number = str(int(parking_number_raw)) if safe_int(parking_number_raw) is not None else None
        parking_type = parking_type_raw if parking_type_raw and parking_type_raw != "0" else None

        status = normalize_status(str(status_raw) if status_raw else None)
        sp = normalize_salesperson(str(sales_raw) if sales_raw else None)
        if sp:
            salespeople.add(sp)

        units.append(Unit(
            project_key="ce",
            tower_name="Principal",
            floor_number=floor_num,
            unit_number=unit_number_text,
            unit_code=None,
            unit_type=unit_type,
            bedrooms=bedrooms,
            is_local=is_local,
            area_interior=area_int,
            area_balcony=area_bal,
            area_terrace=area_ter,
            area_garden=None,
            area_total=area_tot,
            parking_car=park_car,
            parking_tandem=0,
            parking_moto=0,
            parking_type=parking_type,
            parking_number=parking_number,
            parking_level=parking_level_raw,
            bodega_number=bodega_number,
            bodega_area=None,
            price_list=price,
            status=status,
        ))

    wb.close()
    return towers, floors, units, salespeople


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------

def generate_sql(
    all_towers: list[Tower],
    all_floors: list[Floor],
    all_units: list[Unit],
    all_salespeople: set[str],
) -> str:
    lines: list[str] = []
    lines.append("-- ==========================================================================")
    lines.append("-- Production Seed — Generated by scripts/seed_prod.py")
    lines.append("-- ==========================================================================")
    lines.append("-- Safety: single transaction, idempotent (ON CONFLICT DO NOTHING)")
    lines.append("-- Review this file carefully before executing.")
    lines.append("-- ==========================================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # Step 1: Fix project names and slugs
    lines.append("-- Step 1: Fix project names and slugs")
    for key in ["blt", "ce", "b5", "ben"]:
        pid = PROJECT_IDS[key]
        name = PROJECT_NAMES[key]
        slug = PROJECT_SLUGS[key]
        lines.append(
            f"UPDATE projects SET name = {sql_str(name)}, slug = {sql_str(slug)} "
            f"WHERE id = '{pid}';"
        )
    lines.append("")

    # Step 2: Salespeople
    lines.append(f"-- Step 2: Salespeople ({len(all_salespeople)} rows)")
    for sp in sorted(all_salespeople):
        # display_name = first name (or full name for companies)
        display = sp.split()[0] if " " in sp else sp
        lines.append(
            f"INSERT INTO salespeople (full_name, display_name) "
            f"VALUES ({sql_str(sp)}, {sql_str(display)}) "
            f"ON CONFLICT (full_name) DO NOTHING;"
        )
    lines.append("")

    # Step 3: Towers
    lines.append(f"-- Step 3: Towers ({len(all_towers)} rows)")
    for t in all_towers:
        pid = PROJECT_IDS[t.project_key]
        lines.append(
            f"INSERT INTO towers (project_id, name, is_default) "
            f"VALUES ('{pid}', {sql_str(t.name)}, {sql_bool(t.is_default)}) "
            f"ON CONFLICT (project_id, name) DO NOTHING;"
        )
    lines.append("")

    # Step 4: Floors (use subquery for tower_id)
    lines.append(f"-- Step 4: Floors ({len(all_floors)} rows)")
    for f in all_floors:
        pid = PROJECT_IDS[f.project_key]
        lines.append(
            f"INSERT INTO floors (tower_id, number) VALUES ("
            f"(SELECT id FROM towers WHERE project_id = '{pid}' AND name = {sql_str(f.tower_name)}), "
            f"{f.number}"
            f") ON CONFLICT (tower_id, number) DO NOTHING;"
        )
    lines.append("")

    # Step 5: Units (use subquery for floor_id)
    lines.append(f"-- Step 5: Units ({len(all_units)} rows)")
    for u in all_units:
        pid = PROJECT_IDS[u.project_key]
        floor_subq = (
            f"(SELECT f.id FROM floors f "
            f"JOIN towers t ON t.id = f.tower_id "
            f"WHERE t.project_id = '{pid}' "
            f"AND t.name = {sql_str(u.tower_name)} "
            f"AND f.number = {u.floor_number})"
        )
        lines.append(
            f"INSERT INTO rv_units ("
            f"floor_id, unit_number, unit_code, unit_type, bedrooms, is_local, "
            f"area_interior, area_balcony, area_terrace, area_garden, area_total, "
            f"parking_car, parking_tandem, parking_moto, parking_type, parking_number, parking_level, "
            f"bodega_number, bodega_area, price_list, status"
            f") VALUES ("
            f"{floor_subq}, "
            f"{sql_str(u.unit_number)}, {sql_str(u.unit_code)}, {sql_str(u.unit_type)}, "
            f"{u.bedrooms}, {sql_bool(u.is_local)}, "
            f"{sql_num(u.area_interior)}, {sql_num(u.area_balcony)}, "
            f"{sql_num(u.area_terrace)}, {sql_num(u.area_garden)}, {sql_num(u.area_total)}, "
            f"{u.parking_car}, {u.parking_tandem}, {u.parking_moto}, "
            f"{sql_str(u.parking_type)}, {sql_str(u.parking_number)}, {sql_str(u.parking_level)}, "
            f"{sql_str(u.bodega_number)}, {sql_num(u.bodega_area)}, "
            f"{sql_num(u.price_list)}, '{u.status}'::rv_unit_status"
            f") ON CONFLICT (floor_id, unit_number) DO NOTHING;"
        )
    lines.append("")

    lines.append("COMMIT;")
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # Verify all files exist
    for key, path in FILES.items():
        if not path.exists():
            print(f"ERROR: File not found: {path}", file=sys.stderr)
            sys.exit(1)
        print(f"  Found: {path.name}")

    all_towers: list[Tower] = []
    all_floors: list[Floor] = []
    all_units: list[Unit] = []
    all_salespeople: set[str] = set()

    # Extract each project
    for key, extractor in [
        ("blt", extract_blt),
        ("ben", extract_benestare),
        ("b5",  extract_b5),
        ("ce",  extract_ce),
    ]:
        print(f"\nExtracting {PROJECT_NAMES[key]}...")
        t, f, u, s = extractor(FILES[key])
        all_towers.extend(t)
        all_floors.extend(f)
        all_units.extend(u)
        all_salespeople.update(s)
        print(f"  Towers: {len(t)}, Floors: {len(f)}, Units: {len(u)}, Salespeople: {len(s)}")

    print(f"\nTotals:")
    print(f"  Towers:      {len(all_towers)}")
    print(f"  Floors:      {len(all_floors)}")
    print(f"  Units:       {len(all_units)}")
    print(f"  Salespeople: {len(all_salespeople)}")

    # Generate SQL
    sql = generate_sql(all_towers, all_floors, all_units, all_salespeople)

    out_path = ROOT / "scripts" / "seed_prod.sql"
    out_path.write_text(sql, encoding="utf-8")
    print(f"\nSQL written to: {out_path}")
    print(f"  Lines: {len(sql.splitlines())}")
    print(f"  Size: {len(sql):,} bytes")
    print(f"\nREVIEW the SQL file before executing!")


if __name__ == "__main__":
    main()

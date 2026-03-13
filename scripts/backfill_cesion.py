#!/usr/bin/env python3
"""
Production Database Backfill — Cesion de Derechos (Boulevard 5).

Reads the ACTUALIZACION xlsx workbook, extracts cesion-specific fields
(parking areas, price_suggested, PCV block, precalificacion, razon_compra,
tipo_cliente) and generates UPDATE statements for the ~280 B5 units in rv_units.

Safety:
  - UPDATE only (units already exist from seed_prod.py)
  - Idempotent: SET ... WHERE subquery matches unit_number + B5 project
  - Single transaction (BEGIN / COMMIT)
  - Pre-flight check: verifies all target units exist
  - Review backfill_cesion.sql before executing

Usage:
  python3 -u scripts/backfill_cesion.py
  # Review scripts/backfill_cesion.sql
  # Then execute via Supabase Management API
"""

from __future__ import annotations

import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
SSOT = ROOT / "origin" / "SSOT" / "Documentos varios"

SOURCE_FILE = SSOT / "ACTUALIZACION APTOS CESIÓN DE DERECHOS BOULEVARD 5 Vrs.030326.xlsx"
SHEET_NAME = "Cesión de derechos"
OUTPUT_FILE = ROOT / "scripts" / "backfill_cesion.sql"

B5_PROJECT_ID = "019c7d10-8e01-720f-942f-cac0017d83a8"

# Expected headers → semantic field name
HEADER_MAP: dict[str, str] = {
    "no. apartamento":              "unit_number",
    "metraje simple":               "parking_car_area",
    "metraje tandem":               "parking_tandem_area",
    "precio actual sugerido":       "price_suggested",
    "bloque pcv":                   "pcv_block",
    "estatus precalificacion":      "precalificacion_status",
    "estatus precalificación":      "precalificacion_status",
    "comentarios de precalificacion":  "precalificacion_notes",
    "comentarios de precalificación":  "precalificacion_notes",
    "razon de compra":              "razon_compra",
    "razón de compra":              "razon_compra",
    "tipo de cliente":              "tipo_cliente",
}


# ---------------------------------------------------------------------------
# Data class
# ---------------------------------------------------------------------------

@dataclass
class CesionRecord:
    unit_number: str
    parking_car_area: float | None
    parking_tandem_area: float | None
    price_suggested: float | None
    pcv_block: int | None
    precalificacion_status: str | None
    precalificacion_notes: str | None
    razon_compra: str | None
    tipo_cliente: str | None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def strip_accents(s: str) -> str:
    """Remove accents for fuzzy header matching."""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_header(raw: str) -> str:
    """Lowercase, strip whitespace and accents for matching."""
    return strip_accents(raw.strip()).lower()


def cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_str(val) -> str | None:
    """Convert to stripped uppercase string, or None."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s == "0":
        return None
    return s.upper()


def sql_str(val: str | None) -> str:
    if val is None:
        return "NULL"
    escaped = val.replace("'", "''")
    return f"'{escaped}'"


def sql_num(val: float | int | None) -> str:
    if val is None:
        return "NULL"
    return str(val)


def sql_bool(val: bool) -> str:
    return "true" if val else "false"


def normalize_unit_number(raw) -> str | None:
    """Normalize unit number to string. Returns None for non-unit values."""
    if raw is None:
        return None
    s = str(raw).strip()
    # Remove trailing .0 from Excel numeric cells
    if s.endswith(".0"):
        s = s[:-2]
    if not s:
        return None
    # Unit numbers are numeric (e.g., "101", "1905"). Skip summary/footer rows
    # like "Tipo", "A", "B 1 parqueo", "C2 parqueos", "D", "E", etc.
    try:
        int(s)
    except ValueError:
        return None
    return s


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def detect_headers(ws) -> dict[str, int]:
    """Scan row 1 for known headers, return {field_name: column_index}."""
    mapping: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = cell(ws, 1, col_idx)
        if raw is None:
            continue
        normalized = normalize_header(str(raw))
        for pattern, field in HEADER_MAP.items():
            if normalized == normalize_header(pattern):
                mapping[field] = col_idx
                break
    return mapping


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def extract_records(ws, col_map: dict[str, int]) -> list[CesionRecord]:
    """Extract cesion records from the worksheet."""
    records: list[CesionRecord] = []

    for row_idx in range(2, ws.max_row + 1):
        unit_raw = cell(ws, row_idx, col_map["unit_number"])
        unit_num = normalize_unit_number(unit_raw)
        if not unit_num:
            continue

        precal_status = safe_str(cell(ws, row_idx, col_map["precalificacion_status"]))
        # Normalize common variations
        if precal_status == "APROBADA":
            precal_status = "APROBADA"
        elif precal_status == "DENEGADA":
            precal_status = "DENEGADA"
        elif precal_status in ("N/A", "NA"):
            precal_status = "N/A"
        elif precal_status == "DISPONIBLE":
            precal_status = "DISPONIBLE"

        # Precalificacion notes: preserve original casing
        notes_raw = cell(ws, row_idx, col_map["precalificacion_notes"])
        precal_notes: str | None = None
        if notes_raw is not None:
            s = str(notes_raw).strip()
            if s and s != "0":
                precal_notes = s

        records.append(CesionRecord(
            unit_number=unit_num,
            parking_car_area=safe_float(cell(ws, row_idx, col_map["parking_car_area"])),
            parking_tandem_area=safe_float(cell(ws, row_idx, col_map["parking_tandem_area"])),
            price_suggested=safe_float(cell(ws, row_idx, col_map["price_suggested"])),
            pcv_block=safe_int(cell(ws, row_idx, col_map["pcv_block"])),
            precalificacion_status=precal_status,
            precalificacion_notes=precal_notes,
            razon_compra=safe_str(cell(ws, row_idx, col_map["razon_compra"])),
            tipo_cliente=safe_str(cell(ws, row_idx, col_map["tipo_cliente"])),
        ))

    return records


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------

def generate_sql(records: list[CesionRecord]) -> list[str]:
    """Generate UPDATE statements for rv_units."""
    lines: list[str] = []

    lines.append("-- ==========================================================================")
    lines.append("-- Backfill: Cesion de Derechos — Boulevard 5")
    lines.append(f"-- Source: ACTUALIZACION APTOS CESION DE DERECHOS BOULEVARD 5 Vrs.030326.xlsx")
    lines.append(f"-- Records: {len(records)}")
    lines.append("-- ==========================================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # Pre-flight: verify target units exist
    unit_numbers = [r.unit_number for r in records]
    unit_list = ", ".join(f"'{u}'" for u in unit_numbers)
    lines.append("-- Pre-flight: verify all target units exist in rv_units for B5")
    lines.append("DO $$ BEGIN")
    lines.append("  IF (")
    lines.append(f"    SELECT count(*) FROM rv_units ru")
    lines.append(f"    JOIN floors f ON f.id = ru.floor_id")
    lines.append(f"    JOIN towers t ON t.id = f.tower_id")
    lines.append(f"    WHERE t.project_id = '{B5_PROJECT_ID}'")
    lines.append(f"    AND ru.unit_number IN ({unit_list})")
    lines.append(f"  ) < {len(records)}")
    lines.append("  THEN RAISE EXCEPTION")
    lines.append(f"    'Pre-flight failed: expected {len(records)} B5 units, found fewer. Aborting.';")
    lines.append("  END IF;")
    lines.append("END $$;")
    lines.append("")

    # Generate UPDATE for each unit
    for i, r in enumerate(records):
        lines.append(f"-- [{i + 1}/{len(records)}] Unit {r.unit_number}")
        lines.append("UPDATE rv_units SET")
        lines.append(f"  parking_car_area        = {sql_num(r.parking_car_area)},")
        lines.append(f"  parking_tandem_area     = {sql_num(r.parking_tandem_area)},")
        lines.append(f"  price_suggested         = {sql_num(r.price_suggested)},")
        lines.append(f"  is_cesion               = true,")
        lines.append(f"  pcv_block               = {sql_num(r.pcv_block)},")
        lines.append(f"  precalificacion_status  = {sql_str(r.precalificacion_status)},")
        lines.append(f"  precalificacion_notes   = {sql_str(r.precalificacion_notes)},")
        lines.append(f"  razon_compra            = {sql_str(r.razon_compra)},")
        lines.append(f"  tipo_cliente            = {sql_str(r.tipo_cliente)}")
        lines.append("WHERE id = (")
        lines.append(f"  SELECT ru.id FROM rv_units ru")
        lines.append(f"  JOIN floors f ON f.id = ru.floor_id")
        lines.append(f"  JOIN towers t ON t.id = f.tower_id")
        lines.append(f"  WHERE t.project_id = '{B5_PROJECT_ID}'")
        lines.append(f"  AND ru.unit_number = '{r.unit_number}'")
        lines.append(");")
        lines.append("")

    lines.append("COMMIT;")
    return lines


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # Verify source file exists
    if not SOURCE_FILE.exists():
        print(f"ERROR: Source file not found: {SOURCE_FILE}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading: {SOURCE_FILE.name}")
    wb = openpyxl.load_workbook(str(SOURCE_FILE), read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # Detect headers
    col_map = detect_headers(ws)
    required_fields = [
        "unit_number", "parking_car_area", "parking_tandem_area",
        "price_suggested", "pcv_block", "precalificacion_status",
        "precalificacion_notes", "razon_compra", "tipo_cliente",
    ]
    missing = [f for f in required_fields if f not in col_map]
    if missing:
        print(f"ERROR: Missing required columns: {missing}", file=sys.stderr)
        print(f"Detected columns: {col_map}", file=sys.stderr)
        sys.exit(1)

    print(f"Column mapping: {col_map}")

    # Extract records
    records = extract_records(ws, col_map)
    wb.close()

    print(f"Extracted {len(records)} cesion records")

    if not records:
        print("ERROR: No records extracted. Aborting.", file=sys.stderr)
        sys.exit(1)

    # Deduplicate: keep last occurrence per unit_number (later rows supersede)
    seen: dict[str, int] = {}
    for i, r in enumerate(records):
        seen[r.unit_number] = i
    deduped = [records[i] for i in sorted(seen.values())]
    if len(deduped) < len(records):
        print(f"Deduplicated: {len(records)} -> {len(deduped)} (kept last occurrence per unit)")
    records = deduped

    # Summary
    pcv_counts: dict[int | None, int] = {}
    precal_counts: dict[str | None, int] = {}
    for r in records:
        pcv_counts[r.pcv_block] = pcv_counts.get(r.pcv_block, 0) + 1
        precal_counts[r.precalificacion_status] = precal_counts.get(r.precalificacion_status, 0) + 1

    print(f"PCV block distribution: {dict(sorted((k, v) for k, v in pcv_counts.items() if k is not None))}")
    print(f"Precalificacion distribution: {dict(sorted(precal_counts.items(), key=lambda x: str(x[0])))}")

    # Generate SQL
    sql_lines = generate_sql(records)
    OUTPUT_FILE.write_text("\n".join(sql_lines), encoding="utf-8")
    print(f"Wrote {len(sql_lines)} lines to {OUTPUT_FILE}")
    print("Review the SQL file, then execute via Supabase Management API.")


if __name__ == "__main__":
    main()

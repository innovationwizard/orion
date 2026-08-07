"""Regenerate src/lib/mercadeo/spend-snapshot.json from the Meta Ads source Excel.

Reads ARTIFACT/Performance-Report-Puerta-Abierta.xlsx (the Power BI source file).
KEY FINDING (2026-08-07): the spend column is labeled "Importe gastado (USD)" but
the Divisa column shows each ad account bills in its own currency — 4 accounts in
GTQ, 2 in USD. The Power BI model ignores Divisa and mixes currencies. This
extraction preserves the real per-account currency.

Usage: python3 scripts/extract-mercadeo-spend.py
"""
import json
from collections import defaultdict
from openpyxl import load_workbook

SRC = "ARTIFACT/Performance-Report-Puerta-Abierta.xlsx"
OUT = "src/lib/mercadeo/spend-snapshot.json"

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Performance - Data Base PA"]
rows = ws.iter_rows(values_only=True)
header = next(rows)
idx = {h: i for i, h in enumerate(header)}

accounts: dict[str, dict] = {}
for r in rows:
    if r[idx["Fecha"]] is None:
        continue
    acc = r[idx["Nombre de la cuenta"]]
    divisa = r[idx["Divisa"]]
    spend = float(r[idx["Importe gastado (USD)"]] or 0)
    date = str(r[idx["Fecha"]])[:10]
    e = accounts.setdefault(acc, {"currency": divisa, "spend": 0.0, "currencies": defaultdict(float), "from": date, "to": date})
    e["spend"] += spend
    e["currencies"][divisa] += spend
    e["from"] = min(e["from"], date)
    e["to"] = max(e["to"], date)

out_accounts = {}
for acc, e in sorted(accounts.items()):
    currencies = dict(e["currencies"])
    if len(currencies) != 1:
        raise SystemExit(f"Account {acc} mixes currencies {currencies} — extraction assumes one currency per account; fix before regenerating.")
    out_accounts[acc] = {
        "currency": e["currency"],
        "spend": round(e["spend"], 2),
        "from": e["from"],
        "to": e["to"],
    }

out = {
    "_source": (
        "Extracted from ARTIFACT/Performance-Report-Puerta-Abierta.xlsx (Meta Ads export, "
        "Power BI source). The 'Importe gastado (USD)' column label is WRONG for GTQ-billed "
        "accounts — real currency comes from the Divisa column, one per ad account. "
        "Regenerate with scripts/extract-mercadeo-spend.py after replacing the Excel."
    ),
    "accounts": out_accounts,
}
json.dump(out, open(OUT, "w"), ensure_ascii=False, indent=2)
for acc, e in out_accounts.items():
    print(f"{acc:22} {e['currency']} {e['spend']:>12,.2f} | {e['from']} → {e['to']}")
print(f"wrote {OUT}")

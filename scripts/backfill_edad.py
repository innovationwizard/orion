#!/usr/bin/env python3
"""
Backfill rv_client_profiles.edad from SSOT edad (actual integer values).

The original backfill_domains.py skipped birth_date because the SSOT sheets
have edad (age in years), not fecha de nacimiento (date of birth).

This script stores the actual edad integer — no fabrication.
For new clients, the actual birth_date is extracted from DPI photos via OCR.

This script:
  1. Re-reads the Buyer Persona sheets from each project's Reporte de Ventas
  2. Extracts edad (integer) for each unit
  3. Generates SQL UPDATE statements for rv_client_profiles.edad

Usage:
  python3 -u scripts/backfill_edad.py
  # Review scripts/backfill_edad.sql
  # Then execute via Supabase Management API
"""

from __future__ import annotations

import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Constants (same as backfill_domains.py)
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
SSOT = ROOT / "origin" / "SSOT" / "Reservas y Ventas"

REPORTE_FILES = {
    "blt": SSOT / "Bosque Las Tapias" / "2. Reporte  de Ventas Bosque Las Tapias septiembre 2025.xlsx",
    "ben": SSOT / "Benestare" / "2. Reporte  de Ventas BENESTARE septiembre 2025.xlsx",
    "b5":  SSOT / "Boulevard 5" / "2. Reporte  de Ventas BOULEVARD 5 septiembre 2025.xlsx",
    "ce":  SSOT / "Casa Elisa" / "2 .Reporte de Ventas CASA ELISA septiembre 2025.xlsx",
}

PROJECT_IDS = {
    "blt": "019c7d10-8ee5-7999-9881-2cd5ad038aa9",
    "ce":  "019c7d10-8e7b-7db6-93be-6f42d0538233",
    "b5":  "019c7d10-8e01-720f-942f-cac0017d83a8",
    "ben": "019c7d10-8f5a-74c7-b3df-c2151ad8a376",
}

PROJECT_NAMES = {
    "blt": "Bosque Las Tapias",
    "ce":  "Casa Elisa",
    "b5":  "Boulevard 5",
    "ben": "Benestare",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def normalize_unit_number(raw) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if text.upper().startswith("L-"):
        return text.upper()
    n = safe_int(text)
    if n is not None and n >= 100:
        return str(n)
    return None


def sql_str(val: str | None) -> str:
    if val is None:
        return "NULL"
    escaped = val.replace("'", "''")
    return f"'{escaped}'"


# ---------------------------------------------------------------------------
# Extract edad from Buyer Persona sheets
# ---------------------------------------------------------------------------

def extract_edad(project_key: str, wb_path: Path) -> list[dict]:
    """Extract unit_number + edad from the Buyer Persona sheet."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    bp_sheet = None
    for name in wb.sheetnames:
        if "buyer persona" in name.lower():
            bp_sheet = name
            break
    if bp_sheet is None:
        wb.close()
        return []

    ws = wb[bp_sheet]
    records: list[dict] = []

    # Detect header columns
    header_map = {
        "apto": "apto",
        "edad": "edad",
    }
    cols: dict[str, int] = {}

    for row_num in range(1, 5):
        for col in range(1, 30):
            val = cell(ws, row_num, col)
            if val is None:
                continue
            text = strip_accents(str(val).strip().lower())
            for pattern, field_name in header_map.items():
                if text.startswith(pattern):
                    if field_name not in cols:
                        cols[field_name] = col
                    break
        if "apto" in cols and "edad" in cols:
            break

    if "apto" not in cols or "edad" not in cols:
        print(f"    WARN: Missing columns (apto={cols.get('apto')}, edad={cols.get('edad')}) — skipping")
        wb.close()
        return []

    header_row = 3

    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        apto_raw = cell(ws, row, cols["apto"])
        if apto_raw is None:
            continue

        apto_str = str(apto_raw).strip()
        if not apto_str:
            continue

        # Parse apartment number (same logic as backfill_domains.py)
        tower_name = None
        unit_number = None

        if project_key == "ben":
            match = re.match(r'^([A-E])\s+(\d+)$', apto_str, re.IGNORECASE)
            if match:
                tower_name = f"Torre {match.group(1).upper()}"
                unit_number = match.group(2)
            else:
                unit_number = normalize_unit_number(apto_str)
        else:
            unit_number = normalize_unit_number(apto_str)

        if unit_number is None:
            continue

        if project_key in ("b5", "ce"):
            tower_name = "Principal"

        edad = safe_int(cell(ws, row, cols["edad"]))
        if edad is None or edad < 18 or edad > 100:
            continue

        records.append({
            "project_key": project_key,
            "tower_name": tower_name,
            "unit_number": unit_number,
            "edad": edad,
        })

    wb.close()

    # Deduplicate by unit key — keep first occurrence
    seen: set[str] = set()
    deduped: list[dict] = []
    for rec in records:
        key = f"{rec['project_key']}|{rec.get('tower_name', '')}|{rec['unit_number']}"
        if key not in seen:
            seen.add(key)
            deduped.append(rec)

    return deduped


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------

def generate_sql(records: list[dict]) -> str:
    """Generate UPDATE statements to set edad (actual integer from SSOT)."""
    lines: list[str] = []
    lines.append("-- ==========================================================================")
    lines.append("-- Backfill rv_client_profiles.edad from SSOT (actual integer values)")
    lines.append("-- Generated by scripts/backfill_edad.py")
    lines.append("-- Source: Buyer Persona sheets in Reporte de Ventas (Sep 2025)")
    lines.append("-- ==========================================================================")
    lines.append("")

    count = 0
    for rec in records:
        pid = PROJECT_IDS[rec["project_key"]]

        if rec["tower_name"]:
            unit_subq = (
                f"(SELECT ru.id FROM rv_units ru "
                f"JOIN floors f ON f.id = ru.floor_id "
                f"JOIN towers t ON t.id = f.tower_id "
                f"WHERE t.project_id = '{pid}' "
                f"AND t.name = {sql_str(rec['tower_name'])} "
                f"AND ru.unit_number = {sql_str(rec['unit_number'])})"
            )
        else:
            unit_subq = (
                f"(SELECT ru.id FROM rv_units ru "
                f"JOIN floors f ON f.id = ru.floor_id "
                f"JOIN towers t ON t.id = f.tower_id "
                f"WHERE t.project_id = '{pid}' "
                f"AND ru.unit_number = {sql_str(rec['unit_number'])} LIMIT 1)"
            )

        client_subq = (
            f"(SELECT rc.client_id FROM reservation_clients rc "
            f"JOIN reservations r ON r.id = rc.reservation_id "
            f"WHERE r.unit_id = {unit_subq} "
            f"AND rc.is_primary = true "
            f"ORDER BY r.created_at DESC LIMIT 1)"
        )

        lines.append(
            f"UPDATE rv_client_profiles SET edad = {rec['edad']} "
            f"WHERE client_id = {client_subq} AND edad IS NULL;"
        )
        count += 1

    lines.append("")
    lines.append(f"-- Total: {count} UPDATE statements")
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=== Backfill edad → birth_date ===\n")

    for key, path in REPORTE_FILES.items():
        if not path.exists():
            print(f"  ERROR: Not found: {path}", file=sys.stderr)
            sys.exit(1)
    print("  All files found.\n")

    all_records: list[dict] = []
    for key in ["blt", "ben", "b5", "ce"]:
        records = extract_edad(key, REPORTE_FILES[key])
        print(f"  {PROJECT_NAMES[key]}: {len(records)} edad records")
        all_records.extend(records)

    print(f"\n  Total: {len(all_records)} records with edad\n")

    sql = generate_sql(all_records)
    out_path = ROOT / "scripts" / "backfill_edad.sql"
    out_path.write_text(sql, encoding="utf-8")
    print(f"  SQL written to: {out_path}")
    print(f"  Lines: {len(sql.splitlines())}")
    print(f"\n  REVIEW the SQL file before executing!")


if __name__ == "__main__":
    main()

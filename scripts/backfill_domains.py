#!/usr/bin/env python3
"""
Production Database Backfill — Referidos, Valorizacion, Buyer Persona.

Reads the 4 "Reporte de Ventas" xlsx workbooks and the Benestare
Disponibilidad file. Extracts referral records, price history increments,
and buyer persona profiles. Generates idempotent SQL.

Safety:
  - All INSERTs into empty tables (pre-flight check)
  - Single transaction (BEGIN / COMMIT)
  - No DELETEs or UPDATEs
  - Review backfill_domains.sql before executing

Usage:
  python3 -u scripts/backfill_domains.py
  # Review scripts/backfill_domains.sql
  # Then execute via Supabase Management API
"""

from __future__ import annotations

import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
SSOT = ROOT / "origin" / "SSOT" / "Reservas y Ventas"

REPORTE_FILES = {
    "blt": SSOT / "Bosque Las Tapias" / "2. Reporte  de Ventas Bosque Las Tapias septiembre 2025.xlsx",
    "ben": SSOT / "Benestare" / "2. Reporte  de Ventas BENESTARE septiembre 2025.xlsx",
    "b5":  SSOT / "Boulevard 5" / "2. Reporte  de Ventas BOULEVARD 5 septiembre 2025.xlsx",
    "ce":  SSOT / "Casa Elisa" / "2 .Reporte de Ventas CASA ELISA septiembre 2025.xlsx",
}

DISP_FILES = {
    "ben": SSOT / "Benestare" / "Precios y Disponibilidad BENESTARE TA septiembre 25.xlsx",
    "b5":  SSOT / "Boulevard 5" / "DisponibilIdad e Integración B5 septiembre 2025.xlsx",
    "ce":  SSOT / "Casa Elisa" / "CASA ELISA con trazabilidad precios septiembre 2025.xlsx",
}

PROJECT_IDS = {
    "blt": "019c7d10-8ee5-7999-9881-2cd5ad038aa9",
    "ce":  "019c7d10-8e7b-7db6-93be-6f42d0538233",
    "b5":  "019c7d10-8e01-720f-942f-cac0017d83a8",
    "ben": "019c7d10-8f5a-74c7-b3df-c2151ad8a376",
}

PROJECT_NAMES = {
    "blt": "Bosque Las Tapias",
    "ce":  "Casa Elisa",
    "b5":  "Boulevard 5",
    "ben": "Benestare",
}

# Same canonical map from seed_prod.py / backfill_reservations.py
SALESPERSON_CANONICAL: dict[str, str] = {
    "paula hernandez": "Paula Hernandez", "paula hernadez": "Paula Hernandez",
    "paula hernande": "Paula Hernandez", "jose gutierrez": "Jose Gutierrez",
    "antonio rada": "Antonio Rada", "brenda bucaro": "Brenda Bucaro",
    "alexander franco": "Alexander Franco", "erwin cardona": "Erwin Cardona",
    "anahi cisneros": "Anahi Cisneros", "diana alvarez": "Diana Alvarez",
    "sofia paredes": "Sofia Paredes", "laura molina": "Laura Molina",
    "efren sanchez": "Efren Sanchez", "eder veliz": "Eder Veliz",
    "pedro pablo sarti": "Pedro Pablo Sarti", "ronaldo ogaldez": "Ronaldo Ogaldez",
    "pablo marroquin": "Pablo Marroquin", "rony ramirez": "Rony Ramirez",
    "noemi menendez": "Noemi Menendez", "mario rodriguez": "Mario Rodriguez",
    "forma capital": "Forma Capital", "junta directiva": "Junta Directiva",
    "alejandra calderon": "Alejandra Calderon", "andrea gonzalez": "Andrea Gonzalez",
    "abigail garcia": "Abigail Garcia", "ivan castillo": "Ivan Castillo",
    "karina fuentes": "Karina Fuentes", "luis esteban": "Luis Esteban",
    "otto h.": "Otto Herrera", "otto herrera": "Otto Herrera",
}
SALESPERSON_EXCLUDE = {"traslado", "jd"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_salesperson(raw: str | None) -> str | None:
    if raw is None:
        return None
    cleaned = re.sub(r'\s+', ' ', str(raw).strip())
    if not cleaned:
        return None
    for sep in [" / ", "/", ","]:
        if sep in cleaned:
            cleaned = cleaned.split(sep)[0].strip()
            break
    if not cleaned:
        return None
    key = strip_accents(cleaned).lower()
    if key in SALESPERSON_EXCLUDE:
        return None
    return SALESPERSON_CANONICAL.get(key, cleaned)


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def safe_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = str(val).strip()
    if not text or text.lower() in ("pendiente", "pendiente de informar"):
        return None
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def sql_str(val: str | None) -> str:
    if val is None:
        return "NULL"
    escaped = val.replace("'", "''")
    return f"'{escaped}'"


def sql_num(val: float | int | None) -> str:
    if val is None:
        return "NULL"
    return str(val)


def sql_date(val: date | None) -> str:
    if val is None:
        return "NULL"
    return f"'{val.isoformat()}'"


def normalize_unit_number(raw) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if text.upper().startswith("L-"):
        return text.upper()
    n = safe_int(text)
    if n is not None and n >= 100:
        return str(n)
    return None


# ---------------------------------------------------------------------------
# Domain 1: Referidos
# ---------------------------------------------------------------------------

def extract_referidos(project_key: str, wb_path: Path) -> list[dict]:
    """Extract referral records from the Referidos sheet."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "referido" in name.lower():
            sheet = name
            break
    if sheet is None:
        wb.close()
        return []

    ws = wb[sheet]
    records: list[dict] = []

    # Detect headers
    header_row = 0
    cols: dict[str, int] = {}
    for row_num in range(1, 5):
        for col in range(1, 20):
            val = cell(ws, row_num, col)
            if val is None:
                continue
            text = strip_accents(str(val).strip().lower())
            if "cliente" in text and "nombre" not in text:
                cols["cliente"] = col
            elif "numero apartamento" in text or "apartamento" in text.replace("numero ", ""):
                cols["unit"] = col
            elif "precio de lista" in text or text == "precio de lista":
                cols["precio_lista"] = col
            elif "precio de referido" in text or text == "precio de referido":
                cols["precio_referido"] = col
            elif "referido por" in text:
                cols["referido_por"] = col
            elif "fecha" in text and "reserva" in text:
                cols["fecha"] = col
            elif text == "asesor":
                cols["asesor"] = col
        if "cliente" in cols and "unit" in cols:
            header_row = row_num
            break

    if header_row == 0:
        wb.close()
        return []

    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        client_raw = cell(ws, row, cols.get("cliente", 0))
        if client_raw is None or str(client_raw).strip() == "":
            continue

        client_name = str(client_raw).strip()
        if len(client_name) < 3:
            continue

        unit_raw = cell(ws, row, cols.get("unit", 0))
        unit_num = normalize_unit_number(unit_raw)
        if unit_num is None:
            continue

        referido_raw = cell(ws, row, cols.get("referido_por", 0))
        referido = str(referido_raw).strip() if referido_raw else None
        if not referido or len(referido) < 3 or referido.lower() in ("pendiente", "pendiente de informar"):
            continue  # referido_por is NOT NULL in schema

        precio_lista = safe_float(cell(ws, row, cols.get("precio_lista", 0)))
        precio_referido = safe_float(cell(ws, row, cols.get("precio_referido", 0)))
        fecha = safe_date(cell(ws, row, cols.get("fecha", 0)))
        asesor = normalize_salesperson(str(cell(ws, row, cols.get("asesor", 0))) if cols.get("asesor") else None)

        records.append({
            "project_key": project_key,
            "unit_number": unit_num,
            "tower_name": "Principal" if project_key in ("b5", "ce") else None,
            "client_name": client_name,
            "precio_lista": precio_lista,
            "precio_referido": precio_referido,
            "referido_por": referido,
            "fecha": fecha,
            "salesperson": asesor,
        })

    wb.close()
    return records


# For BLT referidos, need tower info — the wider schema has it
def extract_referidos_blt(wb_path: Path) -> list[dict]:
    """BLT has a wider 13-col schema with Torre column."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "referido" in name.lower():
            sheet = name
            break
    if sheet is None:
        wb.close()
        return []

    ws = wb[sheet]
    records: list[dict] = []

    # BLT headers are at row 2
    # Col 2=Cliente, 3=Numero Apartamento, 4=Apto, 5=Torre,
    # 6=Precio de Lista, 7=Precio de Referido, 8=Referido Por,
    # 12=Fecha de Reserva, 13=Asesor
    for row in range(3, (ws.max_row or 0) + 1):
        client_raw = cell(ws, row, 2)
        if client_raw is None or str(client_raw).strip() == "":
            continue

        client_name = str(client_raw).strip()
        if len(client_name) < 3:
            continue

        unit_raw = cell(ws, row, 3)
        unit_num = normalize_unit_number(unit_raw)
        if unit_num is None:
            continue

        torre_raw = cell(ws, row, 5)
        torre = None
        if torre_raw:
            t = str(torre_raw).strip()
            if re.match(r'^[A-E]$', t, re.IGNORECASE):
                torre = f"Torre {t.upper()}"
            elif t.lower().startswith("torre"):
                torre = t

        referido_raw = cell(ws, row, 8)
        referido = str(referido_raw).strip() if referido_raw else None
        if not referido or len(referido) < 3:
            continue

        precio_lista = safe_float(cell(ws, row, 6))
        precio_referido = safe_float(cell(ws, row, 7))
        fecha = safe_date(cell(ws, row, 12))
        asesor = normalize_salesperson(str(cell(ws, row, 13)) if cell(ws, row, 13) else None)

        records.append({
            "project_key": "blt",
            "unit_number": unit_num,
            "tower_name": torre,
            "client_name": client_name,
            "precio_lista": precio_lista,
            "precio_referido": precio_referido,
            "referido_por": referido,
            "fecha": fecha,
            "salesperson": asesor,
        })

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Domain 2: Valorizacion (BEN only — cleanest data)
# ---------------------------------------------------------------------------

def extract_valorizacion_ben(wb_path: Path) -> list[dict]:
    """Extract Benestare price history from the Valorizacion sheet."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # Also read increments from the Precios sheet headers to get dates
    ws_precios = wb["Precios"]

    # The Precios sheet has increment columns (14-19) with dates in the header rows
    # Row 2 has dates for some increments
    increment_dates: dict[int, date | None] = {}
    for col in range(14, 20):
        val = cell(ws_precios, 2, col)
        d = safe_date(val)
        if d:
            increment_dates[col - 13] = d  # 1-indexed: col 14=Incremento 1, etc.

    # The Valorizacion sheet has the aggregate data
    val_sheet = None
    for name in wb.sheetnames:
        if "valoriza" in name.lower():
            val_sheet = name
            break

    if val_sheet is None:
        wb.close()
        return []

    ws = wb[val_sheet]
    records: list[dict] = []

    # Parse the Valorizacion sheet structure:
    # It has blocks of data for each increment, roughly:
    # Row 5-6: Valor inicial (08/2023)
    # Row 9-10: Incremento 1 (07/2025)
    # Row 13-14: Incremento 2 (08/2025)
    # Row 17-18: Incremento 3 (12/2025)
    # Row 21-22: Incremento 4 (01/2026)
    # Row 25-26: Incremento 5 (02/2026)
    # Row 29-30: Incremento 6 (03/2026)
    #
    # Each block has: month, units_remaining, per_unit_increment, total_delta, project_value

    # Scan for rows with increment data
    for row in range(2, (ws.max_row or 0) + 1):
        # Look for cells containing "Incremento" or increment dates
        for col in range(1, 17):
            val = cell(ws, row, col)
            if val is None:
                continue
            text = str(val).strip()

            # Check if this row has an increment reference
            if "incremento" in text.lower() and ":" in text:
                # Format: "Incremento 1: Q10,000 por unidad"
                match = re.search(r'incremento\s+(\d+)', text, re.IGNORECASE)
                if match:
                    inc_num = int(match.group(1))

                    # Look for associated data in nearby cells
                    # Try to find: units_remaining, increment_amount, total
                    units_rem = None
                    inc_amount = None
                    total_delta = None

                    # Search the block (this row and next few rows, cols 1-16)
                    for scan_row in range(max(row - 2, 1), min(row + 4, (ws.max_row or 0) + 1)):
                        for scan_col in range(1, 17):
                            sv = cell(ws, scan_row, scan_col)
                            if sv is None:
                                continue
                            st = str(sv).strip().lower()

                            if "unidad" in st and "disponible" in st:
                                # "240 unidades disponibles" → units_remaining
                                m = re.search(r'(\d+)\s*unidad', st)
                                if m:
                                    units_rem = int(m.group(1))
                            elif isinstance(sv, (int, float)) and scan_col > 1:
                                n = safe_float(sv)
                                if n is not None:
                                    if 1000 <= n <= 50000 and inc_amount is None:
                                        inc_amount = n
                                    elif n > 100000 and total_delta is None:
                                        total_delta = n

    # The structured approach above may miss data. Let me use a simpler approach:
    # Parse the Precios sheet directly to count increments and their values.

    # From the exploration, BEN has these increments in the Precios sheet:
    # Col 13: Precio Inicial
    # Col 14: Incremento 1 (no date in header — ~07/2025)
    # Col 15: Incremento 2 (2025-08-11)
    # Col 16: Incremento 3 (2025-12-01)
    # Col 17: Incremento 4 (2026-01-07)
    # Col 18: Incremento 5 (2026-02-02)
    # Col 19: Incremento 6 (2026-03-01)

    # For each increment column, count non-zero cells and get the increment amount
    records = []
    known_dates = {
        1: date(2025, 7, 1),   # Incremento 1 — approximate
        2: date(2025, 8, 11),  # From header
        3: date(2025, 12, 1),
        4: date(2026, 1, 7),
        5: date(2026, 2, 2),
        6: date(2026, 3, 1),
    }

    for inc_num in range(1, 7):
        col_idx = 13 + inc_num  # Col 14 for inc 1, etc.

        # Override with detected header dates if available
        inc_date = increment_dates.get(inc_num) or known_dates.get(inc_num)
        if inc_date is None:
            continue

        # Count non-zero cells and extract the increment amount
        non_zero_count = 0
        amounts: list[float] = []
        for data_row in range(4, (ws_precios.max_row or 0) + 1):
            val = cell(ws_precios, data_row, col_idx)
            n = safe_float(val)
            if n is not None and n != 0:
                non_zero_count += 1
                amounts.append(n)

        if non_zero_count == 0:
            continue

        # BEN increments are uniform (same amount per unit)
        # Use the most common amount as the increment
        inc_amount = max(set(amounts), key=amounts.count) if amounts else 0
        total = sum(amounts)

        records.append({
            "project_key": "ben",
            "tower_id": None,  # Project-level
            "effective_date": inc_date,
            "units_remaining": non_zero_count,
            "increment_amount": inc_amount,
            "appreciation_total": round(total, 2),
            "notes": f"Incremento {inc_num}",
        })

    wb.close()
    return records


def extract_valorizacion_b5(wb_path: Path) -> list[dict]:
    """Extract Boulevard 5 price adjustments from Matriz Precios A sheet.

    Each adjustment column has per-unit deltas. We aggregate per event:
    sum of deltas = appreciation_total, count of non-zero = units_affected.
    """
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "matriz precios a" in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        wb.close()
        return []

    ws = wb[sheet_name]

    # Map adjustment columns: (column_index, date, label)
    # Row 3 has headers, row 2 may have dates
    # From exploration: headers in row 3
    adjustments: list[tuple[int, date | None, str]] = []
    for col in range(1, 50):
        val3 = cell(ws, 3, col)
        if val3 is None:
            continue
        text = str(val3).strip()
        if not re.search(r'ajuste|descuento|incremento', text, re.IGNORECASE):
            continue

        # Try to get date from row 2 or row 1
        d = safe_date(cell(ws, 2, col)) or safe_date(cell(ws, 1, col))
        adjustments.append((col, d, text))

    # Aggregate per-unit deltas for each adjustment
    records: list[dict] = []
    data_start = 4  # Data starts at row 4

    # Pre-build set of rows that have a unit number in col B (skip summary rows)
    unit_rows: set[int] = set()
    for row in range(data_start, (ws.max_row or 0) + 1):
        if cell(ws, row, 2) is not None:  # Col B = unit number
            unit_rows.add(row)

    for col_idx, adj_date, label in adjustments:
        if adj_date is None:
            continue  # Skip undated adjustments

        amounts: list[float] = []
        for row in unit_rows:
            val = cell(ws, row, col_idx)
            n = safe_float(val)
            if n is not None and n != 0:
                amounts.append(n)

        if not amounts:
            continue

        total = round(sum(amounts), 2)
        records.append({
            "project_key": "b5",
            "tower_id": None,
            "effective_date": adj_date,
            "units_remaining": len(amounts),
            "increment_amount": total,
            "appreciation_total": total,
            "notes": label,
        })

    wb.close()
    return records


def extract_valorizacion_ce(wb_path: Path) -> list[dict]:
    """Extract Casa Elisa price adjustments from Disponibilidad sheet.

    Columns Q-AK have per-unit deltas with dates in rows 2-3.
    Header format is inconsistent: cols Q-AE have category in row 2 and date
    in row 3; cols AF-AJ may swap them.

    Opens without read_only to avoid slow random-access on large sheet.
    """
    wb = openpyxl.load_workbook(wb_path, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "disponibilidad" in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        wb.close()
        return []

    ws = wb[sheet_name]

    # Scan columns Q (17) through AK (37) for adjustment events
    records: list[dict] = []
    data_start = 5
    max_row = min(ws.max_row or 100, 200)  # CE has ~78 units, cap at 200

    # Pre-build set of rows that have a unit number in col B (skip summary rows)
    unit_rows: set[int] = set()
    for row in range(data_start, max_row + 1):
        if cell(ws, row, 2) is not None:  # Col B = unit number
            unit_rows.add(row)

    for col in range(17, 38):
        # Try to extract a date from row 2 or row 3
        val2 = cell(ws, 2, col)
        val3 = cell(ws, 3, col)

        adj_date = safe_date(val2) or safe_date(val3)
        if adj_date is None:
            continue  # Skip undated adjustments

        # Get a label from the other row
        label_val = val3 if safe_date(val2) else val2
        label = str(label_val).strip() if label_val else f"Ajuste col {col}"

        # Aggregate per-unit deltas (only rows with a unit number)
        amounts: list[float] = []
        for row in unit_rows:
            val = cell(ws, row, col)
            n = safe_float(val)
            if n is not None and n != 0:
                amounts.append(n)

        if not amounts:
            continue

        total = round(sum(amounts), 2)
        records.append({
            "project_key": "ce",
            "tower_id": None,
            "effective_date": adj_date,
            "units_remaining": len(amounts),
            "increment_amount": total,
            "appreciation_total": total,
            "notes": label,
        })

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Domain 3: Buyer Persona
# ---------------------------------------------------------------------------

def extract_buyer_persona(project_key: str, wb_path: Path) -> list[dict]:
    """Extract buyer persona profiles from the Buyer Persona sheet."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # Find the Buyer Persona sheet
    bp_sheet = None
    for name in wb.sheetnames:
        if "buyer persona" in name.lower():
            bp_sheet = name
            break
    if bp_sheet is None:
        wb.close()
        return []

    ws = wb[bp_sheet]
    records: list[dict] = []

    # Detect headers (always at row 3 based on exploration)
    cols: dict[str, int] = {}
    header_map = {
        "apto": "apto",
        "habitaciones": "habitaciones",
        "tipo": "tipo",
        "nombre": "nombre",
        "correo": "correo",
        "telefono": "telefono",
        "genero": "genero",
        "edad": "edad",
        "estudios": "estudios",
        "tipo compra": "tipo_compra",
        "estado civil": "estado_civil",
        "hijos si o no": "hijos",
        "cantidad hijos": "cantidad_hijos",
        "departamento": "departamento",
        "zona": "zona",
        "trabajo": "trabajo",
        "industria": "industria",
        "ingreso individual": "ingreso_individual",
        "ingreso familiar": "ingreso_familiar",
        "medio": "medio",
        "asesor": "asesor",
        "direccion vivienda": "direccion_vivienda",
        "direccion trabajo": "direccion_trabajo",
    }

    for row_num in range(1, 5):
        for col in range(1, 30):
            val = cell(ws, row_num, col)
            if val is None:
                continue
            text = strip_accents(str(val).strip().lower())
            for pattern, field_name in header_map.items():
                if text.startswith(pattern):
                    if field_name not in cols:
                        cols[field_name] = col
                    break
        if "apto" in cols and ("nombre" in cols or "genero" in cols):
            break

    if "apto" not in cols:
        print(f"    WARN: No 'Apto' column found — skipping")
        wb.close()
        return []

    header_row = 3  # All buyer persona sheets use row 3

    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        apto_raw = cell(ws, row, cols["apto"])
        if apto_raw is None:
            continue

        # Parse apartment number
        apto_str = str(apto_raw).strip()
        if not apto_str:
            continue

        # BEN format: "B 101" → tower_letter="B", unit_number="101"
        tower_name = None
        unit_number = None

        if project_key == "ben":
            # Parse "B 101", "C 302", etc.
            match = re.match(r'^([A-E])\s+(\d+)$', apto_str, re.IGNORECASE)
            if match:
                tower_name = f"Torre {match.group(1).upper()}"
                unit_number = match.group(2)
            else:
                unit_number = normalize_unit_number(apto_str)
        else:
            unit_number = normalize_unit_number(apto_str)

        if unit_number is None:
            continue

        if project_key in ("b5", "ce"):
            tower_name = "Principal"

        # Extract fields
        def get_str(field: str) -> str | None:
            if field not in cols:
                return None
            v = cell(ws, row, cols[field])
            if v is None:
                return None
            s = str(v).strip()
            return s if s else None

        nombre = get_str("nombre")
        genero_raw = get_str("genero")
        edad_raw = cell(ws, row, cols["edad"]) if "edad" in cols else None
        estudios = get_str("estudios")
        tipo_compra_raw = get_str("tipo_compra")
        estado_civil = get_str("estado_civil")
        hijos_raw = get_str("hijos")
        cantidad_hijos = safe_int(cell(ws, row, cols["cantidad_hijos"])) if "cantidad_hijos" in cols else None
        departamento = get_str("departamento")
        zona = get_str("zona")
        trabajo_raw = get_str("trabajo")
        industria = get_str("industria")
        ingreso_ind = safe_float(cell(ws, row, cols["ingreso_individual"])) if "ingreso_individual" in cols else None
        ingreso_fam = safe_float(cell(ws, row, cols["ingreso_familiar"])) if "ingreso_familiar" in cols else None
        medio = get_str("medio")

        # Must have at least name or gender to be a real record
        if not nombre and not genero_raw:
            continue

        # Normalize gender → 'M', 'F', 'Otro' (schema CHECK constraint)
        gender = None
        if genero_raw:
            g = strip_accents(genero_raw).lower().strip()
            if g in ("m", "masculino", "maculino"):
                gender = "M"
            elif g in ("f", "femenino"):
                gender = "F"
            else:
                gender = "Otro"

        # Normalize purchase_type → 'uso_propio', 'inversion'
        purchase_type = None
        if tipo_compra_raw:
            tc = strip_accents(tipo_compra_raw).lower().strip()
            if "inversion" in tc:
                purchase_type = "inversion"
            elif "usuario" in tc or "vivienda" in tc or "uso" in tc:
                purchase_type = "uso_propio"

        # Normalize occupation_type → 'formal', 'informal', 'independiente', 'empresario'
        occupation_type = None
        if trabajo_raw:
            t = strip_accents(trabajo_raw).lower().strip()
            if "formal" in t and "informal" not in t:
                occupation_type = "formal"
            elif "informal" in t:
                occupation_type = "informal"
            elif "independiente" in t:
                occupation_type = "independiente"
            elif "empresario" in t:
                occupation_type = "empresario"
            # If it's a company name, treat as formal
            elif len(t) > 2:
                occupation_type = "formal"

        # Normalize children_count
        if cantidad_hijos is None and hijos_raw:
            h = hijos_raw.lower().strip()
            if h in ("no", "0", "false"):
                cantidad_hijos = 0
            elif h in ("si", "yes", "true", "1"):
                cantidad_hijos = None  # Has kids but count unknown

        # Parse age → birth_date is not available, just store None
        # We skip birth_date — the sheets have age, not DOB

        records.append({
            "project_key": project_key,
            "tower_name": tower_name,
            "unit_number": unit_number,
            "nombre": nombre,
            "gender": gender,
            "education_level": estudios,
            "purchase_type": purchase_type,
            "marital_status": estado_civil,
            "children_count": cantidad_hijos,
            "department": departamento,
            "zone": zona,
            "occupation_type": occupation_type,
            "industry": industria,
            "monthly_income_individual": ingreso_ind,
            "monthly_income_family": ingreso_fam,
            "acquisition_channel": medio,
        })

    wb.close()

    # Deduplicate by unit key — keep record with most non-NULL fields
    deduped: dict[str, dict] = {}
    for rec in records:
        key = f"{rec['project_key']}|{rec.get('tower_name', '')}|{rec['unit_number']}"
        data_fields = [
            "gender", "education_level", "purchase_type", "marital_status",
            "children_count", "department", "zone", "occupation_type",
            "industry", "monthly_income_individual", "monthly_income_family",
            "acquisition_channel",
        ]
        score = sum(1 for f in data_fields if rec.get(f) is not None)
        if key not in deduped or score > deduped[key][1]:
            deduped[key] = (rec, score)
    return [v[0] for v in deduped.values()]


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------

def generate_sql(
    referidos: list[dict],
    valorizacion: list[dict],
    buyer_personas: list[dict],
) -> str:
    lines: list[str] = []
    lines.append("-- ==========================================================================")
    lines.append("-- Production Backfill — Referidos, Valorizacion, Buyer Persona")
    lines.append("-- Generated by scripts/backfill_domains.py")
    lines.append("-- ==========================================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # Pre-flight
    lines.append("-- Pre-flight: abort if any target table has data")
    lines.append("DO $$ BEGIN")
    for tbl in ["rv_referrals", "rv_price_history", "rv_client_profiles"]:
        lines.append(f"  IF (SELECT count(*) FROM {tbl}) > 0 THEN")
        lines.append(f"    RAISE EXCEPTION '{tbl} is not empty — aborting';")
        lines.append(f"  END IF;")
    lines.append("END $$;")
    lines.append("")

    # ---- Domain 1: Referidos ----
    lines.append(f"-- Domain 1: Referidos ({len(referidos)} rows)")
    for r in referidos:
        pid = PROJECT_IDS[r["project_key"]]

        # Build unit subquery
        if r["tower_name"]:
            unit_subq = (
                f"(SELECT ru.id FROM rv_units ru "
                f"JOIN floors f ON f.id = ru.floor_id "
                f"JOIN towers t ON t.id = f.tower_id "
                f"WHERE t.project_id = '{pid}' "
                f"AND t.name = {sql_str(r['tower_name'])} "
                f"AND ru.unit_number = {sql_str(r['unit_number'])})"
            )
        else:
            # Multi-tower project without tower info — find by unit_number alone
            unit_subq = (
                f"(SELECT ru.id FROM rv_units ru "
                f"JOIN floors f ON f.id = ru.floor_id "
                f"JOIN towers t ON t.id = f.tower_id "
                f"WHERE t.project_id = '{pid}' "
                f"AND ru.unit_number = {sql_str(r['unit_number'])} LIMIT 1)"
            )

        sp_subq = "NULL"
        if r["salesperson"]:
            sp_subq = f"(SELECT id FROM salespeople WHERE full_name = {sql_str(r['salesperson'])} LIMIT 1)"

        lines.append(
            f"INSERT INTO rv_referrals "
            f"(unit_id, client_name, precio_lista, precio_referido, referido_por, "
            f"fecha_reserva, salesperson_id) VALUES ("
            f"{unit_subq}, {sql_str(r['client_name'])}, "
            f"{sql_num(r['precio_lista'])}, {sql_num(r['precio_referido'])}, "
            f"{sql_str(r['referido_por'])}, {sql_date(r['fecha'])}, {sp_subq});"
        )
    lines.append("")

    # ---- Domain 2: Valorizacion ----
    lines.append(f"-- Domain 2: Valorizacion ({len(valorizacion)} rows)")
    for v in valorizacion:
        pid = PROJECT_IDS[v["project_key"]]
        lines.append(
            f"INSERT INTO rv_price_history "
            f"(project_id, tower_id, effective_date, units_remaining, "
            f"increment_amount, appreciation_total, notes) VALUES ("
            f"'{pid}', NULL, {sql_date(v['effective_date'])}, "
            f"{v['units_remaining']}, {sql_num(v['increment_amount'])}, "
            f"{sql_num(v['appreciation_total'])}, {sql_str(v['notes'])}) "
            f"ON CONFLICT (project_id, tower_id, effective_date) DO NOTHING;"
        )
    lines.append("")

    # ---- Domain 3: Buyer Persona ----
    lines.append(f"-- Domain 3: Buyer Persona ({len(buyer_personas)} rows)")
    lines.append("-- Links via unit_number → reservation → client_id")

    for bp in buyer_personas:
        pid = PROJECT_IDS[bp["project_key"]]

        # Build unit subquery
        if bp["tower_name"]:
            unit_subq = (
                f"(SELECT ru.id FROM rv_units ru "
                f"JOIN floors f ON f.id = ru.floor_id "
                f"JOIN towers t ON t.id = f.tower_id "
                f"WHERE t.project_id = '{pid}' "
                f"AND t.name = {sql_str(bp['tower_name'])} "
                f"AND ru.unit_number = {sql_str(bp['unit_number'])})"
            )
        else:
            unit_subq = (
                f"(SELECT ru.id FROM rv_units ru "
                f"JOIN floors f ON f.id = ru.floor_id "
                f"JOIN towers t ON t.id = f.tower_id "
                f"WHERE t.project_id = '{pid}' "
                f"AND ru.unit_number = {sql_str(bp['unit_number'])} LIMIT 1)"
            )

        # Find client_id via: unit → reservation (latest CONFIRMED or any) → reservation_clients → client
        client_subq = (
            f"(SELECT rc.client_id FROM reservation_clients rc "
            f"JOIN reservations r ON r.id = rc.reservation_id "
            f"WHERE r.unit_id = {unit_subq} "
            f"AND rc.is_primary = true "
            f"ORDER BY r.created_at DESC LIMIT 1)"
        )

        lines.append(
            f"INSERT INTO rv_client_profiles "
            f"(client_id, gender, education_level, purchase_type, marital_status, "
            f"children_count, department, zone, occupation_type, industry, "
            f"monthly_income_individual, monthly_income_family, acquisition_channel) "
            f"SELECT {client_subq}, "
            f"{sql_str(bp['gender'])}, {sql_str(bp['education_level'])}, "
            f"{sql_str(bp['purchase_type'])}, {sql_str(bp['marital_status'])}, "
            f"{sql_num(bp['children_count'])}, {sql_str(bp['department'])}, "
            f"{sql_str(bp['zone'])}, {sql_str(bp['occupation_type'])}, "
            f"{sql_str(bp['industry'])}, {sql_num(bp['monthly_income_individual'])}, "
            f"{sql_num(bp['monthly_income_family'])}, {sql_str(bp['acquisition_channel'])} "
            f"WHERE {client_subq} IS NOT NULL "
            f"ON CONFLICT (client_id) DO NOTHING;"
        )
    lines.append("")

    lines.append("COMMIT;")
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=== Backfill Domains: Referidos, Valorizacion, Buyer Persona ===\n")

    # Verify files
    for key, path in REPORTE_FILES.items():
        if not path.exists():
            print(f"  ERROR: Not found: {path}", file=sys.stderr)
            sys.exit(1)
    for key, path in DISP_FILES.items():
        if not path.exists():
            print(f"  ERROR: Not found: {path}", file=sys.stderr)
            sys.exit(1)
    print("  All files found.\n")

    # Domain 1: Referidos
    print("--- Domain 1: Referidos ---")
    all_referidos: list[dict] = []

    # BLT has wider schema
    blt_refs = extract_referidos_blt(REPORTE_FILES["blt"])
    print(f"  BLT: {len(blt_refs)} records")
    all_referidos.extend(blt_refs)

    # B5, CE — standard schema (skip BEN — empty placeholders)
    for key in ["b5", "ce"]:
        refs = extract_referidos(key, REPORTE_FILES[key])
        print(f"  {PROJECT_NAMES[key]}: {len(refs)} records")
        all_referidos.extend(refs)

    print(f"  Total: {len(all_referidos)} referidos")

    # Domain 2: Valorizacion
    print("\n--- Domain 2: Valorizacion ---")
    valorizacion: list[dict] = []

    ben_val = extract_valorizacion_ben(DISP_FILES["ben"])
    print(f"  Benestare: {len(ben_val)} increments")
    for v in ben_val:
        print(f"    {v['effective_date']}: Q{v['increment_amount']:,.0f}/unit × {v['units_remaining']} units = Q{v['appreciation_total']:,.0f}")
    valorizacion.extend(ben_val)

    b5_val = extract_valorizacion_b5(DISP_FILES["b5"])
    print(f"  Boulevard 5: {len(b5_val)} adjustments")
    for v in b5_val:
        print(f"    {v['effective_date']}: {v['notes']} — {v['units_remaining']} units, total Q{v['appreciation_total']:,.0f}")
    valorizacion.extend(b5_val)

    ce_val = extract_valorizacion_ce(DISP_FILES["ce"])
    print(f"  Casa Elisa: {len(ce_val)} adjustments")
    for v in ce_val:
        print(f"    {v['effective_date']}: {v['notes']} — {v['units_remaining']} units, total Q{v['appreciation_total']:,.0f}")
    valorizacion.extend(ce_val)

    # Domain 3: Buyer Persona
    print("\n--- Domain 3: Buyer Persona ---")
    all_personas: list[dict] = []
    for key in ["blt", "ben", "b5", "ce"]:
        personas = extract_buyer_persona(key, REPORTE_FILES[key])
        print(f"  {PROJECT_NAMES[key]}: {len(personas)} profiles")
        all_personas.extend(personas)
    print(f"  Total: {len(all_personas)} profiles")

    # Generate SQL
    print("\n--- Generate SQL ---")
    sql = generate_sql(all_referidos, valorizacion, all_personas)

    out_path = ROOT / "scripts" / "backfill_domains.sql"
    out_path.write_text(sql, encoding="utf-8")
    print(f"  SQL written to: {out_path}")
    print(f"  Lines: {len(sql.splitlines())}")
    print(f"  Size: {len(sql):,} bytes")
    print(f"\n  REVIEW the SQL file before executing!")


if __name__ == "__main__":
    main()
